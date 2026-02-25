"""
================================================================================
VISUAL/CLUSTER_OPTUNA.PY — CLUSTERING DE RESULTADOS OPTUNA
================================================================================
Encuentra vecindarios de parámetros estables mediante HDBSCAN.
Entrada: Excel/CSV con doble encabezado (categorías + columnas).
Salida: Excel _AGRUPADO con columna Clúster_ID, ordenado por clúster.

http://127.0.0.1:8050


 
optuna-dashboard sqlite:////Users/manuel/Desktop/MODELOX/DATABASE/ID2.db
 
Ejecutar desde la raíz del proyecto:  python visual/cluster_optuna.py [archivo]
================================================================================
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

# Evitar que visual/rich.py sombree el paquete 'rich' al ejecutar desde visual/
_script_dir = Path(__file__).resolve().parent
if str(_script_dir) in sys.path:
    sys.path.remove(str(_script_dir))
if str(_script_dir.parent) not in sys.path:
    sys.path.insert(0, str(_script_dir.parent))

import re
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn
from rich.table import Table

# ─── Dependencias opcionales ─────────────────────────────────────────────────
try:
    import hdbscan
    _HAS_HDBSCAN = True
except ImportError:
    try:
        from sklearn.cluster import HDBSCAN
        _HAS_HDBSCAN = True
    except ImportError:
        _HAS_HDBSCAN = False

try:
    from sklearn.preprocessing import StandardScaler
    _HAS_SKLEARN = True
except ImportError:
    _HAS_SKLEARN = False

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

console = Console()

# Colores y constantes (alineados con excel.py)
COLORS = {
    "header_bg_cluster": "0F3460",
    "header_bg_id":      "0F3460",
    "header_bg_metrics": "1A1A2E",
    "header_bg_params":  "16213E",
    "text_dark":         "1A1A2E",
    "border_color":      "E8EDF5",
    "row_alt":           "F7F9FC",
}

# Paleta de fondos claros por clúster (cada clúster un color distinto)
CLUSTER_FILLS = [
    "E8F5E9",  # verde claro
    "E3F2FD",  # azul claro
    "FFF3E0",  # naranja claro
    "F3E5F5",  # violeta claro
    "E0F7FA",  # cyan claro
    "FCE4EC",  # rosa claro
    "FFF8E1",  # amarillo claro
    "E8EAF6",  # índigo claro
]
FONT_TITLE = "Arial"
FONT_BODY  = "Arial"
LABEL_RUIDO = "Sin Grupo"


# =============================================================================
# 1. INGRESO Y LIMPIEZA DE RUTA
# =============================================================================

def _clean_path(raw: str) -> str:
    """
    Limpia la ruta del archivo: elimina comillas y espacios al arrastrar.
    """
    s = str(raw).strip()
    for q in ('"', "'", "`"):
        if s.startswith(q) and s.endswith(q):
            s = s[1:-1].strip()
    return s


def _get_input_path() -> Optional[str]:
    """Obtiene ruta desde sys.argv o input()."""
    if len(sys.argv) >= 2:
        return _clean_path(sys.argv[1])
    try:
        raw = input("Arrastra tu archivo Excel/CSV aquí: ").strip()
        return _clean_path(raw) if raw else None
    except EOFError:
        return None


# =============================================================================
# 2. LECTURA Y DETECCIÓN DINÁMICA DE PARÁMETROS
# =============================================================================

def _read_file(path: str) -> Tuple[pd.DataFrame, List[str], List[str], List[str]]:
    """
    Lee Excel o CSV con doble encabezado.
    Fila 0: categorías (DATOS, MÉTRICAS, PARÁMETROS). Fila 1: nombres de columnas.
    Retorna: (df, categorias_row0, cols_params, todas_columnas)
    """
    ext = Path(path).suffix.lower()
    param_keywords = ("PARÁMETROS", "PARAMETROS", "PARAMETERS", "PARAMS")

    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(path, header=[0, 1])
    else:
        df = pd.read_csv(path, encoding="utf-8-sig", header=[0, 1], low_memory=False)

    if isinstance(df.columns, pd.MultiIndex):
        # Forward-fill categorías (celdas fusionadas dejan NaN en el resto)
        cats = []
        last_cat = ""
        for c in df.columns:
            cat = str(c[0]).strip().upper() if c[0] is not None and str(c[0]) != "nan" else last_cat
            if cat:
                last_cat = cat
            cats.append(last_cat if last_cat else "DATOS")

        cols = [str(c[1]).strip() if c[1] is not None and str(c[1]) != "nan" else str(c[0]) for c in df.columns]
        cols = [c if c else f"Col_{i}" for i, c in enumerate(cols)]
        df.columns = cols
    else:
        cats = [""] * len(df.columns)
        cols = list(df.columns)

    cols_params = [
        df.columns[i] for i, cat in enumerate(cats)
        if any(kw in cat for kw in param_keywords)
    ]

    return df, cats, cols_params, list(df.columns)


def _detect_param_columns(df: pd.DataFrame, header_row0: List[str]) -> List[str]:
    """
    Identifica columnas cuya categoría (fila 0) es PARÁMETROS.
    """
    param_keywords = ("PARÁMETROS", "PARAMETROS", "PARAMETERS", "PARAMS")
    out = []
    for i, cat in enumerate(header_row0):
        if i >= len(df.columns):
            break
        if any(kw in str(cat).upper() for kw in param_keywords):
            out.append(df.columns[i])
    return out


# =============================================================================
# 3. PREPROCESAMIENTO Y ESCALADO
# =============================================================================

def _prepare_params_matrix(
    df: pd.DataFrame,
    cols_params: List[str],
) -> Tuple[np.ndarray, np.ndarray, List[int]]:
    """
    Convierte columnas de parámetros a matriz numérica.
    Maneja NaN y no numéricos (reemplaza por mediana o 0).
    Retorna: (X_numérico, máscara_filas_válidas, índices_columnas_numeric)
    """
    valid_cols = []
    for c in cols_params:
        if c not in df.columns:
            continue
        s = pd.to_numeric(df[c], errors="coerce")
        if s.notna().any():
            valid_cols.append(c)

    if not valid_cols:
        return np.zeros((len(df), 1)), np.ones(len(df), dtype=bool), []

    X = df[valid_cols].copy()
    for c in valid_cols:
        X[c] = pd.to_numeric(X[c], errors="coerce")
    X = X.fillna(X.median(numeric_only=True)).fillna(0)
    X_arr = X.values.astype(np.float64)
    valid_mask = np.isfinite(X_arr).all(axis=1)
    return X_arr, valid_mask, [df.columns.get_loc(c) for c in valid_cols]


# =============================================================================
# 4. FILTRO POST-CLUSTERING (eliminar clústeres con ROI<0 o trades_dia<0.17)
# =============================================================================

APLICAR_FILTRO_METRICAS = True  # <--- NUEVA CONDICIÓN (True = filtra por métricas, False = solo agrupa por parámetros)
ROI_MIN = 0.05
TRADES_DIA_MIN = 0.05

def _find_metric_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """Detecta columnas ROI, TRADES_DIA y SCORE por nombre."""
    roi_col = None
    td_col = None
    score_col = None
    cols_upper = {c: str(c).upper() for c in df.columns}
    for c in df.columns:
        u = cols_upper[c]
        if not roi_col and ("ROI" in u or "RETURN_PCT" in u) and "PARAM" not in u:
            roi_col = c
        if not td_col and ("TRADES_DIA" in u or "TRADES_POR_DIA" in u or "TRADES_PER_DAY" in u):
            td_col = c
        if not score_col and u == "SCORE":
            score_col = c
    return roi_col, td_col, score_col


def _filter_clusters(
    df: pd.DataFrame,
    labels: np.ndarray,
    roi_col: Optional[str],
    td_col: Optional[str],
    score_col: Optional[str] = None,
) -> np.ndarray:
    """
    Máscara booleana: True = mantener trial.
    - Clústeres: eliminar todo el clúster si algún trial tiene ROI<0 o trades_dia<0.17.
    - Clústeres: eliminar si todos los trials del clúster tienen el mismo SCORE (sin varianza).
    - Sin Grupo: eliminar cada trial que falle individualmente.
    """
    keep = np.ones(len(df), dtype=bool)
    
    if not APLICAR_FILTRO_METRICAS:
        return keep

    clusters_to_drop = set()

    for cluster_id in np.unique(labels):
        if str(cluster_id) == LABEL_RUIDO:
            continue
            
        mask = labels == cluster_id
        subset = df.loc[mask]
        
        # 1. Filtro por ROI o Trades/Día negativos/bajos
        fails_metrics = False
        if roi_col and roi_col in df.columns:
            roi_vals = pd.to_numeric(subset[roi_col], errors="coerce").fillna(0)
            if (roi_vals < ROI_MIN).any():
                fails_metrics = True
        if td_col and td_col in df.columns:
            td_vals = pd.to_numeric(subset[td_col], errors="coerce").fillna(0)
            if (td_vals < TRADES_DIA_MIN).any():
                fails_metrics = True
                
        # 2. Restricción de varianza de SCORE: si todos tienen el mismo score, eliminar clúster
        fails_variance = False
        if score_col and score_col in df.columns and len(subset) > 1:
            score_vals = pd.to_numeric(subset[score_col], errors="coerce").fillna(-9999)
            if score_vals.nunique() == 1:
                fails_variance = True

        if fails_metrics or fails_variance:
            clusters_to_drop.add(cluster_id)

    # Procesar "Sin Grupo" individualmente
    mask_ruido = labels == LABEL_RUIDO
    if mask_ruido.any():
        for i in np.where(mask_ruido)[0]:
            r = float(pd.to_numeric(df.iloc[i][roi_col], errors="coerce") or 0) if roi_col and roi_col in df.columns else 0
            t = float(pd.to_numeric(df.iloc[i][td_col], errors="coerce") or 0) if td_col and td_col in df.columns else 0
            if (roi_col and roi_col in df.columns and r < ROI_MIN) or (td_col and td_col in df.columns and t < TRADES_DIA_MIN):
                keep[i] = False

    for cluster_id in clusters_to_drop:
        keep[labels == cluster_id] = False
    return keep


# =============================================================================
# 5. CLUSTERING HDBSCAN
# =============================================================================

def _run_hdbscan(X: np.ndarray, min_cluster_size: int = 2) -> np.ndarray:
    """Ejecuta HDBSCAN. Etiqueta -1 = ruido."""
    if not _HAS_HDBSCAN:
        raise ImportError("Instala hdbscan o sklearn>=1.0 para clustering.")
    try:
        clusterer = hdbscan.HDBSCAN(min_cluster_size=min_cluster_size)
    except NameError:
        from sklearn.cluster import HDBSCAN
        clusterer = HDBSCAN(min_cluster_size=min_cluster_size)
    return clusterer.fit_predict(X)


# =============================================================================
# 5. ESCRITURA EXCEL CON ESTILO (excel.py)
# =============================================================================

def _make_border(color: str = "E8EDF5", style: str = "thin") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_excel_style(
    filepath: str,
    df: pd.DataFrame,
    has_cluster_col: bool = True,
    cols_params: Optional[List[str]] = None,
) -> None:
    """Aplica formato visual: header (nombres col) + fondo distinto por clúster + fila Rango Total Absoluto."""
    if not _HAS_OPENPYXL:
        return
    wb = load_workbook(filepath)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    max_row = ws.max_row

    border = _make_border(COLORS["border_color"])
    fill_cluster = PatternFill("solid", fgColor=COLORS["header_bg_cluster"])
    fill_id = PatternFill("solid", fgColor=COLORS["header_bg_id"])
    fill_met = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
    fill_par = PatternFill("solid", fgColor=COLORS["header_bg_params"])
    cluster_to_fill = {}
    if "Clúster_ID" in df.columns:
        for i, cid in enumerate(df["Clúster_ID"].unique()):
            idx = i % len(CLUSTER_FILLS)
            cluster_to_fill[str(cid)] = PatternFill("solid", fgColor=CLUSTER_FILLS[idx])

    font_hdr = Font(name=FONT_TITLE, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_BODY, size=10, color=COLORS["text_dark"])
    align_c = Alignment(horizontal="center", vertical="center")
    align_cw = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fila 1: nombres de columnas | Fila 2+: datos
    col_hdrs = {c: str(ws.cell(1, c).value or "").upper() for c in range(1, max_col + 1)}
    id_like = {"TRIAL", "ESTRATEGIA", "SCORE"}
    metric_like = {"ROI", "WINRATE", "DRAWDOWN", "SHARPE", "SQN", "PROFIT", "TRADES", "LONG", "SHORT"}

    def _section_fill(col: int) -> PatternFill:
        h = col_hdrs.get(col, "")
        if has_cluster_col and col == 1:
            return fill_cluster
        if any(k in h for k in id_like):
            return fill_id
        if any(k in h for k in metric_like):
            return fill_met
        return fill_par

    ws.row_dimensions[1].height = 30
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = font_hdr
        cell.alignment = align_cw
        cell.border = border
        cell.fill = _section_fill(c)

    for r in range(2, max_row + 1):
        df_idx = r - 2
        if df_idx >= len(df) or "Clúster_ID" not in df.columns:
            rf = PatternFill("solid", fgColor="FFFFFF")
        else:
            cid = str(df.iloc[df_idx]["Clúster_ID"])
            rf = cluster_to_fill.get(cid, PatternFill("solid", fgColor="FFFFFF"))
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_body
            cell.alignment = align_c
            cell.border = border
            cell.fill = rf
            h = col_hdrs.get(c, "")
            if any(k in h for k in ("%", "PCT", "WINRATE", "ROI")):
                cell.number_format = "0.00%"
                try:
                    v = cell.value
                    if v is not None and isinstance(v, (int, float)) and v > 1:
                        cell.value = float(v) / 100.0
                except Exception:
                    pass
            elif any(k in h for k in ("SCORE", "SHARPE", "FACTOR", "SQN")):
                cell.number_format = "0.00"

    # Fila final: Rango Total Absoluto (min-max por parámetro)
    params_set = set(str(p).strip() for p in (cols_params or []))
    row_rango = max_row + 1
    fill_rango = PatternFill("solid", fgColor="E8EAF6")
    font_rango = Font(name=FONT_BODY, size=10, bold=True, color=COLORS["text_dark"])
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_rango, column=c)
        cell.font = font_rango
        cell.alignment = align_c
        cell.border = border
        cell.fill = fill_rango
        df_col = df.columns[c - 1] if c - 1 < len(df.columns) else None
        if c == 1:
            cell.value = "Rango Total Absoluto"
        elif df_col and df_col in params_set and df_col in df.columns:
            col_letter = get_column_letter(c)
            # Usar fórmula de Excel para que se ajuste automáticamente si el usuario borra filas
            formula = (
                f'=IF(COUNT({col_letter}2:{col_letter}{max_row})>0, '
                f'MIN({col_letter}2:{col_letter}{max_row}) & " - " & MAX({col_letter}2:{col_letter}{max_row}), '
                f'"-")'
            )
            cell.value = formula
        else:
            cell.value = ""

    # Anchos de columna (incluye fila Rango)
    for col in range(1, max_col + 1):
        max_len = len(str(ws.cell(1, col).value or ""))
        for r in range(2, min(24, row_rango + 1)):
            v = ws.cell(r, col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min((max_len + 2) * 1.15, 28)

    # Congelar paneles
    ws.freeze_panes = ws.cell(row=2, column=2)
    wb.save(filepath)


# =============================================================================
# MAIN
# =============================================================================

def main() -> None:
    if not _HAS_SKLEARN:
        console.print("[bold red]Error:[/bold red] scikit-learn es requerido.")
        sys.exit(1)
    if not _HAS_HDBSCAN:
        console.print("[bold red]Error:[/bold red] Instala [cyan]hdbscan[/cyan] o [cyan]sklearn>=1.0[/cyan].")
        sys.exit(1)

    path = _get_input_path()
    if not path or not os.path.isfile(path):
        console.print("[bold red]No se proporcionó un archivo válido.[/bold red]")
        sys.exit(1)

    path = os.path.abspath(path)
    base, ext = os.path.splitext(path)
    out_path = f"{base}_AGRUPADO{ext}"

    with Progress(
        SpinnerColumn(),
        TextColumn("[bold blue]{task.description}"),
        console=console,
    ) as progress:
        # ── Fase 1: Leyendo archivo ─────────────────────────────────────
        t1 = progress.add_task("Leyendo archivo...", total=None)
        try:
            df, cats, cols_params, all_cols = _read_file(path)
        except Exception as e:
            console.print(f"[bold red]Error leyendo archivo:[/bold red] {e}")
            sys.exit(1)
        progress.update(t1, completed=True)

        # Si _read_file no devolvió cols_params, detectar por categorías
        if not cols_params and cats:
            cols_params = _detect_param_columns(df, cats)

        # Fallback: todas las columnas numéricas que no sean ID/Métricas obvias
        if not cols_params:
            exclude = {"TRIAL", "ESTRATEGIA", "SCORE", "ROI", "WINRATE", "DRAWDOWN", "SHARPE", "SQN"}
            for c in df.columns:
                if str(c).upper() in exclude:
                    continue
                if pd.api.types.is_numeric_dtype(df[c]) or pd.to_numeric(df[c], errors="coerce").notna().any():
                    cols_params.append(c)

        # ── Fase 2: Detectando parámetros ───────────────────────────────
        t2 = progress.add_task("Detectando columnas de parámetros dinámicamente...", total=None)
        progress.update(t2, completed=True)

        if not cols_params:
            console.print("[bold yellow]No se detectaron columnas de parámetros. Usando todas las numéricas.[/bold yellow]")
            cols_params = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
        
        # Filtrar columnas de Trailing Stop si no se utiliza TRAILING
        is_trailing = False
        exit_col_name = next((c for c in df.columns if str(c).upper() in ["EXIT_TYPE", "PARAM_EXIT_TYPE"]), None)
        if exit_col_name:
            is_trailing = df[exit_col_name].astype(str).str.contains("TRAIL", case=False).any()
        else:
            for t_col in df.columns:
                tc_upper = str(t_col).upper()
                if tc_upper in ["ACT", "DIST", "TRAIL_ACT_PCT", "TRAIL_DIST_PCT"] or "TRAIL_ACT" in tc_upper or "TRAIL_DIST" in tc_upper:
                    try:
                        if (pd.to_numeric(df[t_col], errors='coerce').fillna(0) > 0).any():
                            is_trailing = True; break
                    except Exception:
                        pass
                        
        if not is_trailing:
            cols_to_drop_trail = [
                c for c in df.columns 
                if ("TRAIL_ACT" in str(c).upper() or "TRAIL_DIST" in str(c).upper() or str(c).upper() in ["ACT", "DIST"])
            ]
            if cols_to_drop_trail:
                df.drop(columns=cols_to_drop_trail, inplace=True, errors='ignore')
                cols_params = [c for c in cols_params if c not in cols_to_drop_trail]

        if not cols_params:
            console.print("[bold red]No hay columnas numéricas para clusterizar.[/bold red]")
            sys.exit(1)

        # ── Fase 3: Escalando y HDBSCAN ─────────────────────────────────
        t3 = progress.add_task("Escalando datos y aplicando HDBSCAN...", total=None)
        X_arr, valid_mask, _ = _prepare_params_matrix(df, cols_params)
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X_arr)
        labels_raw = _run_hdbscan(X_scaled, min_cluster_size=2)
        labels_str = np.array([LABEL_RUIDO if L == -1 else str(int(L)) for L in labels_raw])
        progress.update(t3, completed=True)

        # ── Filtrar clústeres: eliminar si ROI<0 o trades_dia<0.17 ───────
        t3b = progress.add_task("Filtrando clústeres (ROI≥0, trades/día≥0.17, varianza SCORE)...", total=None)
        roi_col, td_col, score_col = _find_metric_columns(df)
        keep_mask = _filter_clusters(df, labels_str, roi_col, td_col, score_col)
        df_out = df[keep_mask].copy()
        labels_str = labels_str[keep_mask]
        df_out.insert(0, "Clúster_ID", labels_str)
        n_sin_grupo_excl = int((df_out["Clúster_ID"] == LABEL_RUIDO).sum())
        # Excluir "Sin Grupo" del Excel (no se guardan)
        df_out = df_out[df_out["Clúster_ID"] != LABEL_RUIDO].reset_index(drop=True)
        labels_str = df_out["Clúster_ID"].values
        progress.update(t3b, completed=True)
        # Ordenar por número de clúster
        def _sort_key(ser):
            def _key(x):
                try:
                    return int(float(x))
                except (ValueError, TypeError):
                    return 999999
            return ser.map(_key)
        df_out = df_out.sort_values("Clúster_ID", key=_sort_key).reset_index(drop=True)

        # ── Fase 4: Guardando Excel ────────────────────────────────────
        t4 = progress.add_task("Guardando Excel...", total=None)
        df_out.to_excel(out_path, index=False, engine="openpyxl")
        _apply_excel_style(out_path, df_out, has_cluster_col=True, cols_params=cols_params)
        progress.update(t4, completed=True)

    # ── Resumen en consola ─────────────────────────────────────────────
    n_clusters = len(set(labels_str))
    n_eliminados_filtro = len(df) - len(df_out) - n_sin_grupo_excl

    tbl = Table(show_header=True, header_style="bold cyan", border_style="blue")
    tbl.add_column("Métrica", style="cyan", width=28)
    tbl.add_column("Valor", justify="right", style="green")
    tbl.add_row("Archivo de entrada", path)
    tbl.add_row("Archivo de salida", out_path)
    tbl.add_row("Total trials (entrada)", str(len(df)))
    tbl.add_row("Eliminados por filtro (ROI/trades_dia)", str(n_eliminados_filtro))
    tbl.add_row("Sin grupo (excluidos del Excel)", str(n_sin_grupo_excl))
    tbl.add_row("Trials guardados en Excel", str(len(df_out)))
    tbl.add_row("Columnas de parámetros", str(len(cols_params)))
    tbl.add_row("Clústeres en Excel", str(n_clusters))
    if not roi_col or not td_col:
        missing = []
        if not roi_col:
            missing.append("ROI")
        if not td_col:
            missing.append("trades_dia")
        tbl.add_row("[dim]Columnas filtro no encontradas[/dim]", f"[dim]{', '.join(missing)}[/dim]")

    console.print()
    console.print(Panel(tbl, title="[bold white]✅ Clustering completado[/bold white]", border_style="green"))
    console.print()


if __name__ == "__main__":
    main()