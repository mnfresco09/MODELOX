"""
================================================================================
VISUAL/EXCEL.PY — DASHBOARD QUANT EN EXCEL  (v7.1 — ANCHO 10 XL)
================================================================================

CAMBIOS v7.1:
  - Gráfico: anclado en col B, ancho 1000px (ancho 10)
  - Tabla:   columna R (índice 17, 0-based), fuente 13pt, fila 2x
  - Sin solape: chart 1000px, tabla desplazada a la derecha
================================================================================
"""

import os
import re
import math
import datetime
import logging
from copy import deepcopy
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Protocol
import csv

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

try:
    import xlsxwriter
    from xlsxwriter.utility import xl_col_to_name
    _HAS_XW = True
except ImportError:
    _HAS_XW = False

logger = logging.getLogger(__name__)

# ==============================================================================
# COLORES Y CONSTANTES
# ==============================================================================

COLORS = {
    "header_bg_metrics": "2D3436",
    "header_bg_params":  "636E72",
    "header_bg_id":      "2D3436",
    "text_white":        "FFFFFF",
    "text_dark":         "2D3436",
    "border_color":      "DFE6E9",
    "success_bg":        "F0FFF4",
    "danger_bg":         "FFF5F5",
    "accent_green":      "38A169",
    "accent_red":        "E53E3E",
    "row_alt":           "F8F9FA",
    "section_border":    "A0AEC0",
    "table_header_bg":   "EDF2F7",
}

FONT_TITLE = "Calibri"
FONT_BODY  = "Calibri"

METRICS_ORDER = [
    "TRADES_DIA", "LONG", "SHORT", "PROFIT_FACTOR",
    "ROI_PCT", "MAX_DD_PCT", "EXPECTATIVA",
    "SALDO_SIN_COMISIONES", "PNL_NETO", "SALDO_ACTUAL", "COMISIONES_TOTAL",
]

# Alias y métricas extra para detectar correctamente columnas de métricas
# que llegan con variantes de nombre desde metrics.py/reporting.
METRIC_ALIASES_TO_CANONICAL = {
    "ROI": "ROI_PCT",
    "ROI%": "ROI_PCT",
    "RETURN": "ROI_PCT",
    "RETURN%": "ROI_PCT",
    "RETURN_PCT": "ROI_PCT",
    "WINRATE": "WINRATE_PCT",
    "WIN_RATE": "WINRATE_PCT",
    "WIN_RATE_PCT": "WINRATE_PCT",
    "MAX_DRAWDOWN": "MAX_DD_PCT",
    "MAX_DRAWDOWN_PCT": "MAX_DD_PCT",
    "DD": "MAX_DD_PCT",
    "DD_PCT": "MAX_DD_PCT",
    "DRAWDOWN": "MAX_DD_PCT",
    "N_TRADES": "TOTAL_TRADES",
    "NUM_TRADES": "TOTAL_TRADES",
    "COUNT_TRADES": "TOTAL_TRADES",
    "TRADES_POR_DIA": "TRADES_DIA",
}

# Nombres para mostrar en las cabeceras del Excel (interno → display)
METRIC_DISPLAY_NAMES = {
    "TRADES_DIA":           "TRADES DIA",
    "LONG":                 "LONG",
    "SHORT":                "SHORT",
    "PROFIT_FACTOR":        "PROFIT FACTOR",
    "ROI_PCT":              "ROI",
    "MAX_DD_PCT":           "MAX DD",
    "EXPECTATIVA":          "EXPECTATIVA",
    "SALDO_SIN_COMISIONES": "BEN BRUTO",
    "PNL_NETO":             "BEN NETO",
    "SALDO_ACTUAL":         "SALDO ACTUAL",
    "COMISIONES_TOTAL":     "COMISIONES",
}

ALL_METRICS_ORDER = METRICS_ORDER

# Todas las métricas conocidas (para excluirlas de la sección PARÁMETROS)
ALL_KNOWN_METRICS = {
    "TOTAL_TRADES", "TRADES_DIA", "LONG", "SHORT",
    "PROFIT_FACTOR", "ROI_PCT", "WINRATE_PCT", "MAX_DD_PCT",
    "SHARPE", "SQN", "EXPECTATIVA", "SORTINO", "CALMAR",
    "PAYOFF_RATIO", "NET_PNL", "PNL_NETO", "SALDO_ACTUAL",
    "SALDO_MAX", "SALDO_MIN", "COMISIONES_TOTAL", "SALDO_SIN_COMISIONES",
    "ROI%", "WINRATE%", "MAX_DD%", "WIN_STREAK", "LOSS_STREAK",
    "AVG_TRADE", "TRADES_POR_DIA", "PNL_NETO_POR_DIA_OPERADO",
    "DURATION_MEAN_MIN", "RACHA_GANADORA", "RACHA_PERDEDORA",
    "PORC_GANADORAS", "PORC_PERDEDORAS", "SALDO_MEAN",
    "MAX_GANANCIA", "MAX_PERDIDA",
}

ID_COLS = ["TRIAL", "ESTRATEGIA", "SCORE"]

EXCLUDED_PARAMS = {
    "NOMBRE_COMBO", "EXIT_TYPE", "CANTIDAD",
    "ACTIVO", "TIMEFRAME", "TF", "ASSET", "SYMBOL",
    "RESULTADO", "METRICS", "COMBO", "ESTATEGIA",
    "SALDO", "VOLUMEN", "APALANCAMIENTO",
}

METRIC_KEYWORDS_TO_DROP = [
    "PROFIT", "LOSS", "PNL", "NET", "GROSS", "SALDO", "BALANCE", "RETORNO", "RETURN",
    "ROI", "BENEFICIO", "RIESGO", "RISK", "REWARD", "COMISION", "FEES",
    "WIN", "GANADORA", "PERDEDORA", "ACIERTO", "RATE", "PCT", "PORC_", "PERCENT",
    "DRAWDOWN", "DD", "RACHA", "STREAK", "UNDERWATER",
    "RATIO", "FACTOR", "SHARPE", "SORTINO", "CALMAR", "SQN", "EXPECTATIVA", "KELLY",
    "AVG", "MEAN", "MEDIAN", "STD", "VAR", "MAX", "MIN", "SUM", "TOTAL",
    "ESTABILIDAD", "COUNT", "NUM_", "N_", "TRADES", "LONGS", "SHORTS", "CANTIDAD_OP",
    "METRIC", "RESULT", "BEST", "WORST", "DIA_OPERADO", "DURATION", "TIME"
]

PREFIXES_TO_CLEAN = [
    "ESTRATEGIA_PARAMS_", "STRATEGY_PARAMS_", "PARAM_", "PARAMS_",
    "INDICATOR_", "CONFIG_", "METRICS_"
]

# ==============================================================================
# EXCEL REPORTER
# ==============================================================================

class ReporterProtocol(Protocol):
    def needs_dataframe(self, score: float) -> bool: ...
    def on_trial_end(self, artifacts: Any) -> None: ...
    def on_strategy_end(self, strategy_name: str, study: Any) -> None: ...


@dataclass
class ExcelReporter:
    resumen_path: str = "resultados/excel/resumen.xlsx"
    trades_base_dir: str = "resultados/excel"
    max_archivos: int = 5
    use_fast_mode: bool = True
    # Datos de precio para la gráfica comparativa
    datos_dir: str = "datos"
    fecha_inicio: Optional[str] = None
    fecha_fin: Optional[str] = None
    formato_datos: str = "feather"

    _csv_resumen_path: Optional[str] = field(default=None, init=False, repr=False)
    _resumen_rows: List[Dict[str, Any]] = field(default_factory=list, init=False, repr=False)
    _trade_candidates: List[Dict[str, Any]] = field(default_factory=list, init=False, repr=False)
    _min_candidate_score: float = field(default=float("-inf"), init=False, repr=False)
    _activo: Optional[str] = field(default=None, init=False, repr=False)
    _final_excel_path: Optional[str] = field(default=None, init=False, repr=False)

    def needs_dataframe(self, score: float) -> bool:
        return False

    @staticmethod
    def _safe_activo_name(activo: str) -> str:
        return str(activo).strip().replace(" ", "_").upper() if activo else "DEFAULT"

    def _update_min_score(self):
        self._min_candidate_score = (
            min(c["score"] for c in self._trade_candidates)
            if self._trade_candidates else float("-inf")
        )

    def on_trial_end(self, artifacts) -> None:
        params_src = getattr(artifacts, "params_reporting", None) or artifacts.params
        activo = None
        if isinstance(params_src, dict):
            activo = params_src.get("__activo") or params_src.get("ACTIVO") or params_src.get("activo")

        self._activo = activo
        score  = artifacts.score if artifacts.score is not None else 0.0
        params = dict(params_src)
        params["NOMBRE_COMBO"] = artifacts.strategy_name

        self._resumen_rows.append({
            "trial_number":  artifacts.trial_number,
            "score":         score,
            "metrics":       deepcopy(artifacts.metrics) if artifacts.metrics else {},
            "params":        {k: v for k, v in params.items() if not str(k).startswith("__")},
            "strategy_name": artifacts.strategy_name,
        })

        try:
            base_dir = self.trades_base_dir
            os.makedirs(base_dir, exist_ok=True)
            if not self._csv_resumen_path:
                self._csv_resumen_path = os.path.join(base_dir, "RESUMEN.csv")
            self._write_resumen_csv(self._csv_resumen_path)
        except Exception:
            pass

        is_candidate = (
            len(self._trade_candidates) < self.max_archivos or
            score > self._min_candidate_score
        )
        if is_candidate and artifacts.trades is not None:
            self._trade_candidates.append({
                "score":        score,
                "trial_number": artifacts.trial_number,
                "trades":       artifacts.trades,
                "params":       params,
                "metrics":      artifacts.metrics,
            })
            if len(self._trade_candidates) > self.max_archivos:
                self._trade_candidates.sort(key=lambda x: x["score"], reverse=True)
                self._trade_candidates.pop()
            self._update_min_score()

    def on_strategy_end(self, strategy_name: str, study) -> None:
        if not self._resumen_rows:
            return

        activo   = self._activo
        base_dir = self.trades_base_dir
        os.makedirs(base_dir, exist_ok=True)

        activo_safe = self._safe_activo_name(str(activo) if activo else "DEFAULT")
        csv_path    = os.path.join(base_dir, "RESUMEN.csv")
        self._write_resumen_csv(csv_path)

        self._trade_candidates.sort(key=lambda x: x["score"], reverse=True)
        for candidate in self._trade_candidates[:self.max_archivos]:
            try:
                self._write_trades_excel(base_dir, candidate)
            except Exception as e:
                logger.warning(f"Error guardando trades trial {candidate['trial_number']}: {e}")

        # Intentar leer fechas desde configuracion.py si no se pasaron
        _fecha_inicio = self.fecha_inicio
        _fecha_fin    = self.fecha_fin
        if _fecha_inicio is None:
            try:
                from general.configuracion import FECHA_INICIO, FECHA_FIN
                _fecha_inicio = FECHA_INICIO
                _fecha_fin    = FECHA_FIN
            except Exception:
                pass

        try:
            self._final_excel_path = convertir_resumen_csv_a_excel(
                csv_path=csv_path,
                strategy_name=strategy_name,
                activo=activo_safe,
                output_dir=base_dir,
                excel_path=self.resumen_path,
                datos_dir=self.datos_dir,
                fecha_inicio=_fecha_inicio,
                fecha_fin=_fecha_fin,
                formato_datos=self.formato_datos,
            )
        except Exception as e:
            logger.warning(f"Error generando Dashboard Excel: {e}")

        self._resumen_rows        = []
        self._trade_candidates    = []
        self._min_candidate_score = float("-inf")

    def _write_resumen_csv(self, csv_path: str):
        if not self._resumen_rows:
            return
        all_keys = {"trial", "score", "strategy"}
        for row in self._resumen_rows:
            if row.get("metrics"):
                all_keys.update(row["metrics"].keys())
            if row.get("params"):
                all_keys.update(f"param_{k}" for k in row["params"].keys())

        columns = ["trial", "score", "strategy"]
        columns.extend(sorted(k for k in all_keys if k not in {"trial","score","strategy"} and not k.startswith("param_")))
        columns.extend(sorted(k for k in all_keys if k.startswith("param_")))

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            for row in self._resumen_rows:
                csv_row = {"trial": row["trial_number"], "score": row["score"], "strategy": row["strategy_name"]}
                if row.get("metrics"):
                    csv_row.update(row["metrics"])
                if row.get("params"):
                    csv_row.update({f"param_{k}": v for k, v in row["params"].items()})
                writer.writerow(csv_row)

    def _write_trades_excel(self, trades_dir: str, candidate: Dict[str, Any]):
        trades = candidate["trades"]
        if trades is None or (hasattr(trades, "empty") and trades.empty):
            return

        df_trades = trades.to_pandas() if hasattr(trades, "to_pandas") else trades

        # --- Limpiar timezone ---
        try:
            df_trades = df_trades.copy()
            if isinstance(df_trades.index, pd.DatetimeIndex) and df_trades.index.tz is not None:
                df_trades.index = df_trades.index.tz_localize(None)
            for col in list(df_trades.columns):
                try:
                    if isinstance(df_trades[col].dtype, pd.DatetimeTZDtype):
                        df_trades[col] = df_trades[col].dt.tz_localize(None)
                except Exception:
                    continue
            for col in list(df_trades.columns):
                if df_trades[col].dtype != object:
                    continue
                s = df_trades[col]
                try:
                    sample = next((v for v in s.head(50).tolist() if v is not None), None)
                    if sample and getattr(sample, "tzinfo", None):
                        df_trades[col] = s.apply(
                        )
                except Exception:
                    continue
        except Exception:
            pass

        exit_type = str(candidate["params"].get("__exit_type", candidate["params"].get("exit_type", "FIXED"))).upper()
        df_export = _preparar_df_trades(df_trades, exit_type)

        saldo = candidate['params'].get('__saldo_usado') or 0

        apal  = candidate['params'].get('__apalancamiento_max') or 0
        vol   = saldo * apal

        filename = f"TRIAL {candidate['trial_number']} - {int(candidate['score'])}.xlsx"
        filepath = os.path.join(trades_dir, filename)

        try:
            _escribir_trades_xlsxwriter(filepath, df_export, saldo, vol, apal)
        except Exception as e:
            logger.warning(f"Error xlsxwriter, fallback openpyxl: {e}")
            _escribir_trades_openpyxl_fallback(filepath, df_export, saldo, vol, apal)


# ==============================================================================
# PREPARACIÓN DEL DATAFRAME
# ==============================================================================

def _preparar_df_trades(df_trades: pd.DataFrame, exit_type: str = "") -> pd.DataFrame:
    rename_map = {

        "entry_time": "ENTRY_TIME", "exit_time": "EXIT_TIME",
        "type": "POSICIÓN", "entry_price": "ENTRY_PRICE", "exit_price": "EXIT_PRICE",
        "qty": "CANTIDAD", "saldo_usado": "SALDO",
        "pnl_bruto": "PNL BRUTO", "comision": "COMISIONES",
        "pnl_neto": "PNL NETO", "pnl_pct": "ROI",
        "saldo_antes": "BALANCE_PRE", "saldo_despues": "BALANCE",
        "reason": "EXIT_REASON", "trail_act_price": "TRAIL_ACT_PRICE",
        "trail_act_time": "TRAIL_ACT_TIME"
    }
    df = df_trades.rename(columns=rename_map)

    cols_to_drop = [
        "ENTRY_IDX", "EXIT_IDX", "SIDE_INT", "entry_idx", "exit_idx", "side_int",
        "BALANCE_PRE", "TRAIL_ACT_IDX", "trail_act_idx",
        "ROI", "SALDO", "VOLUMEN", "APALANCAMIENTO"
    ]
    df.drop(columns=[c for c in cols_to_drop if c in df.columns], inplace=True)

    reason_map = {1: "SL", 2: "TP", 3: "TRAIL", 4: "TIME", 0: "END", 5: "CUSTOM"}
    if "EXIT_REASON" in df.columns:
        df["EXIT_REASON"] = df["EXIT_REASON"].map(reason_map).fillna("UNKNOWN")

    for c in ["POSICIÓN", "EXIT_REASON"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.upper()

    df.columns = [c.upper() for c in df.columns]

    order = [
        "ENTRY_TIME", "ENTRY_PRICE", "TRAIL_ACT_PRICE", "TRAIL_ACT_TIME",
        "EXIT_TIME", "EXIT_PRICE", "POSICIÓN", "EXIT_REASON", "CANTIDAD",
        "PNL BRUTO", "COMISIONES", "PNL NETO", "BALANCE"
    ]
    
    # Si la salida es FIXED (no TRAILING), no mostramos columnas de activación de Trail
    if "TRAIL" not in exit_type:
        order = [c for c in order if c not in ("TRAIL_ACT_PRICE", "TRAIL_ACT_TIME")]

    # Truncar precios a 2 decimales según petición del usuario

    for price_col in ["ENTRY_PRICE", "EXIT_PRICE", "TRAIL_ACT_PRICE"]:
        if price_col in df.columns:
            df[price_col] = df[price_col].apply(lambda x: int(x * 100) / 100.0 if pd.notnull(x) else x)

    return df[[c for c in order if c in df.columns]]


# ==============================================================================
# HELPER: major_unit "bonito" para el eje Y
# ==============================================================================

def _nice_major_unit(data_range: float, target_ticks: int = 6) -> float:
    """Calcula un major_unit legible para el eje Y."""
    if data_range <= 0:
        return 1.0
    raw = data_range / target_ticks
    magnitude = math.pow(10, math.floor(math.log10(raw)))
    normalized = raw / magnitude
    if normalized <= 1:
        nice = 1
    elif normalized <= 2:
        nice = 2
    elif normalized <= 5:
        nice = 5
    else:
        nice = 10
    return nice * magnitude


# ==============================================================================
# ESCRITURA ULTRA RÁPIDA CON XLSXWRITER  (v7 — posición y escalado corregidos)
# ==============================================================================

def _escribir_trades_xlsxwriter(
    filepath: str,
    df: pd.DataFrame,
    val_saldo: float = 0,
    val_volumen: float = 0,
    val_apal: float = 0,
):
    """
    Layout final:
      · Datos:  A1 → L(max_row+1)
      · Gráfico: anclado en B(max_row+3), ancho 500 px × alto 430 px
                 → ocupa aproximadamente B → I (8 cols × ~65 px = 520 px)
      · Tabla:  anclada en J(max_row+3), inmediatamente junto al gráfico
                fuente 13 pt, filas 28 px (tamaño 2x)
    """
    n_rows = len(df)
    cols   = list(df.columns)

    wb = xlsxwriter.Workbook(filepath, {
        'nan_inf_to_errors': True,
        'strings_to_numbers': False,
        # Necesario para que xlsxwriter serialice datetime correctamente
        'default_date_format': 'dd/mm/yy hh:mm',
    })
    ws = wb.add_worksheet('Trades')
    ws.hide_gridlines(2)
    ws.freeze_panes(1, 0)
    ws.set_row(0, 28)

    # ── BASE de formato ──────────────────────────────────────────────────────
    _BASE = dict(
        font_name='Calibri', font_size=10, font_color='#2D3436',
        align='center', valign='vcenter',
        bottom=1, bottom_color='#E2E8F0',
        top=0, left=0, right=0,
    )

    def _fmt(bg='#FFFFFF', num_format=None, **kw):
        p = {**_BASE, 'bg_color': bg}
        if num_format:
            p['num_format'] = num_format
        p.update(kw)
        return wb.add_format(p)

    hdr_fmt = wb.add_format({
        'bold': True, 'font_name': 'Calibri', 'font_size': 10,
        'font_color': '#718096', 'bg_color': '#F7FAFC',
        'align': 'center', 'valign': 'vcenter',
        'bottom': 2, 'bottom_color': '#A0AEC0',
        'top': 0, 'left': 0, 'right': 0,
        'text_wrap': True,
    })

    # Pares de formato (fila normal, fila alternada)
    FMT = {
        'gen':   (_fmt('#FFFFFF'),                          _fmt('#F7FAFC')),
        'dt':    (_fmt('#FFFFFF', 'dd/mm/yy hh:mm'),        _fmt('#F7FAFC', 'dd/mm/yy hh:mm')),
        'price': (_fmt('#FFFFFF', '#,##0.00'),               _fmt('#F7FAFC', '#,##0.00')),
        'money': (_fmt('#FFFFFF', '#,##0.00'),               _fmt('#F7FAFC', '#,##0.00')),
        'pct':   (_fmt('#FFFFFF', '0.00%'),                  _fmt('#F7FAFC', '0.00%')),
        'apal':  (_fmt('#FFFFFF', '0.00"x"'),                _fmt('#F7FAFC', '0.00"x"')),
    }

    def _fmt_key(hdr: str) -> str:
        h = hdr.upper()
        if 'TIME' in h or 'DATE' in h:                          return 'dt'
        if 'PRICE' in h:                                         return 'price'
        if 'PNL' in h or 'BALANCE' in h or 'COMISIONES' in h:  return 'money'
        if h == 'ROI':                                           return 'pct'
        if 'APALANCAMIENTO' in h:                                return 'apal'
        return 'gen'

    col_keys = [_fmt_key(c) for c in cols]

    # Índices de columnas especiales
    pnl_neto_idx    = next((i for i, c in enumerate(cols) if 'PNL' in c and 'NETO' in c), None)
    balance_idx     = next((i for i, c in enumerate(cols) if c == 'BALANCE'), None)
    entry_time_idx  = next((i for i, c in enumerate(cols) if c == 'ENTRY_TIME'), None)
    comisiones_idx  = next((i for i, c in enumerate(cols) if 'COMISIONES' in c), None)
    entry_price_idx = next((i for i, c in enumerate(cols) if c == 'ENTRY_PRICE'), None)
    exit_price_idx  = next((i for i, c in enumerate(cols) if c == 'EXIT_PRICE'), None)

    # ── Hoja auxiliar oculta para datos de gráficos ─────────────────────
    # Columnas: A=ENTRY_TIME, B=BAL_BRUTO, C=STRAT_ROI%, D=BH_ROI%
    has_bruto_data = False
    has_roi_data   = False
    if n_rows > 0:
        ws_aux = wb.add_worksheet('_ChartData')
        ws_aux.hide()
        ws_aux.write_string(0, 0, 'ENTRY_TIME')
        ws_aux.write_string(0, 1, 'BAL_BRUTO')
        ws_aux.write_string(0, 2, 'STRAT_IDX')
        ws_aux.write_string(0, 3, 'BH_IDX')
        ws_aux.write_string(0, 4, 'CUM_COMISIONES')
        ws_aux.write_string(0, 5, 'CUM_PNL_NETO')

        dt_fmt = wb.add_format({'num_format': 'dd/mm/yy hh:mm'})

        # Primer precio del activo (para Buy & Hold)
        first_entry_price = None
        _saldo_ini = 0.0  # Se calcula en la primera iteración
        if entry_price_idx is not None:
            try:
                first_entry_price = float(df.iloc[0, entry_price_idx])
            except (ValueError, TypeError):
                first_entry_price = None

        comisiones_acum = 0.0
        pnl_neto_acum   = 0.0
        has_fee_data    = False
        for r_idx in range(n_rows):
            # Col A: Timestamp
            if entry_time_idx is not None:
                try:
                    et_val = df.iloc[r_idx, entry_time_idx]
                    if isinstance(et_val, (pd.Timestamp, datetime.datetime)):
                        dt = et_val.to_pydatetime() if hasattr(et_val, 'to_pydatetime') else et_val
                        dt = dt.replace(tzinfo=None)
                        ws_aux.write_datetime(r_idx + 1, 0, dt, dt_fmt)
                    else:
                        ws_aux.write(r_idx + 1, 0, str(et_val))
                except Exception:
                    ws_aux.write(r_idx + 1, 0, r_idx)

            # Col B: Balance Bruto
            if balance_idx is not None and comisiones_idx is not None:
                try:
                    comision_val = float(df.iloc[r_idx, comisiones_idx])
                except (ValueError, TypeError):
                    comision_val = 0.0
                comisiones_acum += comision_val
                try:
                    balance_val = float(df.iloc[r_idx, balance_idx])
                except (ValueError, TypeError):
                    balance_val = 0.0
                bruto_val = balance_val + comisiones_acum
                ws_aux.write_number(r_idx + 1, 1, bruto_val)
                has_bruto_data = True

            # Col C: Strategy ROI % (0% = punto de partida)
            # saldo_inicial = BALANCE[0] - PNL_NETO[0] (balance antes del primer trade)
            if balance_idx is not None and pnl_neto_idx is not None:
                try:
                    bal = float(df.iloc[r_idx, balance_idx])
                except (ValueError, TypeError):
                    bal = 0.0
                if r_idx == 0:
                    try:
                        first_pnl = float(df.iloc[0, pnl_neto_idx])
                        first_bal = float(df.iloc[0, balance_idx])
                        _saldo_ini = first_bal - first_pnl
                    except (ValueError, TypeError):
                        _saldo_ini = first_bal if first_bal > 0 else 1.0
                if _saldo_ini > 0:
                    strat_roi = (bal / _saldo_ini - 1.0) * 100.0
                    ws_aux.write_number(r_idx + 1, 2, strat_roi)
                    has_roi_data = True

            # Col D: Buy & Hold ROI %
            if entry_price_idx is not None and first_entry_price and first_entry_price > 0:
                # Usar exit_price del último trade, entry_price para el resto
                if r_idx == n_rows - 1 and exit_price_idx is not None:
                    try:
                        price_now = float(df.iloc[r_idx, exit_price_idx])
                    except (ValueError, TypeError):
                        price_now = first_entry_price
                else:
                    try:
                        price_now = float(df.iloc[r_idx, entry_price_idx])
                    except (ValueError, TypeError):
                        price_now = first_entry_price
                # ROI % (0% = inicio)
                bh_roi = (price_now / first_entry_price - 1.0) * 100.0
                ws_aux.write_number(r_idx + 1, 3, bh_roi)

            # Col E: Comisiones acumuladas
            if comisiones_idx is not None:
                try:
                    com_val = float(df.iloc[r_idx, comisiones_idx])
                except (ValueError, TypeError):
                    com_val = 0.0
                # Solo sumar si no se sumó ya arriba (evitar doble suma)
                # comisiones_acum ya fue actualizada en Col B
                ws_aux.write_number(r_idx + 1, 4, comisiones_acum)
                has_fee_data = True

            # Col F: PNL Neto acumulado
            if pnl_neto_idx is not None:
                try:
                    pnl_val = float(df.iloc[r_idx, pnl_neto_idx])
                except (ValueError, TypeError):
                    pnl_val = 0.0
                pnl_neto_acum += pnl_val
                ws_aux.write_number(r_idx + 1, 5, pnl_neto_acum)

    # ── Anchos de columna: muestrea 20 filas, O(cols) ───────────────────────
    for i, col_name in enumerate(cols):
        max_len = len(str(col_name))
        if n_rows > 0:
            sample_len = df.iloc[:20, i].astype(str).str.len().max()
            if pd.notna(sample_len):
                max_len = max(max_len, int(sample_len))
        ws.set_column(i, i, min((max_len + 2) * 1.15, 28))

    # ── HEADERS ─────────────────────────────────────────────────────────────
    for i, col_name in enumerate(cols):
        ws.write(0, i, col_name, hdr_fmt)

    # ── DATOS: O(n) en C nativo, ~20-50x más rápido que openpyxl ───────────
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=1):
        alt = (r_idx % 2 == 0)
        ws.set_row(r_idx, 18)

        for c_idx, val in enumerate(row_data):
            fmt = FMT[col_keys[c_idx]][1 if alt else 0]

            if val is None or (isinstance(val, float) and pd.isna(val)):
                ws.write_blank(r_idx, c_idx, None, fmt)
            elif isinstance(val, (pd.Timestamp, datetime.datetime)):
                try:
                    dt = val.to_pydatetime() if hasattr(val, 'to_pydatetime') else val
                    dt = dt.replace(tzinfo=None)
                    ws.write_datetime(r_idx, c_idx, dt, fmt)
                except Exception:
                    ws.write_string(r_idx, c_idx, str(val), fmt)
            elif isinstance(val, bool):
                ws.write_boolean(r_idx, c_idx, val, fmt)
            elif isinstance(val, (int, float)):
                ws.write_number(r_idx, c_idx, float(val), fmt)
            else:
                ws.write_string(r_idx, c_idx, str(val), fmt)

    # ── FORMATO CONDICIONAL PNL NETO ────────────────────────────────────────
    if pnl_neto_idx is not None and n_rows > 0:
        fmt_green = wb.add_format({**_BASE,
                                    'font_color': '#38A169', 'bold': True})
        fmt_red   = wb.add_format({**_BASE,
                                    'font_color': '#E53E3E', 'bold': True})
        ws.conditional_format(1, pnl_neto_idx, n_rows, pnl_neto_idx,
                               {'type': 'cell', 'criteria': '>', 'value': 0, 'format': fmt_green})
        ws.conditional_format(1, pnl_neto_idx, n_rows, pnl_neto_idx,
                               {'type': 'cell', 'criteria': '<', 'value': 0, 'format': fmt_red})

    # ── FILA DE SEPARACIÓN entre datos y gráfico/tabla ──────────────────────
    BLOCK_ROW = n_rows + 2   # fila 0-indexed donde empieza bloque inferior

    CHARTS_HEIGHT_ROWS = 0  # filas que ocupan los gráficos (para calcular TABLE_ROW)

    # ── GRÁFICO 1: EVOLUCIÓN DEL BALANCE (solo Balance Neto) ────────────────
    if balance_idx is not None and n_rows > 0:
        bal_series = df.iloc[:, balance_idx].dropna()
        y_min_raw  = float(bal_series.min())
        y_max_raw  = float(bal_series.max())
        data_range = y_max_raw - y_min_raw if y_max_raw != y_min_raw else max(abs(y_max_raw) * 0.1, 1.0)

        margin     = data_range * 0.03
        y_min_axis = y_min_raw - margin
        y_max_axis = y_max_raw + margin
        major_unit = _nice_major_unit(data_range, target_ticks=6)

        chart1 = wb.add_chart({'type': 'line'})

        bal_col_letter = xl_col_to_name(balance_idx)
        series_neto = {
            'name':   'Balance Neto',
            'values': f"=Trades!${bal_col_letter}$2:${bal_col_letter}${n_rows + 1}",
            'line':   {'color': '#4A5568', 'width': 1.5},
            'marker': {'type': 'none'},
        }
        if entry_time_idx is not None:
            et_letter = xl_col_to_name(entry_time_idx)
            series_neto['categories'] = f"=Trades!${et_letter}$2:${et_letter}${n_rows + 1}"
        chart1.add_series(series_neto)

        chart1.set_title({'name': 'Balance', 'name_font': {'size': 11, 'bold': True, 'color': '#2D3436', 'name': 'Calibri'}})
        chart1.set_y_axis({
            'num_format':      '#,##0',
            'num_font':        {'size': 8, 'color': '#718096', 'name': 'Calibri'},
            'min':             y_min_axis,
            'max':             y_max_axis,
            'major_unit':      major_unit,
            'major_gridlines': {'visible': True, 'line': {'color': '#EDF2F7', 'width': 0.5}},
            'minor_gridlines': {'visible': False},
            'line':            {'none': True},
        })
        chart1.set_x_axis({
            'num_format':      'MMM yy',
            'num_font':        {'size': 7, 'color': '#A0AEC0', 'name': 'Calibri'},
            'major_gridlines': {'visible': False},
            'major_tick_mark': 'none',
            'minor_tick_mark': 'none',
            'line':            {'color': '#E2E8F0', 'width': 0.5},
        })
        chart1.set_legend({'none': True})
        chart1.set_plotarea({'border': {'none': True}, 'fill': {'color': '#FFFFFF'}})
        chart1.set_chartarea({'border': {'none': True}, 'fill': {'color': '#FFFFFF'}})
        chart1.set_size({'width': 1000, 'height': 400})

        ws.insert_chart(BLOCK_ROW, 1, chart1, {'x_offset': 2, 'y_offset': 5})
        CHARTS_HEIGHT_ROWS += 24  # ~430px ≈ 24 filas

    # ── GRÁFICO 2: BALANCE NETO vs BALANCE BRUTO (comparación) ──────────────
    # Usa datos de _ChartData (hoja oculta) para Balance Bruto
    if balance_idx is not None and has_bruto_data and n_rows > 0:
        # Calcular rango Y incluyendo ambas series
        bal_series = df.iloc[:, balance_idx].dropna()
        y_min_raw2 = float(bal_series.min())
        y_max_raw2 = float(bal_series.max())

        comisiones_acum = 0.0
        for r_idx in range(n_rows):
            try:
                comision_val = float(df.iloc[r_idx, comisiones_idx])
            except (ValueError, TypeError):
                comision_val = 0.0
            comisiones_acum += comision_val
            try:
                balance_val = float(df.iloc[r_idx, balance_idx])
            except (ValueError, TypeError):
                balance_val = 0.0
            bruto_val = balance_val + comisiones_acum
            y_min_raw2 = min(y_min_raw2, bruto_val)
            y_max_raw2 = max(y_max_raw2, bruto_val)

        data_range2 = y_max_raw2 - y_min_raw2 if y_max_raw2 != y_min_raw2 else max(abs(y_max_raw2) * 0.1, 1.0)
        margin2     = data_range2 * 0.03
        y_min_axis2 = y_min_raw2 - margin2
        y_max_axis2 = y_max_raw2 + margin2
        major_unit2 = _nice_major_unit(data_range2, target_ticks=6)

        chart2 = wb.add_chart({'type': 'line'})

        # Serie 1 — BALANCE NETO (azul) — desde hoja Trades
        bal_col_letter = xl_col_to_name(balance_idx)
        series_neto2 = {
            'name':   'Balance Neto',
            'values': f"=Trades!${bal_col_letter}$2:${bal_col_letter}${n_rows + 1}",
            'line':   {'color': '#4A5568', 'width': 1.5},
            'marker': {'type': 'none'},
        }
        if entry_time_idx is not None:
            et_letter = xl_col_to_name(entry_time_idx)
            series_neto2['categories'] = f"=Trades!${et_letter}$2:${et_letter}${n_rows + 1}"
        chart2.add_series(series_neto2)

        # Serie 2 — BALANCE BRUTO (verde) — desde hoja _ChartData
        series_bruto2 = {
            'name':       'Balance Bruto',
            'values':     f"='_ChartData'!$B$2:$B${n_rows + 1}",
            'categories': f"='_ChartData'!$A$2:$A${n_rows + 1}",
            'line':       {'color': '#CBD5E0', 'width': 1.5, 'dash_type': 'dash'},
            'marker':     {'type': 'none'},
        }
        chart2.add_series(series_bruto2)

        chart2.set_title({'name': 'Neto vs Bruto', 'name_font': {'size': 11, 'bold': True, 'color': '#2D3436', 'name': 'Calibri'}})
        chart2.set_y_axis({
            'num_format':      '#,##0',
            'num_font':        {'size': 8, 'color': '#718096', 'name': 'Calibri'},
            'min':             y_min_axis2,
            'max':             y_max_axis2,
            'major_unit':      major_unit2,
            'major_gridlines': {'visible': True, 'line': {'color': '#EDF2F7', 'width': 0.5}},
            'minor_gridlines': {'visible': False},
            'line':            {'none': True},
        })
        chart2.set_x_axis({
            'num_format':      'MMM yy',
            'num_font':        {'size': 7, 'color': '#A0AEC0', 'name': 'Calibri'},
            'major_gridlines': {'visible': False},
            'major_tick_mark': 'none',
            'minor_tick_mark': 'none',
            'line':            {'color': '#E2E8F0', 'width': 0.5},
        })
        chart2.set_legend({'position': 'bottom', 'font': {'size': 9, 'color': '#718096', 'name': 'Calibri'}})
        chart2.set_plotarea({'border': {'none': True}, 'fill': {'color': '#FFFFFF'}})
        chart2.set_chartarea({'border': {'none': True}, 'fill': {'color': '#FFFFFF'}})
        chart2.set_size({'width': 1000, 'height': 400})

        ws.insert_chart(BLOCK_ROW + CHARTS_HEIGHT_ROWS, 1, chart2, {'x_offset': 2, 'y_offset': 5})
        CHARTS_HEIGHT_ROWS += 24

    # ── GRÁFICO 3: ESTRATEGIA vs BUY & HOLD (índice base 100, escala log) ───
    if has_roi_data and entry_price_idx is not None and n_rows > 0:
        chart3 = wb.add_chart({'type': 'line'})

        # Serie 1 — Estrategia (gris oscuro)
        series_strat = {
            'name':       'Estrategia',
            'values':     f"='_ChartData'!$C$2:$C${n_rows + 1}",
            'categories': f"='_ChartData'!$A$2:$A${n_rows + 1}",
            'line':       {'color': '#4A5568', 'width': 1.5},
            'marker':     {'type': 'none'},
        }
        chart3.add_series(series_strat)

        # Serie 2 — Buy & Hold (gris claro)
        series_bh = {
            'name':       'Buy & Hold',
            'values':     f"='_ChartData'!$D$2:$D${n_rows + 1}",
            'categories': f"='_ChartData'!$A$2:$A${n_rows + 1}",
            'line':       {'color': '#CBD5E0', 'width': 1.5},
            'marker':     {'type': 'none'},
        }
        chart3.add_series(series_bh)

        chart3.set_title({'name': 'Estrategia vs Buy & Hold', 'name_font': {'size': 11, 'bold': True, 'color': '#2D3436', 'name': 'Calibri'}})
        chart3.set_y_axis({
            'num_format':      '#,##0"%"',
            'num_font':        {'size': 8, 'color': '#718096', 'name': 'Calibri'},
            'major_gridlines': {'visible': True, 'line': {'color': '#EDF2F7', 'width': 0.5}},
            'minor_gridlines': {'visible': False},
            'line':            {'none': True},
            'crossing':        0,
        })
        chart3.set_x_axis({
            'num_format':      'MMM yy',
            'num_font':        {'size': 7, 'color': '#A0AEC0', 'name': 'Calibri'},
            'major_gridlines': {'visible': False},
            'major_tick_mark': 'none',
            'minor_tick_mark': 'none',
            'line':            {'color': '#E2E8F0', 'width': 0.5},
            'label_position':  'low',
        })
        chart3.set_legend({'position': 'bottom', 'font': {'size': 9, 'color': '#718096', 'name': 'Calibri'}})
        chart3.set_plotarea({'border': {'none': True}, 'fill': {'color': '#FFFFFF'}})
        chart3.set_chartarea({'border': {'none': True}, 'fill': {'color': '#FFFFFF'}})
        chart3.set_size({'width': 1000, 'height': 400})

        ws.insert_chart(BLOCK_ROW + CHARTS_HEIGHT_ROWS, 1, chart3, {'x_offset': 2, 'y_offset': 5})
        CHARTS_HEIGHT_ROWS += 24

    # ── GRÁFICO 4: COMISIONES vs BENEFICIO ──────────────────────────────────
    if has_fee_data and pnl_neto_idx is not None and n_rows > 0:
        chart4 = wb.add_chart({'type': 'line'})

        # Serie 1 — PNL Neto acumulado (gris oscuro sólido)
        chart4.add_series({
            'name':       'Beneficio Neto',
            'values':     f"='_ChartData'!$F$2:$F${n_rows + 1}",
            'categories': f"='_ChartData'!$A$2:$A${n_rows + 1}",
            'line':       {'color': '#4A5568', 'width': 1.5},
            'marker':     {'type': 'none'},
        })

        # Serie 2 — Comisiones acumuladas (gris claro punteado)
        chart4.add_series({
            'name':       'Comisiones',
            'values':     f"='_ChartData'!$E$2:$E${n_rows + 1}",
            'categories': f"='_ChartData'!$A$2:$A${n_rows + 1}",
            'line':       {'color': '#CBD5E0', 'width': 1.5, 'dash_type': 'dash'},
            'marker':     {'type': 'none'},
        })

        chart4.set_title({'name': 'Comisiones vs Beneficio', 'name_font': {'size': 11, 'bold': True, 'color': '#2D3436', 'name': 'Calibri'}})
        chart4.set_y_axis({
            'num_format':      '#,##0',
            'num_font':        {'size': 8, 'color': '#718096', 'name': 'Calibri'},
            'major_gridlines': {'visible': True, 'line': {'color': '#EDF2F7', 'width': 0.5}},
            'minor_gridlines': {'visible': False},
            'line':            {'none': True},
        })
        chart4.set_x_axis({
            'num_format':      'MMM yy',
            'num_font':        {'size': 7, 'color': '#A0AEC0', 'name': 'Calibri'},
            'major_gridlines': {'visible': False},
            'major_tick_mark': 'none',
            'minor_tick_mark': 'none',
            'line':            {'color': '#E2E8F0', 'width': 0.5},
            'label_position':  'low',
        })
        chart4.set_legend({'position': 'bottom', 'font': {'size': 9, 'color': '#718096', 'name': 'Calibri'}})
        chart4.set_plotarea({'border': {'none': True}, 'fill': {'color': '#FFFFFF'}})
        chart4.set_chartarea({'border': {'none': True}, 'fill': {'color': '#FFFFFF'}})
        chart4.set_size({'width': 1000, 'height': 400})

        ws.insert_chart(BLOCK_ROW + CHARTS_HEIGHT_ROWS, 1, chart4, {'x_offset': 2, 'y_offset': 5})
        CHARTS_HEIGHT_ROWS += 24

    # ── TABLA DE PARÁMETROS ─────────────────────────────────────────────────
    TABLE_COL = 1
    TABLE_ROW = BLOCK_ROW + CHARTS_HEIGHT_ROWS

    _TBL = dict(font_name='Calibri', font_size=11, valign='vcenter', indent=1)

    title_fmt = wb.add_format({**_TBL,
        'bold': True, 'font_color': '#2D3436', 'bg_color': '#EDF2F7',
        'align': 'left',
        'bottom': 2, 'bottom_color': '#A0AEC0',
        'top': 0, 'left': 0, 'right': 0,
    })

    def _lbl_fmt(is_last=False):
        return wb.add_format({**_TBL,
            'bold': False, 'font_color': '#718096', 'bg_color': '#FFFFFF',
            'align': 'left',
            'bottom': 1 if not is_last else 0,
            'bottom_color': '#EDF2F7',
            'top': 0, 'left': 0, 'right': 0,
        })

    def _val_fmt(num_fmt='#,##0.00', is_last=False):
        return wb.add_format({**_TBL,
            'bold': True, 'font_color': '#2D3436', 'bg_color': '#FFFFFF',
            'align': 'right', 'num_format': num_fmt,
            'bottom': 1 if not is_last else 0,
            'bottom_color': '#EDF2F7',
            'top': 0, 'left': 0, 'right': 0,
        })

    items = [
        ("Saldo Usado",    val_saldo,   '#,##0.00 $', False),
        ("Volumen Máx.",   val_volumen, '#,##0.00 $', False),
        ("Apalancamiento", val_apal,    '0.00"x"',    True),
    ]

    ws.set_row(TABLE_ROW, 28)
    ws.merge_range(TABLE_ROW, TABLE_COL, TABLE_ROW, TABLE_COL + 1,
                   'Parámetros', title_fmt)

    for i, (label, value, num_fmt, is_last) in enumerate(items):
        r = TABLE_ROW + i + 1
        ws.set_row(r, 24)
        ws.write(r, TABLE_COL,     label, _lbl_fmt(is_last))
        ws.write(r, TABLE_COL + 1, value, _val_fmt(num_fmt, is_last))

    ws.set_column(TABLE_COL,     TABLE_COL,     22)
    ws.set_column(TABLE_COL + 1, TABLE_COL + 1, 18)

    wb.close()


# ==============================================================================
# FALLBACK OPENPYXL
# ==============================================================================

def _escribir_trades_openpyxl_fallback(
    filepath: str,
    df: pd.DataFrame,
    val_saldo: float = 0,
    val_volumen: float = 0,
    val_apal: float = 0,
):
    from openpyxl.formatting.rule import CellIsRule

    df.to_excel(filepath, index=False, sheet_name="Trades", engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    max_row = ws.max_row

    border   = _make_border_op(COLORS["border_color"])
    fill_hdr = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
    fill_alt = PatternFill("solid", fgColor=COLORS["row_alt"])
    fill_w   = PatternFill("solid", fgColor="FFFFFF")
    font_hdr = Font(name=FONT_TITLE, size=11, bold=True, color="FFFFFF")
    font_b   = Font(name=FONT_BODY,  size=10, color=COLORS["text_dark"])
    align_c  = Alignment(horizontal='center', vertical='center')
    align_cw = Alignment(horizontal='center', vertical='center', wrap_text=True)

    col_headers = {c: str(ws.cell(1, c).value or "").upper() for c in range(1, max_col + 1)}
    col_fmt = {}
    pnl_col = None
    for c, hdr in col_headers.items():
        if "TIME" in hdr or "DATE" in hdr:
            col_fmt[c] = "DD/MM/YY HH:MM"
        elif "PRICE" in hdr:
            col_fmt[c] = "#,##0.00"
        elif "PNL" in hdr or "BALANCE" in hdr or "COMISIONES" in hdr:
            col_fmt[c] = "#,##0.00"
        else:
            col_fmt[c] = "General"
        if "PNL" in hdr and "NETO" in hdr:
            pnl_col = c

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font_hdr; cell.alignment = align_cw
        cell.fill = fill_hdr; cell.border = border
    ws.row_dimensions[1].height = 28

    for r in range(2, max_row + 1):
        rf = fill_alt if r % 2 == 0 else fill_w
        ws.row_dimensions[r].height = 18
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_b; cell.alignment = align_c
            cell.border = border; cell.fill = rf
            cell.number_format = col_fmt.get(c, "General")

    if pnl_col:
        col_letter = get_column_letter(pnl_col)
        rng = f"{col_letter}2:{col_letter}{max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='greaterThan', formula=['0'],
            fill=PatternFill("solid", fgColor=COLORS["success_bg"]),
            font=Font(name=FONT_BODY, size=10, color=COLORS["accent_green"], bold=True)
        ))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='lessThan', formula=['0'],
            fill=PatternFill("solid", fgColor=COLORS["danger_bg"]),
            font=Font(name=FONT_BODY, size=10, color=COLORS["accent_red"], bold=True)
        ))

    _col_widths_fast_op(ws, max_col, max_row)
    ws.freeze_panes = ws.cell(row=2, column=1)

    # Gráfico openpyxl con eje Y escalado
    col_map = {str(ws.cell(1, c).value or "").upper().strip(): c for c in range(1, max_col + 1)}
    if "BALANCE" in col_map:
        from openpyxl.chart import LineChart, Reference
        bal_col_idx = col_map["BALANCE"]
        bal_data = [ws.cell(r, bal_col_idx).value for r in range(2, max_row + 1)
                    if isinstance(ws.cell(r, bal_col_idx).value, (int, float))]
        if bal_data:
            data_range = max(bal_data) - min(bal_data) or 1
            margin = data_range * 0.03
            chart2 = LineChart()
            chart2.title  = "EVOLUCIÓN DEL BALANCE"
            chart2.legend = None
            chart2.y_axis.numFmt = '#,##0.00'
            chart2.y_axis.scaling.min = min(bal_data) - margin
            chart2.y_axis.scaling.max = max(bal_data) + margin
            chart2.y_axis.majorGridlines = None
            chart2.x_axis.majorGridlines = None
            data = Reference(ws, min_col=bal_col_idx, min_row=1, max_row=max_row)
            chart2.add_data(data, titles_from_data=True)
            s1 = chart2.series[0]
            s1.graphicalProperties.line.solidFill = "3A86FF"
            s1.graphicalProperties.line.width = 22860
            s1.smooth = True
            s1.marker.symbol = "none"
            chart2.height = 14; chart2.width = 44  # Ampliado (ancho 10 eq)
            ws.add_chart(chart2, f"B{max_row + 3}")

    # Tabla fallback debajo del gráfico (B)
    TABLE_COL_OP = 2   # openpyxl 1-indexed = B
    TABLE_ROW_OP = max_row + 33  # Justo debajo del gráfico (que es alto)
    for i, (label, value) in enumerate([("SALDO", val_saldo), ("VOLUMEN", val_volumen), ("APALANCAMIENTO", val_apal)]):
        r = TABLE_ROW_OP + i
        ws.cell(r, TABLE_COL_OP).value = label
        ws.cell(r, TABLE_COL_OP + 1).value = value

    wb.save(filepath)


# ==============================================================================
# FUNCIÓN PRINCIPAL (CSV → DASHBOARD RESUMEN)
# ==============================================================================

def _load_price_series(
    activo: str,
    datos_dir: str,
    fecha_inicio: Optional[str],
    fecha_fin: Optional[str],
    formato_datos: str,
    n_points: int,
) -> Optional[List[float]]:
    """
    Carga precios de cierre del activo desde el archivo de datos OHLCV y los
    remuestrea a exactamente n_points valores con espaciado uniforme.

    Retorna None si el archivo no existe o hay algún error.
    """
    try:
        import glob as _glob
        import polars as _pl

        activo_up = str(activo).strip().upper()
        base_dir  = datos_dir or "datos"

        # Candidatos en orden de preferencia
        candidates = [
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.feather"),
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.fthr"),
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.parquet"),
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.pq"),
            os.path.join(base_dir, f"{activo_up.lower()}_ohlcv_1m.feather"),
            os.path.join(base_dir, f"{activo_up.lower()}_ohlcv_1m.parquet"),
        ]

        filepath = next((p for p in candidates if os.path.exists(p)), None)
        if filepath is None:
            # Búsqueda glob como fallback
            for pat in [f"{activo_up}_ohlcv_1m.*", f"{activo_up.lower()}_ohlcv_1m.*"]:
                hits = _glob.glob(os.path.join(base_dir, pat))
                if hits:
                    filepath = hits[0]
                    break

        if filepath is None:
            return None

        ext = os.path.splitext(filepath)[1].lower()
        cols = ["timestamp", "close"]
        if ext in (".feather", ".fthr"):
            df_raw = _pl.read_ipc(filepath, columns=cols)
        elif ext in (".parquet", ".pq"):
            df_raw = _pl.read_parquet(filepath, columns=cols)
        else:
            return None

        # Normalizar timestamp a datetime[μs] naive para evitar errores de comparación
        ts_col = df_raw.get_column("timestamp")
        if ts_col.dtype == _pl.Datetime("ns", "UTC") or str(ts_col.dtype).startswith("Datetime"):
            try:
                df_raw = df_raw.with_columns(
                    _pl.col("timestamp").cast(_pl.Datetime("us")).alias("timestamp")
                )
            except Exception:
                try:
                    df_raw = df_raw.with_columns(
                        _pl.col("timestamp").dt.replace_time_zone(None).cast(_pl.Datetime("us")).alias("timestamp")
                    )
                except Exception:
                    pass

        # Filtrado por rango de fechas
        if fecha_inicio:
            try:
                dt_start = _pl.Series([fecha_inicio]).str.to_datetime(format="%Y-%m-%d", strict=False).item()
                df_raw = df_raw.filter(_pl.col("timestamp") >= dt_start)
            except Exception:
                pass
        if fecha_fin:
            try:
                dt_end = _pl.Series([fecha_fin]).str.to_datetime(format="%Y-%m-%d", strict=False).item()
                df_raw = df_raw.filter(_pl.col("timestamp") <= dt_end)
            except Exception:
                pass

        if df_raw.height < 2:
            return None

        prices = df_raw.get_column("close").to_list()
        n_src  = len(prices)

        if n_points <= 1:
            return [prices[0]]

        # Remuestreo: selección de índices equidistantes
        indices  = [int(round(i * (n_src - 1) / (n_points - 1))) for i in range(n_points)]
        resampled = [prices[i] for i in indices]
        return resampled

    except Exception as _e:
        logger.debug(f"_load_price_series: {_e}")
        return None


def convertir_resumen_csv_a_excel(
    *,
    csv_path: str,
    strategy_name: str,
    activo: Optional[str] = None,
    timeframe: Optional[str] = None,
    saldo_inicial: float = 300.0,
    excel_path: Optional[str] = None,
    output_dir: Optional[str] = None,
    # Datos de precio para la gráfica
    datos_dir: str = "datos",
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    formato_datos: str = "feather",
) -> str:
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"No existe el CSV: {csv_path}")

    if output_dir is None and excel_path:
        output_dir = os.path.dirname(str(excel_path)) or None

    activo    = activo or "UNKNOWN"
    timeframe = timeframe or "UNKNOWN"

    df = pd.read_csv(csv_path)
    df, known_param_names = _normalizar_nombres(df, strategy_name)
    df_final, cols_metrics, cols_params = _organizar_y_filtrar_columnas(df, known_param_names)
    df_final = _ordenar_filas(df_final)

    final_excel_path = _generar_nombre_archivo(csv_path, output_dir, str(activo), strategy_name, str(timeframe))
    os.makedirs(os.path.dirname(final_excel_path) or ".", exist_ok=True)

    with pd.ExcelWriter(final_excel_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, startrow=1)

    _aplicar_estilo_avanzado(
        final_excel_path, df_final, cols_metrics, cols_params, saldo_inicial,
    )

    path_to_return = final_excel_path
    if excel_path:
        import shutil
        try:
            if os.path.abspath(excel_path) != os.path.abspath(final_excel_path):
                shutil.move(final_excel_path, excel_path)
                path_to_return = excel_path
        except Exception:
            pass

    try:
        if os.path.exists(csv_path):
            os.remove(csv_path)
    except Exception:
        pass

    return path_to_return


# ==============================================================================
# LÓGICA DE PROCESAMIENTO
# ==============================================================================

def _normalizar_nombres(df: pd.DataFrame, strategy_name: str) -> tuple:
    """
    Normaliza nombres de columnas y detecta cuáles son parámetros de estrategia.

    DETECCIÓN ROBUSTA DE PARÁMETROS:
    ================================
    En el CSV intermedio, los parámetros de la estrategia se guardan con el
    prefijo 'PARAM_' (ej: PARAM_OLS_WINDOW, PARAM_EXIT_SL_PCT). Usamos
    este prefijo como FUENTE DE VERDAD para saber qué es un parámetro,
    en lugar de depender de heurísticas de keywords que fallan con nuevas
    estrategias.

    Returns:
        tuple: (df_normalizado, known_param_names) donde known_param_names
               es un set con los nombres limpios de las columnas que
               originalmente tenían prefijo PARAM_.
    """
    df.columns = [str(c).upper().strip() for c in df.columns]

    # ── FASE 1: Registrar columnas que vienen con prefijo PARAM_ ──────────
    # Estas son DEFINITIVAMENTE parámetros de la estrategia.
    # Guardamos el nombre limpio (sin prefijo) para usarlo después.
    _PARAM_PREFIXES = ("PARAM_", "PARAMS_", "ESTRATEGIA_PARAMS_", "STRATEGY_PARAMS_")
    known_param_names: set = set()
    for col in df.columns:
        for pfx in _PARAM_PREFIXES:
            if col.startswith(pfx):
                clean_name = col[len(pfx):]
                # Excluir params internos (__activo, __saldo, etc.)
                if not clean_name.startswith("_"):
                    known_param_names.add(clean_name)
                break

    # ── FASE 2: Renombrado estándar de métricas y columnas conocidas ──────
    rename_map = {
        "STRATEGY": "ESTRATEGIA", "SHARPE_RATIO": "SHARPE", "TRADES_POR_DIA": "TRADES_DIA",
        "N_TRADES": "TOTAL_TRADES", "NUM_TRADES": "TOTAL_TRADES", "COUNT_TRADES": "TOTAL_TRADES",
        "AVG_TRADE_DURATION": "AVG_TRADE", "DURATION_MEAN_MIN": "AVG_TRADE",
        "RETORNO_PROMEDIO": "AVG_TRADE",
        "WIN_RATE_PCT": "WINRATE_PCT", "WINRATE": "WINRATE_PCT",
        "PORC_GANADORAS": "WINRATE_PCT", "WIN_RATE": "WINRATE_PCT",
        "RACHA_GANADORA": "WIN_STREAK", "RACHA_PERDEDORA": "LOSS_STREAK",
        "MAX_DRAWDOWN_PCT": "MAX_DD_PCT", "MAX_DRAWDOWN": "MAX_DD_PCT",
        "DRAWDOWN": "MAX_DD_PCT", "DD": "MAX_DD_PCT", "DD_PCT": "MAX_DD_PCT", "MAX_DD": "MAX_DD_PCT",
        "RETURN_PCT": "ROI_PCT", "ROI": "ROI_PCT",
        "COUNT_LONGS": "LONG", "N_LONGS": "LONG", "LONGS": "LONG", "NUM_LONGS": "LONG",
        "N_TRADES_LONG": "LONG", "TRADES_LONG": "LONG",
        "COUNT_SHORTS": "SHORT", "N_SHORTS": "SHORT", "SHORTS": "SHORT", "NUM_SHORTS": "SHORT",
        "N_TRADES_SHORT": "SHORT", "TRADES_SHORT": "SHORT",
        "EXIT_SL_PCT": "SL", "P_SL": "SL", "SL_PCT": "SL",
        "EXIT_TP_PCT": "TP", "P_TP": "TP", "TP_PCT": "TP",
        "EXIT_TRAIL_ACT_PCT": "ACT", "TRAIL_ACT": "ACT",
        "EXIT_TRAIL_DIST_PCT": "DIST", "TRAIL_DIST": "DIST", "DISTANCE": "DIST",
        # --- Mapeos para las métricas del nuevo resumen ---
        "NET_PNL": "PNL_NETO",
        "SALDO_BRUTO": "SALDO_SIN_COMISIONES",
        "COMISIONES": "COMISIONES_TOTAL",
        "COMISION_TOTAL": "COMISIONES_TOTAL",
    }

    for old, new in rename_map.items():
        if old in df.columns and new not in df.columns:
            df.rename(columns={old: new}, inplace=True)
        elif old in df.columns and new in df.columns:
            df.drop(columns=[old], inplace=True)

    if "ESTRATEGIA" not in df.columns:
        df.insert(0, "ESTRATEGIA", strategy_name.upper())
    if "TRIAL" not in df.columns and df.index.name != "TRIAL":
        df.insert(0, "TRIAL", range(len(df)))

    # ── FASE 3: Limpiar prefijos y normalizar nombres ─────────────────────
    # Mantenemos un mapeo old_name → clean_name para actualizar known_param_names
    def _canon_metric(col_name: str) -> str:
        return METRIC_ALIASES_TO_CANONICAL.get(col_name, col_name)

    new_cols = []
    updated_param_names: set = set()
    for col in df.columns:
        if col in METRICS_ORDER or col in ID_COLS or col in {"SL", "TP", "ACT", "DIST"}:
            # SL/TP/ACT/DIST son parámetros de salida conocidos, marcarlos
            if col in {"SL", "TP", "ACT", "DIST"}:
                updated_param_names.add(col)
            new_cols.append(col)
            continue

        clean = col
        was_param = False
        for p in PREFIXES_TO_CLEAN:
            if clean.startswith(p):
                # Si tenía prefijo de parámetro, marcarlo
                if p in ("PARAM_", "PARAMS_", "ESTRATEGIA_PARAMS_", "STRATEGY_PARAMS_"):
                    was_param = True
                clean = clean[len(p):]

        # Excluir params internos del sistema (__, doble guión bajo)
        if clean.startswith("_"):
            # Son params internos (__activo, __saldo, etc.), no mostrar
            new_cols.append(clean)
            continue

        # Aplicar rename_map DESPUÉS de limpiar prefijos para atrapar
        # casos como PARAM_EXIT_SL_PCT → (strip PARAM_) → EXIT_SL_PCT → SL
        if clean in rename_map:
            clean = rename_map[clean]
        clean = clean.replace("_PCT", "%").replace("PERCENTAGE", "%")
        clean = _canon_metric(clean)

        # Si el nombre original (sin limpiar prefijo) estaba en known_param_names,
        # o si fue detectado por tener prefijo PARAM_, marcar el nombre limpio
        original_without_prefix = col
        for pfx in _PARAM_PREFIXES:
            if original_without_prefix.startswith(pfx):
                original_without_prefix = original_without_prefix[len(pfx):]
                break
        if was_param or original_without_prefix in known_param_names:
            updated_param_names.add(clean)

        new_cols.append(clean)

    df.columns = new_cols

    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()]

    return df, updated_param_names


def _organizar_y_filtrar_columnas(df: pd.DataFrame, known_param_names: set = None):
    """
    Organiza columnas en: ID | MÉTRICAS | PARÁMETROS.

    DETECCIÓN ROBUSTA:
    ==================
    1. PRIORIDAD 1 — known_param_names: Columnas que vinieron con prefijo PARAM_
       en el CSV original. Estas son DEFINITIVAMENTE parámetros, sin importar
       su nombre. Esto garantiza que cualquier parámetro de cualquier estrategia
       nueva sea detectado correctamente.

    2. PRIORIDAD 2 — Fallback heurístico: Para columnas que no están en
       known_param_names (ej: datos legacy sin prefijo), se usa la heurística
       de keywords existente como respaldo.
    """
    if known_param_names is None:
        known_param_names = set()

    cols = list(df.columns)
    current_ids = [c for c in ID_COLS if c in cols]

    current_metrics = []
    for m in ALL_METRICS_ORDER:
        for t in [m, m.replace("_PCT", "%")]:
            if t in cols:
                current_metrics.append(t)
                break
    current_metrics = list(dict.fromkeys(current_metrics))

    excluded = set(current_ids + current_metrics) | ALL_KNOWN_METRICS
    candidates = [c for c in cols if c not in excluded]

    exit_cols = {"SL", "TP", "ACT", "DIST"}

    # Determinar si es TRAILING o FIXED
    is_trailing = False
    exit_col_name = next((c for c in df.columns if str(c).upper() in ["EXIT_TYPE", "PARAM_EXIT_TYPE", "PARAM_EXIT_TYPE"]), None)
    if exit_col_name:
        is_trailing = df[exit_col_name].astype(str).str.contains("TRAIL", case=False).any()
    else:
        for t_col in ["ACT", "DIST"]:
            if t_col in df.columns:
                try:
                    if (pd.to_numeric(df[t_col], errors='coerce').fillna(0) > 0).any():
                        is_trailing = True; break
                except Exception:
                    pass

    # ── Clasificación de parámetros con doble vía ─────────────────────────
    very_bad = {"PROFIT", "WIN", "SALDO", "BALANCE", "DRAWDOWN", "DD",
                "ROI", "RETORNO", "NUM_", "COUNT", "TRADES", "RESULT", "METRIC"}
    exceptions = {"STOP", "SL", "TP", "TRAIL", "TIME", "PERIOD", "LEN", "FAST",
                  "SLOW", "SIGNAL", "LIMIT", "THRESHOLD", "SIGMA", "OFFSET", "ATR", "ACT", "DIST"}

    current_params = []
    for c in candidates:
        # Saltar columnas vacías o internas del sistema
        if df[c].astype(str).str.strip().eq("").all() or c.startswith("_"):
            continue
        if c in EXCLUDED_PARAMS or c.replace("%", "_PCT") in EXCLUDED_PARAMS:
            continue

        # ── VÍA 1: ¿Está en known_param_names? → ES PARÁMETRO (definitivo) ──
        if c in known_param_names:
            # Respetar lógica de trailing para ACT/DIST
            if c in {"ACT", "DIST"} and not is_trailing:
                continue
            current_params.append(c)
            continue

        # ── VÍA 2: Exit cols conocidas ───────────────────────────────────────
        if c in exit_cols:
            if c in {"ACT", "DIST"} and not is_trailing:
                continue
            current_params.append(c)
            continue

        # ── VÍA 3: Fallback heurístico (para datos legacy sin prefijo) ──────
        c_upper = c.upper()
        is_garbage = False
        for kw in METRIC_KEYWORDS_TO_DROP:
            if kw in c_upper:
                if any(bw in c_upper for bw in very_bad):
                    is_garbage = True; break
                if not any(ex in c_upper for ex in exceptions):
                    is_garbage = True; break
        if not is_garbage:
            current_params.append(c)

    current_params.sort()
    final_cols = current_ids + current_metrics + current_params
    return df[final_cols], current_metrics, current_params


def _ordenar_filas(df: pd.DataFrame) -> pd.DataFrame:
    if "SCORE" in df.columns:
        return df.sort_values("SCORE", ascending=False).reset_index(drop=True)
    if "SALDO_ACTUAL" in df.columns:
        return df.sort_values("SALDO_ACTUAL", ascending=False).reset_index(drop=True)
    return df


def _generar_nombre_archivo(csv, out_dir, activo, est, tf) -> str:
    fname = (f"RESUMEN_{activo}_{re.sub(r'[^A-Z0-9]', '', est.upper())}"
             f"_{re.sub(r'[^a-zA-Z0-9]', '', tf.lower())}.xlsx")
    return os.path.join(out_dir or os.path.dirname(csv), fname)


# ==============================================================================
# HELPERS OPENPYXL
# ==============================================================================

def _make_border_op(color: str, style: str = 'thin') -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _col_widths_fast_op(ws, max_col: int, max_row: int, header_row: int = 1, sample: int = 20):
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = len(str(ws.cell(row=header_row, column=col).value or ""))
        for r in range(header_row + 1, min(header_row + sample + 1, max_row + 1)):
            try:
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min((max_len + 2) * 1.15, 28)


# ==============================================================================
# ESTILOS DASHBOARD RESUMEN
# ==============================================================================

def _aplicar_estilo_avanzado(
    filepath, df, metrics_cols, params_cols, saldo_ini,
):
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import column_index_from_string
    from openpyxl.formatting.rule import CellIsRule

    wb = load_workbook(filepath)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    max_row = ws.max_row

    n_ids         = len([c for c in ID_COLS if c in df.columns])
    n_metrics     = len(metrics_cols)
    n_params      = len(params_cols)
    start_metrics = n_ids + 1
    end_metrics   = start_metrics + n_metrics - 1
    start_params  = end_metrics + 1
    end_params    = start_params + n_params - 1

    # ── Colores minimalistas ────────────────────────────────────────────────
    C_HDR      = "F7FAFC"   # fondo cabecera (gris muy claro)
    C_HDR_FONT = "718096"   # texto cabecera (gris medio)

    border_thin = _make_border_op("E2E8F0", style='thin')
    border_bot  = Border(
        bottom=Side(style='medium', color='A0AEC0'),
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
    )

    fill_hdr     = PatternFill("solid", fgColor=C_HDR)
    fill_section = PatternFill("solid", fgColor="EDF2F7")
    fill_alt     = PatternFill("solid", fgColor=COLORS["row_alt"])
    fill_w       = PatternFill("solid", fgColor="FFFFFF")

    font_section = Font(name=FONT_TITLE, size=9, bold=True, color="A0AEC0")
    font_hdr     = Font(name=FONT_TITLE, size=10, bold=True, color=C_HDR_FONT)
    font_body    = Font(name=FONT_BODY,  size=10, color=COLORS["text_dark"])
    font_id      = Font(name=FONT_BODY,  size=10, bold=True, color=COLORS["text_dark"])
    align_c      = Alignment(horizontal='center', vertical='center')
    align_l      = Alignment(horizontal='left',   vertical='center', indent=1)
    align_cw     = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Fila 1: Encabezados de sección (texto gris sutil) ───────────────────
    ws.row_dimensions[1].height = 20

    def _section(c_start, c_end, label, fill):
        c = ws.cell(row=1, column=c_start)
        c.value = label; c.fill = fill; c.font = font_section; c.alignment = align_c
        if c_end > c_start:
            ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)

    if n_ids > 0:
        _section(1, n_ids, "DATOS", fill_section)
    if n_metrics > 0:
        _section(start_metrics, end_metrics, "MÉTRICAS", fill_section)
    if n_params > 0:
        _section(start_params, end_params, "PARÁMETROS", fill_section)

    # ── Fila 2: Nombres de columna ──────────────────────────────────────────
    ws.row_dimensions[2].height = 32

    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = font_hdr; cell.alignment = align_cw
        cell.border = border_bot
        cell.fill = fill_hdr

    # ── Aplicar nombres de display en fila 2 ────────────────────────────────
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        orig = str(cell.value or "").strip()
        if orig in METRIC_DISPLAY_NAMES:
            cell.value = METRIC_DISPLAY_NAMES[orig]

    # Leer cabeceras ya con display names
    col_hdrs = {c: str(ws.cell(2, c).value or "").strip() for c in range(1, max_col + 1)}

    # ── Filas de datos ───────────────────────────────────────────────────────
    # Columnas monetarias y porcentuales (tras renombrar)
    MONEY_HDRS = {"BEN BRUTO", "BEN NETO", "SALDO ACTUAL", "COMISIONES"}
    PCT_HDRS   = {"ROI", "MAX DD"}
    INT_HDRS   = {"LONG", "SHORT", "TRIAL"}

    for r in range(3, max_row + 1):
        rf = fill_alt if r % 2 == 0 else fill_w
        ws.row_dimensions[r].height = 22

        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            hdr  = col_hdrs[c]
            is_id_col = (c < start_metrics)

            cell.font      = font_id if is_id_col else font_body
            cell.alignment = align_l if (hdr == "ESTRATEGIA") else align_c
            cell.border    = border_thin
            cell.fill      = rf

            if hdr == "TRADES DIA":
                cell.number_format = "0.00"
            elif hdr in INT_HDRS:
                cell.number_format = "0"
            elif hdr in PCT_HDRS:
                cell.number_format = "0.00%"
                try:
                    cell.value = float(cell.value) / 100.0
                except Exception:
                    pass
            elif hdr == "PROFIT FACTOR":
                cell.number_format = "0.00"
            elif hdr == "EXPECTATIVA":
                cell.number_format = "#,##0.00"
            elif hdr in MONEY_HDRS:
                cell.number_format = "#,##0.00"
            elif hdr == "SCORE":
                cell.number_format = "0.00"

    _col_widths_fast_op(ws, max_col, max_row, header_row=2, sample=20)

    # ── Mapa columna-display-name → letra Excel ──────────────────────────────
    col_map = {col_hdrs[c]: get_column_letter(c) for c in range(1, max_col + 1)}

    # ── Formato condicional ──────────────────────────────────────────────────
    data_start = f"3:{max_row}"

    if "BEN NETO" in col_map:
        rng = f"{col_map['BEN NETO']}3:{col_map['BEN NETO']}{max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='greaterThan', formula=['0'],
            font=Font(name=FONT_BODY, size=10, color=COLORS["accent_green"], bold=True)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='lessThan', formula=['0'],
            font=Font(name=FONT_BODY, size=10, color=COLORS["accent_red"], bold=True)))

    ws.freeze_panes = ws.cell(row=3, column=start_metrics)

    # ── SIN GRÁFICO EN RESUMEN — La gráfica comparativa está en los trials ──

    wb.save(filepath)


# ==============================================================================
# EXPORTACIÓN RÁPIDA
# ==============================================================================

def exportar_trades_excel_rapido(
    df_trades: pd.DataFrame,
    resumen_csv_path: str,
    metrics: dict,
    params: dict,
    trial_number: int,
    trades_actual_base: str = "trades_trial",
    score: float = None,
    max_archivos: int = 5,
    skip_trades_file: bool = False,
):
    params = dict(params or {})
    fila = {
        "TRIAL":      trial_number,
        "SCORE":      score if score is not None else 0,
        "ESTRATEGIA": params.get("NOMBRE_COMBO", "UNKNOWN"),
    }
    for k, v in metrics.items():
        fila[k.upper()] = v

    def _aplanar(d, prefix=""):
        out = {}
        for k, v in d.items():
            if isinstance(v, dict):
                out.update(_aplanar(v, f"{prefix}{k.upper()}_"))
            else:
                out[f"{prefix}{k.upper()}"] = v
        return out

    fila.update(_aplanar(params))
    df_fila = pd.DataFrame([fila])

    if int(trial_number) == 0 and os.path.exists(resumen_csv_path):
        try:
            os.remove(resumen_csv_path)
        except Exception:
            pass

    mode = "w" if not os.path.exists(resumen_csv_path) else "a"
    df_fila.to_csv(resumen_csv_path, index=False, mode=mode, header=(mode == "w"))

    if not skip_trades_file:
        _gestionar_archivos_trades(df_trades, trades_actual_base, trial_number, score, max_archivos, params=params)


def _gestionar_archivos_trades(df, base_path, trial, score, max_files, params=None):
    trades_dir = os.path.dirname(base_path) or "."
    os.makedirs(trades_dir, exist_ok=True)

    s_val = score if score is not None else -999
    fpath = os.path.join(trades_dir, f"TRADES_TRIAL{trial}_SCORE{s_val:.2f}.xlsx")

    df_export = df.copy()
    for col in df_export.select_dtypes(include=["datetime64[ns, UTC]", "datetime64[ns]"]).columns:
        if hasattr(df_export[col].dt, "tz") and df_export[col].dt.tz is not None:
            df_export[col] = df_export[col].dt.tz_localize(None)

    df_export = _preparar_df_trades(df_export)

    saldo = (params or {}).get('__saldo_usado') or 0
    apal  = (params or {}).get('__apalancamiento_max') or 0

    try:
        _escribir_trades_xlsxwriter(fpath, df_export, saldo, saldo * apal, apal)
    except Exception:
        _escribir_trades_openpyxl_fallback(fpath, df_export, saldo, saldo * apal, apal)

    files = [f for f in os.listdir(trades_dir) if f.startswith("TRADES_TRIAL") and f.endswith(".xlsx")]
    if len(files) > max_files:
        scored = []
        for f in files:
            try:
                s = float(re.search(r"SCORE(-?\d+\.?\d*)", f).group(1))
                scored.append((s, f))
            except Exception:
                scored.append((-9999, f))
        scored.sort(key=lambda x: x[0], reverse=True)
        for _, f_del in scored[max_files:]:
            try:
                os.remove(os.path.join(trades_dir, f_del))
            except Exception:
                pass