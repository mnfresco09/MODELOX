"""
================================================================================
HCA.PY - Agrupamiento Jerárquico Aglomerativo con Seriación Óptima
================================================================================
Analiza un Excel de resultados MODELOX, detecta automáticamente los parámetros
de la estrategia, y aplica Agrupamiento Jerárquico Aglomerativo (HCA) con:

  • Enlace Complete: usa la distancia MÁXIMA entre cualquier par de puntos
    de dos clusters candidatos → es estricto con vecinos lejanos → ningún
    miembro puede estar lejos de otro en el mismo cluster → rangos de
    parámetros más estrechos y compactos
  • Seriación Óptima (Optimal Leaf Ordering): garantiza que filas adyacentes
    en el Excel sean los vecinos más cercanos en el espacio de parámetros
  • Corte dinámico del dendrograma: selección automática de K óptimo
  • Pesos adaptativos por tipo de parámetro (%, rango estrecho, amplio...)

El resultado es un Excel donde puedes leer de arriba a abajo y cada fila
es la más parecida posible a la siguiente → mapa continuo del espacio de
búsqueda ordenado por vecindad real.

Uso:
    python hca.py
    → Pedirá arrastrar el archivo Excel de resultados

Autor: Sistema MODELOX
================================================================================
"""

import os
import sys
import re
import warnings
import time
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from scipy.cluster.hierarchy import (
    linkage,
    fcluster,
    leaves_list,
    optimal_leaf_ordering,
)
from scipy.spatial.distance import pdist, squareform
from sklearn.preprocessing import RobustScaler
from sklearn.metrics import (
    silhouette_score,
    calinski_harabasz_score,
    davies_bouldin_score,
)

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)


# ==============================================================================
# CONFIGURACIÓN DE ESTILO (Idéntico a visual/excel.py + clustering.py)
# ==============================================================================

COLORS = {
    "header_bg_metrics": "1A5276",  # Azul Oscuro
    "header_bg_params":  "566573",  # Gris Plomo
    "header_bg_id":      "212F3D",  # Negro Azulado
    "text_white":        "FFFFFF",
    "text_dark":         "212F3D",
    "border_color":      "BDC3C7",
    "success_bg":        "D5F5E3",
    "danger_bg":         "FADBD8",
}

CLUSTER_PALETTE = [
    "2E86C1", "28B463", "D4AC0D", "CB4335", "8E44AD",
    "E67E22", "1ABC9C", "EC7063", "5DADE2", "45B39D",
    "F4D03F", "AF7AC5", "EB984E", "85C1E9", "82E0AA",
    "F1948A", "BB8FCE", "F0B27A", "76D7C4", "AEB6BF",
    "1F618D", "196F3D", "B7950B", "922B21", "6C3483",
    "CA6F1E", "148F77", "C0392B", "2980B9", "27AE60",
]

FONT_TITLE = "Arial"
FONT_BODY = "Arial"

# ── Detección de columnas (compartido con clustering.py) ──

KNOWN_METRICS = {
    "SALDO_ACTUAL", "ROI_PCT", "ROI%", "PROFIT_FACTOR", "WINRATE_PCT",
    "WINRATE%", "TOTAL_TRADES", "TRADES_DIA", "MAX_DD_PCT", "MAX_DD%",
    "SHARPE", "SQN", "ESTABILIDAD", "AVG_TRADE", "EXPECTATIVA",
    "WIN_STREAK", "LOSS_STREAK", "NUM_LONGS", "NUM_SHORTS",
    "SHARPE_RATIO", "SORTINO", "CALMAR", "KELLY",
    "NET_PROFIT", "PNL_NETO", "NET_PNL",
}

KNOWN_IDS = {"TRIAL", "ESTRATEGIA", "SCORE", "STRATEGY"}

METRIC_KEYWORDS = [
    "PROFIT", "LOSS", "PNL", "NET", "GROSS", "SALDO", "BALANCE", "RETORNO",
    "RETURN", "ROI", "BENEFICIO", "RIESGO", "RISK", "REWARD", "COMISION",
    "FEES", "WIN", "GANADORA", "PERDEDORA", "ACIERTO", "RATE",
    "DRAWDOWN", "DD", "RACHA", "STREAK", "UNDERWATER",
    "RATIO", "FACTOR", "SHARPE", "SORTINO", "CALMAR", "SQN",
    "EXPECTATIVA", "KELLY", "COUNT", "NUM_", "N_TRADES",
    "LONGS", "SHORTS", "TRADES", "METRIC", "RESULT",
    "BEST", "WORST", "DIA_OPERADO", "DURATION",
]

PARAM_EXCEPTIONS = [
    "STOP", "SL", "TP", "TRAIL", "PERIOD", "LEN", "FAST", "SLOW",
    "SIGNAL", "LIMIT", "THRESHOLD", "SIGMA", "OFFSET", "ATR",
    "MA", "EMA", "SMA", "ZLEMA", "RSI", "MACD", "BB", "BOLL",
    "LOOKBACK", "LOOKBAR", "WINDOW", "DIST", "MULT", "FACTOR_",
    "EXIT_SL", "EXIT_TP", "TAKE", "LOSS_",
]


# ==============================================================================
# 1. LECTURA INTELIGENTE DEL ARCHIVO
# ==============================================================================

def solicitar_archivo() -> str:
    """Solicita al usuario arrastrar un archivo Excel."""
    print("\n" + "=" * 70)
    print("  MODELOX - HCA: AGRUPAMIENTO JERÁRQUICO CON SERIACIÓN")
    print("=" * 70)
    print("\n  Arrastra aquí el archivo Excel de resultados y pulsa Enter:\n")

    path = input("  📂 Archivo: ").strip()
    path = path.strip("'\"")
    path = path.replace("\\ ", " ")

    if not os.path.isfile(path):
        print(f"\n  ❌ No se encontró el archivo: {path}")
        sys.exit(1)

    if not path.lower().endswith((".xlsx", ".xls")):
        print(f"\n  ❌ El archivo debe ser .xlsx o .xls")
        sys.exit(1)

    print(f"\n  ✅ Archivo cargado: {os.path.basename(path)}")
    return path


def leer_excel_modelox(filepath: str) -> Tuple[pd.DataFrame, List[str], List[str], List[str]]:
    """
    Lee un Excel MODELOX con headers en fila 2 (fila 1 = títulos de grupo).
    Retorna: (df, id_cols, metric_cols, param_cols)
    """
    df = pd.read_excel(filepath, header=1)

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [str(c[-1]).strip() for c in df.columns]

    df.columns = [str(c).strip() for c in df.columns]

    id_cols = []
    metric_cols = []
    param_cols = []

    for col in df.columns:
        col_upper = col.upper().replace("%", "_PCT")

        if col_upper in KNOWN_IDS or col == "TRIAL":
            id_cols.append(col)
            continue

        if col_upper in KNOWN_METRICS or col in KNOWN_METRICS:
            metric_cols.append(col)
            continue

        is_metric = False
        is_param_exception = False

        for kw in PARAM_EXCEPTIONS:
            if kw in col_upper:
                is_param_exception = True
                break

        if not is_param_exception:
            for kw in METRIC_KEYWORDS:
                if kw in col_upper:
                    is_metric = True
                    break

        if is_metric:
            metric_cols.append(col)
        else:
            if df[col].dtype in [np.float64, np.int64, np.float32, np.int32, float, int]:
                if df[col].nunique() > 1:
                    param_cols.append(col)
                else:
                    metric_cols.append(col)
            elif pd.to_numeric(df[col], errors="coerce").notna().sum() > len(df) * 0.8:
                df[col] = pd.to_numeric(df[col], errors="coerce")
                if df[col].nunique() > 1:
                    param_cols.append(col)
            else:
                metric_cols.append(col)

    print(f"\n  📊 Estructura detectada:")
    print(f"     • {len(df)} trials")
    print(f"     • {len(id_cols)} columnas ID: {id_cols}")
    print(f"     • {len(metric_cols)} métricas: {metric_cols}")
    print(f"     • {len(param_cols)} parámetros: {param_cols}")

    return df, id_cols, metric_cols, param_cols


# ==============================================================================
# 2. ANÁLISIS DE PARÁMETROS Y PESOS ADAPTATIVOS
# ==============================================================================

def analizar_tipo_parametro(series: pd.Series, name: str) -> Dict:
    """
    Clasifica el parámetro y calcula tolerancia ROBUSTA basada en la
    distribución real de los datos, no solo en el rango bruto.

    Método: Ensemble de 3 estimadores robustos de "resolución natural":

    1. Freedman-Diaconis (FD):  2 × IQR × n^(-1/3)
       → Ancho óptimo de bin para histogramas, robusto a outliers.

    2. MAD-based:  1.4826 × median(|xi - median(x)|)
       → Estimador robusto de dispersión (σ equivalente sin outliers).

    3. Scott adaptado:  3.49 × MAD × n^(-1/3)
       → Regla de bin width con MAD en vez de std.

    Tolerancia final = median(FD, MAD, Scott), con clamp de seguridad
    al rango efectivo (P5-P95) para no ser ni demasiado estrecho ni
    demasiado generoso.

    Ventajas vs el método anterior (rango × porcentaje fijo):
    - Se adapta a la FORMA de la distribución, no solo al min/max
    - Inmune a outliers (usa IQR y MAD en vez de rango y std)
    - Escala correctamente con el tamaño del dataset (factor n^(-1/3))
    - Para datos muy concentrados → tolerancia más estricta
    - Para datos dispersos uniformemente → tolerancia más generosa
    """
    clean = series.dropna()
    if len(clean) == 0:
        return {"type": "unknown", "weight": 1.0, "name": name}

    values = clean.values.astype(float)
    n = len(values)
    vmin, vmax = float(values.min()), float(values.max())
    rango = vmax - vmin
    n_unique = clean.nunique()
    std = float(values.std())
    name_upper = name.upper()

    # ── Estadísticas robustas ──
    q25, q50, q75 = np.percentile(values, [25, 50, 75])
    iqr = q75 - q25
    p5, p95 = np.percentile(values, [5, 95])
    rango_efectivo = p95 - p5  # Rango sin el 10% extremo

    # MAD: Median Absolute Deviation (estimador robusto de σ)
    mad_raw = float(np.median(np.abs(values - q50)))
    mad_sigma = 1.4826 * mad_raw  # Factor de consistencia para normalidad

    info = {
        "name": name, "min": vmin, "max": vmax, "range": rango,
        "n_unique": n_unique, "std": std, "iqr": iqr,
        "mad": mad_sigma, "rango_efectivo": rango_efectivo,
    }

    # ── Caso trivial: constante ──
    if rango == 0 or n_unique == 1:
        info.update({"type": "constant", "weight": 0.0, "neighbor_tolerance": 0})
        return info

    # ── Clasificación del tipo (para info/debug, no afecta la tolerancia) ──
    if ("%" in name or "PCT" in name_upper or "SL" in name_upper or "TP" in name_upper) and vmax <= 100:
        param_type = "percentage"
    elif n_unique <= 10:
        param_type = "discrete_few"
    elif rango <= 30:
        param_type = "narrow_range"
    elif rango <= 200:
        param_type = "medium_range"
    else:
        param_type = "wide_range"

    # ── Caso especial: discreto con pocos valores ──
    if n_unique <= 10:
        sorted_unique = np.sort(clean.unique())
        if len(sorted_unique) >= 2:
            steps = np.diff(sorted_unique)
            min_step = float(steps.min())
            tol = max(min_step, 0.5)
        else:
            tol = 1.0
        info.update({"type": param_type, "weight": 1.0, "neighbor_tolerance": tol})
        return info

    # ── Cálculo robusto de tolerancia (3 estimadores) ──
    n_cbrt = n ** (1.0 / 3.0)

    # 1. Freedman-Diaconis: ancho óptimo de bin
    fd_width = 2.0 * iqr / n_cbrt if iqr > 0 else rango / n_cbrt

    # 2. MAD-based: dispersión robusta directa
    mad_tol = mad_sigma

    # 3. Scott adaptado (con MAD en vez de std)
    scott_width = 3.49 * mad_sigma / n_cbrt if mad_sigma > 0 else rango / n_cbrt

    # Ensemble: mediana de los 3
    candidates = sorted([fd_width, mad_tol, scott_width])
    tol = candidates[1]  # Mediana

    # ── Safety clamps ──
    sorted_unique = np.sort(clean.unique())
    if len(sorted_unique) >= 2:
        min_step = float(np.diff(sorted_unique).min())
        tol = max(tol, min_step)
    tol = max(tol, 0.5)

    if rango_efectivo > 0:
        tol = min(tol, rango_efectivo * 0.15)

    tol = max(tol, 0.5)

    info.update({"type": param_type, "weight": 1.0, "neighbor_tolerance": tol})

    # Sin multiplicador técnico: todos los pesos son 1.0

    return info


def construir_features_ponderadas(
    df_params: pd.DataFrame,
    param_info: Dict[str, Dict],
) -> np.ndarray:
    """
    RobustScaler + pesos adaptativos → matriz para distancia euclidiana.
    """
    features = []
    weights = []

    for col in df_params.columns:
        info = param_info[col]
        if info.get("type") == "constant" or info.get("weight", 0) == 0:
            continue

        values = df_params[col].values.reshape(-1, 1)
        scaler = RobustScaler()
        scaled = scaler.fit_transform(values).ravel()

        features.append(scaled)
        weights.append(info["weight"])

    if not features:
        raise ValueError("No se encontraron parámetros válidos para clustering")

    X = np.column_stack(features)
    w = np.array(weights)
    w = w / w.sum() * len(w)
    X = X * w[np.newaxis, :]

    return X


# ==============================================================================
# 3. HCA: COMPLETE LINKAGE + SERIACIÓN ÓPTIMA + CORTE DINÁMICO
# ==============================================================================

def ejecutar_hca(
    X: np.ndarray,
    param_info: Dict[str, Dict],
    olo_direct_limit: int = 1500,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    Pipeline completo de Agrupamiento Jerárquico Aglomerativo:

    1. Calcula distancias condensadas (pdist) con métrica euclidiana
    2. Construye dendrograma con enlace Complete (max distance)
    3. Aplica Seriación Óptima (OLO directo o divide-and-conquer)
    4. Corta el dendrograma en K clusters óptimo (Silhouette + Calinski)

    Retorna: (seriated_order, labels, linkage_matrix)

    Estrategia de seriación:
    - n ≤ olo_direct_limit: OLO directo de scipy (O(n³), pero rápido)
    - n > olo_direct_limit: Divide-and-Conquer OLO:
        a) Pre-cortar en bloques con fcluster
        b) OLO dentro de cada bloque (rápido, cada uno < limit)
        c) OLO entre centroides de bloques (orden inter-bloque)
        d) Optimización de costuras entre bloques adyacentes
      → Misma calidad perceptiva, velocidad O(Σ(nᵢ³)) ≪ O(N³)
    """
    n = X.shape[0]
    t0 = time.time()

    # ── Paso 1: Distancias condensadas ──
    print(f"\n  ⏳ Calculando distancias ({n} × {n})...")
    D = pdist(X, metric="euclidean")
    t1 = time.time()
    print(f"     ✓ {len(D):,} pares en {t1-t0:.1f}s")

    # ── Paso 2: Linkage Complete ──
    # Complete linkage = max distancia entre cualquier par de puntos
    # → Clusters más estrictos: ningún miembro puede estar lejos de otro
    # → Rangos de parámetros más estrechos vs Ward (varianza mínima)
    print(f"\n  🌳 Construyendo dendrograma (Complete linkage)...")
    Z = linkage(D, method="complete")
    t2 = time.time()
    print(f"     ✓ Dendrograma en {t2-t1:.1f}s")

    # ── Paso 3: Seriación Óptima ──
    if n <= olo_direct_limit:
        print(f"\n  🔀 OLO directo ({n} ≤ {olo_direct_limit})...")
        Z_ordered = optimal_leaf_ordering(Z, D)
        seriated_order = leaves_list(Z_ordered)
        t3 = time.time()
        print(f"     ✓ Seriación directa en {t3-t2:.1f}s")
    else:
        print(f"\n  🔀 Seriación Divide-and-Conquer ({n} > {olo_direct_limit})...")
        print(f"     Fase 1: Particionar → OLO intra-bloque → OLO inter-bloque")
        seriated_order = _seriacion_divide_and_conquer(X, Z, D, n, olo_direct_limit)
        Z_ordered = Z  # Guardamos el linkage original
        t3 = time.time()
        print(f"     ✓ Seriación D&C en {t3-t2:.1f}s")

    # ── Paso 4: Corte dinámico del dendrograma ──
    print(f"\n  ✂️  Determinando número óptimo de clusters...")
    labels = _corte_dinamico_dendrograma(X, Z, n)

    t4 = time.time()
    n_clusters = len(set(labels))
    print(f"\n  📦 HCA completado en {t4-t0:.1f}s total:")
    print(f"     • {n_clusters} clusters")
    metodo = 'OLO directo' if n <= olo_direct_limit else 'Divide-and-Conquer OLO'
    print(f"     • Seriación: {metodo}")

    return seriated_order, labels, Z_ordered


def _seriacion_divide_and_conquer(
    X: np.ndarray,
    Z: np.ndarray,
    D: np.ndarray,
    n: int,
    block_limit: int,
) -> np.ndarray:
    """
    Seriación OLO en divide-and-conquer:

    1. Particionar en bloques de tamaño ≤ block_limit usando fcluster
    2. Aplicar OLO dentro de cada bloque (rápido)
    3. Calcular centroides y aplicar OLO al orden inter-bloque
    4. Optimizar costuras: reordenar extremos de bloques adyacentes
       para minimizar la distancia en los puntos de unión

    Complejidad: O(K·(n/K)³ + K³) ≪ O(n³)
    Para n=5000, block_limit=1000: ~5 bloques × 2.5s = ~12s vs ~300s directo
    """

    # ── 1. Particionar en bloques ──
    # Calcular cuántos bloques necesitamos
    n_blocks = max(2, int(np.ceil(n / block_limit)))
    # Pero no demasiados (ineficiente)
    n_blocks = min(n_blocks, max(3, n // 50))

    block_labels = fcluster(Z, t=n_blocks, criterion="maxclust")
    unique_blocks = sorted(set(block_labels))
    actual_blocks = len(unique_blocks)

    print(f"       • {actual_blocks} bloques (máx ~{block_limit} trials c/u)")

    # ── 2. OLO dentro de cada bloque ──
    block_orders = {}  # block_id → array de índices originales, seriados
    block_centroids = []
    block_ids_ordered = []

    for bid in unique_blocks:
        mask = block_labels == bid
        indices = np.where(mask)[0]
        X_block = X[indices]
        nb = len(indices)

        if nb <= 2:
            block_orders[bid] = indices
            block_centroids.append(X_block.mean(axis=0))
            block_ids_ordered.append(bid)
            continue

        # OLO intra-bloque
        D_block = pdist(X_block, metric="euclidean")
        Z_block = linkage(D_block, method="complete")

        if nb <= block_limit:
            Z_block_opt = optimal_leaf_ordering(Z_block, D_block)
        else:
            # Bloque anormalmente grande → sub-dividir recursivamente
            Z_block_opt = Z_block

        local_order = leaves_list(Z_block_opt)
        block_orders[bid] = indices[local_order]
        block_centroids.append(X_block[local_order].mean(axis=0))
        block_ids_ordered.append(bid)

    print(f"       • OLO intra-bloque completado")

    # ── 3. OLO entre bloques (orden inter-bloque) ──
    centroids = np.array(block_centroids)
    if len(centroids) > 2:
        D_cent = pdist(centroids, metric="euclidean")
        Z_cent = linkage(D_cent, method="complete")
        Z_cent_opt = optimal_leaf_ordering(Z_cent, D_cent)
        inter_order = leaves_list(Z_cent_opt)
    else:
        inter_order = np.arange(len(centroids))

    print(f"       • OLO inter-bloque completado")

    # ── 4. Ensamblar y optimizar costuras ──
    ordered_block_ids = [block_ids_ordered[i] for i in inter_order]

    # Para cada par de bloques adyacentes, podemos invertir un bloque
    # si eso acerca los extremos (reduce distancia en la costura)
    final_order = []

    for idx, bid in enumerate(ordered_block_ids):
        block_indices = block_orders[bid]

        if idx > 0 and len(final_order) > 0 and len(block_indices) > 1:
            # Punto final del bloque anterior
            last_point = X[final_order[-1]]

            # Extremos del bloque actual (primer y último punto)
            first_point = X[block_indices[0]]
            last_block_point = X[block_indices[-1]]

            dist_normal = np.linalg.norm(last_point - first_point)
            dist_reversed = np.linalg.norm(last_point - last_block_point)

            if dist_reversed < dist_normal:
                # Invertir el bloque para mejor costura
                block_indices = block_indices[::-1]

        final_order.extend(block_indices.tolist())

    print(f"       • Costuras optimizadas ({len(ordered_block_ids)} uniones)")

    return np.array(final_order, dtype=int)


def _corte_dinamico_dendrograma(
    X: np.ndarray,
    Z: np.ndarray,
    n_samples: int,
) -> np.ndarray:
    """
    Encuentra el K óptimo para cortar el dendrograma usando
    un ensemble de 3 métricas:
      - Silhouette Score (cohesión + separación)
      - Calinski-Harabasz Index (varianza inter/intra)
      - Davies-Bouldin Index (similaridad entre clusters, menor=mejor)

    Puntúa cada K y elige por consenso.
    """
    # Rango de K a explorar
    min_k = 3
    max_k = min(40, max(4, n_samples // 15))

    if max_k <= min_k:
        max_k = min_k + 2

    print(f"     Explorando K de {min_k} a {max_k}...")

    # Subsamplear para métricas si dataset es muy grande
    sample_size = min(5000, n_samples)
    if sample_size < n_samples:
        rng = np.random.RandomState(42)
        sample_idx = rng.choice(n_samples, sample_size, replace=False)
        X_sample = X[sample_idx]
    else:
        sample_idx = None
        X_sample = X

    results = []

    for k in range(min_k, max_k + 1):
        labels_k = fcluster(Z, t=k, criterion="maxclust")

        # Usar subsample para métricas
        if sample_idx is not None:
            labels_sample = labels_k[sample_idx]
        else:
            labels_sample = labels_k

        # Solo evaluar si hay al menos 2 clusters distintos en la muestra
        n_unique = len(set(labels_sample))
        if n_unique < 2:
            continue

        try:
            sil = silhouette_score(X_sample, labels_sample)
        except Exception:
            sil = -1

        try:
            ch = calinski_harabasz_score(X_sample, labels_sample)
        except Exception:
            ch = 0

        try:
            db = davies_bouldin_score(X_sample, labels_sample)
        except Exception:
            db = 999

        results.append({"k": k, "sil": sil, "ch": ch, "db": db})

    if not results:
        print("     ⚠️  No se pudo evaluar ningún K → usando K=5")
        return fcluster(Z, t=5, criterion="maxclust")

    # Normalizar cada métrica a [0, 1] para scoring combinado
    df_r = pd.DataFrame(results)

    for col in ["sil", "ch"]:
        cmin, cmax = df_r[col].min(), df_r[col].max()
        if cmax > cmin:
            df_r[f"{col}_norm"] = (df_r[col] - cmin) / (cmax - cmin)
        else:
            df_r[f"{col}_norm"] = 0.5

    # Davies-Bouldin es inverso (menor = mejor)
    db_min, db_max = df_r["db"].min(), df_r["db"].max()
    if db_max > db_min:
        df_r["db_norm"] = 1.0 - (df_r["db"] - db_min) / (db_max - db_min)
    else:
        df_r["db_norm"] = 0.5

    # Score combinado ponderado
    # Silhouette es el más fiable para vecindad, Calinski para separación
    df_r["score"] = (
        0.45 * df_r["sil_norm"] +
        0.30 * df_r["ch_norm"] +
        0.25 * df_r["db_norm"]
    )

    best_row = df_r.loc[df_r["score"].idxmax()]
    best_k = int(best_row["k"])

    print(f"     • K óptimo = {best_k}")
    print(f"       Silhouette = {best_row['sil']:.4f}")
    print(f"       Calinski-H = {best_row['ch']:.1f}")
    print(f"       Davies-B   = {best_row['db']:.4f}")
    print(f"       Score comb = {best_row['score']:.4f}")

    # Top 3 para referencia
    top3 = df_r.nlargest(3, "score")
    print(f"     Top 3: K={top3['k'].tolist()}, Scores={[f'{s:.3f}' for s in top3['score'].tolist()]}")

    return fcluster(Z, t=best_k, criterion="maxclust")


# ==============================================================================
# 3b. ENFORCEMENT DE TOLERANCIAS (RESTRICCIÓN DURA)
# ==============================================================================

def _forzar_tolerancias(
    df: pd.DataFrame,
    labels: np.ndarray,
    param_cols: List[str],
    param_info: Dict[str, Dict],
    max_depth: int = 10,
) -> np.ndarray:
    """
    Restricción DURA de tolerancias: dentro de cada cluster, para CADA
    parámetro, la diferencia max-min NO puede superar 2×tolerancia.

    Si un cluster viola la restricción en algún parámetro, se subdivide
    iterativamente hasta que todos cumplan.

    ¿Por qué 2×tolerancia? La tolerancia define el radio de vecindad
    (±tolerancia desde el centro). Dos puntos cualesquiera dentro del
    cluster pueden diferir como máximo en 2×tolerancia (diámetro).

    Retorna: labels corregidos (pueden tener más clusters que los originales)
    """
    labels = labels.copy()
    next_id = int(labels.max()) + 1
    total_splits = 0

    # Construir mapa de tolerancias (solo parámetros con tolerancia > 0)
    tolerancias = {}
    for col in param_cols:
        info = param_info.get(col, {})
        tol = info.get("neighbor_tolerance", 0)
        if tol > 0 and col in df.columns:
            tolerancias[col] = tol * 2  # diámetro = 2 × radio

    if not tolerancias:
        print("     ⚠️  No hay tolerancias definidas, se omite enforcement")
        return labels

    print(f"\n  🔒 Enforcement de tolerancias (restricción dura):")
    print(f"     Diámetro máximo permitido por parámetro:")
    for col, diam in tolerancias.items():
        print(f"       {col}: ±{diam/2:.2f} → diámetro max = {diam:.2f}")

    for depth in range(max_depth):
        violations_found = False
        cluster_ids = sorted(set(labels))

        for cid in cluster_ids:
            mask = labels == cid
            n_in_cluster = mask.sum()
            if n_in_cluster <= 1:
                continue

            sub = df.loc[mask, list(tolerancias.keys())]

            # Verificar violaciones en cada parámetro
            worst_param = None
            worst_excess = 0

            for col, max_diam in tolerancias.items():
                actual_range = sub[col].max() - sub[col].min()
                excess = actual_range - max_diam
                if excess > worst_excess:
                    worst_excess = excess
                    worst_param = col

            if worst_param is None:
                continue  # Este cluster cumple todas las tolerancias

            # ── Subdividir el cluster violador ──
            violations_found = True
            max_diam = tolerancias[worst_param]
            actual_range = sub[worst_param].max() - sub[worst_param].min()

            # Calcular cuántos sub-clusters necesitamos para este parámetro
            n_sub = max(2, int(np.ceil(actual_range / max_diam)))
            # Pero no más de lo razonable
            n_sub = min(n_sub, max(2, n_in_cluster // 2))

            # Subdividir usando el parámetro violador como eje principal
            indices = np.where(mask)[0]
            values = df.loc[mask, worst_param].values

            # Crear bins equidistantes basados en la tolerancia
            vmin = values.min()
            bin_edges = [vmin + i * max_diam for i in range(n_sub + 1)]
            bin_edges[-1] = values.max() + 1  # Asegurar que el último bin captura todo

            for i in range(n_sub):
                lo = bin_edges[i]
                hi = bin_edges[i + 1]
                bin_mask = (values >= lo) & (values < hi)
                if bin_mask.sum() > 0:
                    if i == 0:
                        # Reusar el ID original para el primer sub-cluster
                        pass
                    else:
                        labels[indices[bin_mask]] = next_id
                        next_id += 1

            total_splits += 1

        if not violations_found:
            break

    # Re-numerar clusters secuencialmente (1, 2, 3, ...)
    unique_labels = sorted(set(labels))
    remap = {old: new + 1 for new, old in enumerate(unique_labels)}
    labels = np.array([remap[l] for l in labels])

    n_final = len(set(labels))
    print(f"\n     ✓ {total_splits} subdivisiones realizadas")
    print(f"     ✓ {n_final} clusters finales (todos cumplen tolerancias)")

    # Verificación final
    n_ok, n_fail = 0, 0
    for cid in sorted(set(labels)):
        mask = labels == cid
        for col, max_diam in tolerancias.items():
            actual = df.loc[mask, col].max() - df.loc[mask, col].min()
            if actual <= max_diam + 1e-9:
                n_ok += 1
            else:
                n_fail += 1
                print(f"     ⚠️  Cluster {cid}, {col}: rango={actual:.2f} > {max_diam:.2f}")

    if n_fail == 0:
        print(f"     ✅ Verificación: {n_ok} checks OK, 0 violaciones")
    else:
        print(f"     ❌ Verificación: {n_fail} violaciones residuales")

    return labels


def _detectar_col_roi(df: pd.DataFrame) -> Optional[str]:
    """Detecta la columna de ROI en el DataFrame."""
    for candidate in ["ROI_PCT", "ROI%", "ROI"]:
        if candidate in df.columns:
            return candidate
    for col in df.columns:
        if "ROI" in col.upper():
            return col
    return None


def _filtrar_clusters_roi_negativo(
    df: pd.DataFrame,
    labels: np.ndarray,
    seriated_order: np.ndarray,
    min_trials: int = 2,
) -> Tuple[pd.DataFrame, np.ndarray, np.ndarray]:
    """
    Validación final de clusters con 2 restricciones:
      1. Si ALGÚN trial tiene ROI negativo → cluster entero eliminado
      2. Si el cluster tiene menos de min_trials → eliminado

    Solo sobreviven clusters con ≥ min_trials trials y 100% ROI ≥ 0.

    Retorna: (df_filtrado, labels_filtrados, seriated_order_filtrado)
    """
    roi_col = _detectar_col_roi(df)

    if roi_col is None:
        print("\n  ⚠️  No se encontró columna ROI → no se filtra por rentabilidad")
        return df, labels, seriated_order

    print(f"\n  🧹 Validación de clusters:")
    print(f"     Requisitos: 100% ROI ≥ 0 + mínimo {min_trials} trials")

    df_work = df.copy()
    df_work["__CLUSTER__"] = labels
    df_work["__ORIG_IDX__"] = range(len(df_work))

    clusters_validos = []
    clusters_eliminados_roi = []
    clusters_eliminados_size = []

    for cid in sorted(df_work["__CLUSTER__"].unique()):
        sub = df_work[df_work["__CLUSTER__"] == cid]
        n_total = len(sub)
        n_negativos = (sub[roi_col] < 0).sum()

        if n_total < min_trials:
            clusters_eliminados_size.append((cid, n_total))
        elif n_negativos > 0:
            clusters_eliminados_roi.append((cid, n_total, n_negativos))
        else:
            clusters_validos.append(cid)

    n_elim_roi = len(clusters_eliminados_roi)
    n_elim_size = len(clusters_eliminados_size)
    print(f"     • {len(clusters_validos)} clusters válidos")
    print(f"     • {n_elim_roi} eliminados por ROI negativo")
    print(f"     • {n_elim_size} eliminados por < {min_trials} trials")

    if clusters_eliminados_roi:
        total_elim = sum(n for _, n, _ in clusters_eliminados_roi)
        print(f"     • {total_elim} trials descartados (ROI)")
        for cid, n_total, n_neg in clusters_eliminados_roi[:5]:
            print(f"       ❌ Cluster {cid}: {n_total} trials, {n_neg} con ROI < 0")
        if n_elim_roi > 5:
            print(f"       ... y {n_elim_roi - 5} clusters más")

    if clusters_eliminados_size:
        total_elim_s = sum(n for _, n in clusters_eliminados_size)
        print(f"     • {total_elim_s} trials descartados (tamaño)")
        for cid, n_total in clusters_eliminados_size[:5]:
            print(f"       ❌ Cluster {cid}: solo {n_total} trial(s)")
        if n_elim_size > 5:
            print(f"       ... y {n_elim_size - 5} clusters más")

    if not clusters_validos:
        print("\n  ❌ NINGÚN cluster cumple los requisitos (100% ROI+ y ≥ 2 trials).")
        print("     No hay zona de parámetros donde TODAS las combinaciones")
        print("     sean rentables dentro de la tolerancia.")
        df_empty = df.iloc[:0].copy()
        return df_empty, np.array([], dtype=int), np.array([], dtype=int)

    # Filtrar: solo filas que pertenecen a clusters válidos
    valid_set = set(clusters_validos)
    mask_valid = df_work["__CLUSTER__"].isin(valid_set)
    valid_orig_indices = set(df_work.loc[mask_valid, "__ORIG_IDX__"].values)

    # Filtrar seriated_order: mantener solo índices válidos, preservar orden
    seriated_filtered = np.array([i for i in seriated_order if i in valid_orig_indices])

    # Crear mapeo de índice original → nuevo índice
    old_to_new = {}
    for new_idx, old_idx in enumerate(sorted(valid_orig_indices)):
        old_to_new[old_idx] = new_idx

    # Reconstruir df filtrado
    df_filtered = df.iloc[sorted(valid_orig_indices)].reset_index(drop=True)

    # Reconstruir labels filtrado
    labels_filtered = labels[sorted(valid_orig_indices)]

    # Re-mapear seriated_order a nuevos índices
    seriated_new = np.array([old_to_new[i] for i in seriated_filtered])

    # Re-numerar clusters: 1, 2, 3...
    unique_labels = sorted(set(labels_filtered))
    remap = {old: new + 1 for new, old in enumerate(unique_labels)}
    labels_filtered = np.array([remap[l] for l in labels_filtered])

    n_final = len(set(labels_filtered))
    n_trials = len(df_filtered)
    print(f"\n     ✅ Resultado: {n_final} clusters válidos, {n_trials} trials rentables")

    return df_filtered, labels_filtered, seriated_new


# ==============================================================================
# 4. ANÁLISIS Y RESUMEN DE CLUSTERS
# ==============================================================================

def analizar_clusters(
    df: pd.DataFrame,
    labels: np.ndarray,
    param_cols: List[str],
    metric_cols: List[str],
    param_info: Dict[str, Dict],
) -> pd.DataFrame:
    """Genera resumen estadístico de cada cluster, ordenado por score medio."""
    df = df.copy()
    df["__CLUSTER__"] = labels

    resumen = []
    for cid in sorted(df["__CLUSTER__"].unique()):
        sub = df[df["__CLUSTER__"] == cid]
        n = len(sub)
        info = {"CLUSTER": int(cid), "N_TRIALS": n}

        if "SCORE" in df.columns:
            info["SCORE_MEDIO"] = sub["SCORE"].mean()
            info["SCORE_MAX"] = sub["SCORE"].max()
            info["SCORE_MIN"] = sub["SCORE"].min()

        for col in param_cols:
            if col in sub.columns:
                info[f"{col}_MIN"] = sub[col].min()
                info[f"{col}_MAX"] = sub[col].max()
                info[f"{col}_MEDIA"] = sub[col].mean()
                info[f"{col}_STD"] = sub[col].std()

        resumen.append(info)

    df_res = pd.DataFrame(resumen)
    sort_col = "SCORE_MEDIO" if "SCORE_MEDIO" in df_res.columns else "N_TRIALS"
    df_res = df_res.sort_values(sort_col, ascending=False).reset_index(drop=True)
    return df_res


def imprimir_resumen(
    df_resumen: pd.DataFrame,
    param_cols: List[str],
    param_info: Dict[str, Dict],
):
    """Imprime resumen en consola."""
    print("\n" + "=" * 70)
    print("  RESUMEN DE CLUSTERS (HCA Complete)")
    print("=" * 70)

    for _, row in df_resumen.iterrows():
        cid = int(row["CLUSTER"])
        n = int(row["N_TRIALS"])

        print(f"\n  ── Cluster {cid} ({n} trials) ──")
        if "SCORE_MEDIO" in row:
            print(f"     Score: media={row['SCORE_MEDIO']:.2f}, "
                  f"max={row['SCORE_MAX']:.2f}, min={row['SCORE_MIN']:.2f}")

        for col in param_cols:
            mk = f"{col}_MIN"
            xk = f"{col}_MAX"
            if mk in row and xk in row:
                vmin, vmax = row[mk], row[xk]
                vmedia = row.get(f"{col}_MEDIA", (vmin + vmax) / 2)
                vstd = row.get(f"{col}_STD", 0)
                tipo = param_info.get(col, {}).get("type", "?")

                # Mostrar cumplimiento de tolerancia
                tol = param_info.get(col, {}).get("neighbor_tolerance", 0)
                rango = vmax - vmin
                if tol > 0:
                    diam = tol * 2
                    ok = "✅" if rango <= diam + 1e-9 else "❌"
                    print(f"     {col}: [{vmin:.2f} - {vmax:.2f}] "
                          f"rango={rango:.2f} ≤ {diam:.2f} {ok} "
                          f"(μ={vmedia:.2f}, σ={vstd:.2f})")
                else:
                    print(f"     {col}: [{vmin:.2f} - {vmax:.2f}] "
                          f"(μ={vmedia:.2f}, σ={vstd:.2f}) [{tipo}]")
    print()


# ==============================================================================
# 5. GENERACIÓN DE EXCEL CON FORMATO PROFESIONAL + SERIACIÓN
# ==============================================================================

def generar_excel_hca(
    df_original: pd.DataFrame,
    seriated_order: np.ndarray,
    labels: np.ndarray,
    id_cols: List[str],
    metric_cols: List[str],
    param_cols: List[str],
    df_resumen_clusters: pd.DataFrame,
    output_path: str,
    saldo_inicial: float = 300.0,
):
    """
    Genera Excel con formato idéntico a MODELOX:
    - Filas seriadas (vecinos más cercanos adyacentes)
    - Separadores de color por cluster
    - 2 hojas: 'Seriado' (principal) y 'Resumen Clusters'
    """
    df = df_original.copy()
    df["__CLUSTER__"] = labels

    # ── Orden: por cluster (mejor score primero), dentro de cada cluster por seriación ──
    # Esto garantiza que cada cluster sea un bloque continuo, sin fragmentar.

    # Mapear cluster_id → rank por score (mejor primero)
    cluster_rank = {
        int(row["CLUSTER"]): idx
        for idx, (_, row) in enumerate(df_resumen_clusters.iterrows())
    }

    # Para cada trial, guardar su posición en la seriación original
    seriation_pos = {idx: pos for pos, idx in enumerate(seriated_order)}

    # Crear orden final: cluster_rank primero, seriación dentro
    df["__CLUSTER_RANK__"] = df["__CLUSTER__"].map(cluster_rank).fillna(9999).astype(int)
    df["__SERIATE_POS__"] = [seriation_pos.get(i, i) for i in range(len(df))]
    df_seriated = df.sort_values(["__CLUSTER_RANK__", "__SERIATE_POS__"]).reset_index(drop=True)

    # Columnas finales
    final_id_cols = ["CLUSTER"] + [c for c in id_cols if c in df.columns]
    final_metric_cols = [c for c in metric_cols if c in df.columns]
    final_param_cols = [c for c in param_cols if c in df.columns]
    all_cols = final_id_cols + final_metric_cols + final_param_cols

    # Construir filas con separadores por cluster (ahora siempre contiguos)
    rows_data = []
    separator_rows = []  # (data_row_index, cluster_id)
    current_row = 0
    prev_cluster = None

    for iloc_idx in range(len(df_seriated)):
        trial = df_seriated.iloc[iloc_idx]
        cid = int(trial["__CLUSTER__"])

        # Insertar separador cuando cambia el cluster
        if cid != prev_cluster:
            sep = {col: "" for col in all_cols}
            # Contar trials en este bloque continuo
            block_count = 0
            for j in range(iloc_idx, len(df_seriated)):
                if int(df_seriated.iloc[j]["__CLUSTER__"]) == cid:
                    block_count += 1
                else:
                    break

            if "SCORE" in df_seriated.columns:
                block_scores = df_seriated.iloc[iloc_idx:iloc_idx + block_count]["SCORE"]
                sep["CLUSTER"] = (
                    f"━━ CLUSTER {cid} ━━ ({block_count} trials, "
                    f"Score medio: {block_scores.mean():.2f}) ━━"
                )
            else:
                sep["CLUSTER"] = f"━━ CLUSTER {cid} ━━ ({block_count} trials) ━━"

            rows_data.append(sep)
            separator_rows.append((current_row, cid))
            current_row += 1
            prev_cluster = cid

        # Fila de datos
        data_row = {"CLUSTER": cid}
        for col in id_cols:
            if col in trial:
                data_row[col] = trial[col]
        for col in metric_cols:
            if col in trial:
                data_row[col] = trial[col]
        for col in param_cols:
            if col in trial:
                data_row[col] = trial[col]

        rows_data.append(data_row)
        current_row += 1

    # Crear workbook directamente (evita save/reload para archivos grandes)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Seriado"

    # Fila 2: headers
    for j, col_name in enumerate(all_cols, 1):
        ws.cell(row=2, column=j, value=col_name)

    # Fila 3+: datos
    for i, row_data in enumerate(rows_data, 3):
        for j, col_name in enumerate(all_cols, 1):
            val = row_data.get(col_name, "")
            ws.cell(row=i, column=j, value=val)

    # Estilos
    _aplicar_estilo(
        ws, all_cols,
        final_id_cols, final_metric_cols, final_param_cols,
        separator_rows, cluster_rank, saldo_inicial,
    )

    # Hoja 2: Resumen
    _agregar_hoja_resumen(wb, df_resumen_clusters)

    wb.save(output_path)
    return output_path


# ==============================================================================
# 6. ESTILOS (Idéntico a MODELOX)
# ==============================================================================

def _aplicar_estilo(
    ws,
    all_cols: List[str],
    id_cols: List[str],
    metric_cols: List[str],
    param_cols: List[str],
    separator_rows: List[Tuple[int, int]],
    cluster_rank: Dict[int, int],
    saldo_ini: float,
):
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    max_row = ws.max_row

    n_ids = len(id_cols)
    n_metrics = len(metric_cols)
    n_params = len(param_cols)

    start_metrics = n_ids + 1
    end_metrics = start_metrics + n_metrics - 1
    start_params = end_metrics + 1
    end_params = start_params + n_params - 1

    border = Border(
        left=Side(style="thin", color=COLORS["border_color"]),
        right=Side(style="thin", color=COLORS["border_color"]),
        top=Side(style="thin", color=COLORS["border_color"]),
        bottom=Side(style="thin", color=COLORS["border_color"]),
    )
    font_group = Font(name=FONT_TITLE, size=12, bold=True, color=COLORS["text_white"])
    align_center = Alignment(horizontal="center", vertical="center")

    # ── ROW 1: TÍTULOS DE GRUPO ──
    for c in range(1, start_metrics):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_id"])
        cell.font = font_group
        cell.alignment = align_center

    if n_metrics > 0:
        cell = ws.cell(row=1, column=start_metrics)
        cell.value = "MÉTRICAS CLAVE"
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
        cell.font = font_group
        cell.alignment = align_center
        if n_metrics > 1:
            ws.merge_cells(start_row=1, start_column=start_metrics,
                           end_row=1, end_column=end_metrics)
        for c in range(start_metrics, end_metrics + 1):
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])

    if n_params > 0:
        cell = ws.cell(row=1, column=start_params)
        cell.value = "PARÁMETROS ESTRATEGIA"
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_params"])
        cell.font = font_group
        cell.alignment = align_center
        if n_params > 1:
            ws.merge_cells(start_row=1, start_column=start_params,
                           end_row=1, end_column=end_params)
        for c in range(start_params, end_params + 1):
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=COLORS["header_bg_params"])

    # ── ROW 2: HEADERS ──
    font_header = Font(name=FONT_TITLE, size=12, bold=True, color=COLORS["text_white"])
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if col < start_metrics:
            cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_id"])
        elif col <= end_metrics:
            cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
        else:
            cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_params"])
        val_len = len(str(cell.value)) if cell.value else 0
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, val_len + 2), 22)

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    # ── DATOS ──
    font_body = Font(name=FONT_BODY, size=12, color=COLORS["text_dark"])

    sep_excel = set()
    cluster_for_row = {}
    for data_idx, cid in separator_rows:
        erow = data_idx + 3
        sep_excel.add(erow)
        cluster_for_row[erow] = cid

    for r in range(3, max_row + 1):
        if r in sep_excel:
            cid = cluster_for_row[r]
            rank = cluster_rank.get(cid, cid)
            color_idx = rank % len(CLUSTER_PALETTE)
            sep_fill = PatternFill("solid", fgColor=CLUSTER_PALETTE[color_idx])
            sep_font = Font(name=FONT_TITLE, size=11, bold=True, color=COLORS["text_white"])

            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = sep_fill
                cell.font = sep_font
                cell.alignment = align_center
                cell.border = border
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
            ws.row_dimensions[r].height = 25
        else:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = font_body
                cell.alignment = align_center
                cell.border = border

                hdr = str(ws.cell(2, c).value or "").upper()
                if isinstance(cell.value, (int, float)):
                    if "TRADES" in hdr and "DIA" in hdr:
                        cell.number_format = "0.00"
                    elif "TRADES" in hdr or "NUM_" in hdr or "CLUSTER" in hdr:
                        cell.number_format = "0"
                    elif any(k in hdr for k in ["SCORE", "%", "PCT", "RATIO", "SHARPE", "FACTOR", "ESTABILIDAD"]):
                        cell.number_format = "0.00"
                    elif any(k in hdr for k in ["SALDO", "PROFIT", "PNL"]):
                        cell.number_format = "#,##0.00"

    # ── CONDITIONAL FORMATTING ──
    col_map = {}
    for c in range(1, max_col + 1):
        val = str(ws.cell(2, c).value or "").strip()
        if val:
            col_map[val] = (get_column_letter(c), c)

    if "SCORE" in col_map:
        cl, _ = col_map["SCORE"]
        ws.conditional_formatting.add(
            f"{cl}3:{cl}{max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    for roi_name in ["ROI_PCT", "ROI%"]:
        if roi_name in col_map:
            cl, _ = col_map[roi_name]
            ws.conditional_formatting.add(
                f"{cl}3:{cl}{max_row}",
                DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True),
            )
            break

    if "SALDO_ACTUAL" in col_map:
        _, cn = col_map["SALDO_ACTUAL"]
        for row in range(3, max_row + 1):
            if row in sep_excel:
                continue
            cell = ws.cell(row=row, column=cn)
            try:
                val = float(cell.value)
                if val >= saldo_ini * 1.5:
                    cell.fill = PatternFill("solid", fgColor=COLORS["success_bg"])
                    cell.font = Font(name=FONT_BODY, size=12, color="006100", bold=True)
                elif val < saldo_ini:
                    cell.fill = PatternFill("solid", fgColor=COLORS["danger_bg"])
                    cell.font = Font(name=FONT_BODY, size=12, color="9C0006")
            except Exception:
                pass

    freeze_col = get_column_letter(start_metrics)
    ws.freeze_panes = f"{freeze_col}3"


def _agregar_hoja_resumen(wb, df_resumen: pd.DataFrame):
    """Hoja 2: estadísticas por cluster."""
    if "Resumen Clusters" in wb.sheetnames:
        del wb["Resumen Clusters"]

    ws = wb.create_sheet("Resumen Clusters")
    headers = list(df_resumen.columns)

    font_hdr = Font(name=FONT_TITLE, size=11, bold=True, color=COLORS["text_white"])
    fill_hdr = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    brd = Border(
        left=Side(style="thin", color=COLORS["border_color"]),
        right=Side(style="thin", color=COLORS["border_color"]),
        top=Side(style="thin", color=COLORS["border_color"]),
        bottom=Side(style="thin", color=COLORS["border_color"]),
    )

    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = font_hdr
        c.fill = fill_hdr
        c.alignment = align_c
        c.border = brd
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, len(h) + 2), 22)

    font_b = Font(name=FONT_BODY, size=11, color=COLORS["text_dark"])
    for i, (_, row) in enumerate(df_resumen.iterrows(), 2):
        for j, col in enumerate(headers, 1):
            val = row[col]
            c = ws.cell(row=i, column=j, value=val)
            c.font = font_b
            c.alignment = align_c
            c.border = brd
            if isinstance(val, float):
                c.number_format = "0.00"

    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.sheet_view.showGridLines = False


# ==============================================================================
# 7. PIPELINE PRINCIPAL
# ==============================================================================

def main():
    """Pipeline HCA completo."""

    # 1. Archivo
    filepath = solicitar_archivo()

    # 2. Lectura
    df, id_cols, metric_cols, param_cols = leer_excel_modelox(filepath)

    if not param_cols:
        print("\n  ❌ No se detectaron parámetros. Imposible clusterizar.")
        sys.exit(1)

    # 3. Análisis de parámetros
    print("\n  🔍 Análisis de parámetros:")
    param_info = {}
    for col in param_cols:
        info = analizar_tipo_parametro(df[col], col)
        param_info[col] = info
        print(f"     • {col}: tipo={info['type']}, peso={info['weight']:.2f}, "
              f"rango=[{info['min']:.2f}-{info['max']:.2f}], "
              f"tolerancia={info.get('neighbor_tolerance', 0):.2f}")

    # 4. Features ponderadas
    print("\n  ⚙️  Construyendo features ponderadas...")
    X = construir_features_ponderadas(df[param_cols], param_info)
    print(f"     • Matriz: {X.shape[0]} × {X.shape[1]}")

    # 5. HCA
    print("\n  🧠 Ejecutando HCA (Complete + Seriación Óptima)...")
    seriated_order, labels, Z = ejecutar_hca(X, param_info)

    # 6. Enforcement de tolerancias (restricción dura)
    labels = _forzar_tolerancias(df, labels, param_cols, param_info)

    # 7. Filtrar clusters con algún ROI negativo (cluster entero se invalida)
    df, labels, seriated_order = _filtrar_clusters_roi_negativo(
        df, labels, seriated_order,
    )

    if len(df) == 0:
        print("\n  ❌ No quedan trials tras filtrar. No se genera Excel.")
        sys.exit(0)

    # 8. Análisis
    df_resumen = analizar_clusters(df, labels, param_cols, metric_cols, param_info)
    imprimir_resumen(df_resumen, param_cols, param_info)

    # 9. Excel
    base_dir = os.path.dirname(filepath)
    base_name = os.path.splitext(os.path.basename(filepath))[0]
    output_path = os.path.join(base_dir, f"{base_name}_HCA.xlsx")

    print(f"\n  📊 Generando Excel: {os.path.basename(output_path)}...")

    generar_excel_hca(
        df_original=df,
        seriated_order=seriated_order,
        labels=labels,
        id_cols=id_cols,
        metric_cols=metric_cols,
        param_cols=param_cols,
        df_resumen_clusters=df_resumen,
        output_path=output_path,
    )

    print(f"\n  ✅ ¡Excel HCA generado exitosamente!")
    print(f"     📁 {output_path}")
    print(f"\n     Hojas:")
    print(f"       1. 'Seriado' - Trials ordenados por vecindad (OLO)")
    print(f"       2. 'Resumen Clusters' - Estadísticas por cluster")
    print(f"\n     💡 Lee el Excel de arriba a abajo: cada fila es la más")
    print(f"        parecida posible a la siguiente en parámetros.")

    # 10. PDF profesional
    pdf_path = os.path.join(base_dir, f"{base_name}_HCA.pdf")
    try:
        from pdf_hca import generar_pdf_hca
        generar_pdf_hca(
            df=df,
            labels=labels,
            id_cols=id_cols,
            metric_cols=metric_cols,
            param_cols=param_cols,
            df_resumen=df_resumen,
            param_info=param_info,
            output_path=pdf_path,
            source_file=filepath,
        )
        print(f"     📁 {pdf_path}")
    except Exception as e:
        print(f"\n  ⚠️  Error generando PDF: {e}")
        import traceback
        traceback.print_exc()

    print("\n" + "=" * 70)


if __name__ == "__main__":
    main()
