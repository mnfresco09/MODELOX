from __future__ import annotations

# ============================================================================
# [CRÍTICO] CONFIGURACIÓN DE ALTO RENDIMIENTO PARA VM (CLOUD)
# ============================================================================
import os
import sys

# CONFIGURACIÓN DE HILOS: FORZAMOS A 1 HILO POR PROCESO.
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"
os.environ["POLARS_MAX_THREADS"] = "1"

# VARIABLES DE DEPURACIÓN
os.environ.setdefault("MODELOX_TIMINGS", "1")
os.environ.setdefault("MODELOX_TIMINGS_PRINT_EVERY", "10")

# ============================================================================
# IMPORTS DEL SISTEMA
# ============================================================================
import csv
import warnings
import logging
import atexit
import shutil
from copy import deepcopy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Protocol, Tuple
import re

# CAPTURA SEGURA DE LA RUTA RAÍZ AL INICIO (EVITA ERROR EN ATEXIT)
try:
    _PROJECT_ROOT = Path(__file__).resolve().parent
except NameError:
    # Fallback si __file__ no está definido (ej: shell interactiva)
    _PROJECT_ROOT = Path.cwd()

# SILENCIAR ADVERTENCIAS
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

# ============================================================================
# LOGGING SETUP
# ============================================================================
def setup_logging(level: int = logging.INFO) -> None:
    logging.basicConfig(
        level=level,
        format="[%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True,
    )
    logging.getLogger("matplotlib").setLevel(logging.WARNING)
    logging.getLogger("PIL").setLevel(logging.WARNING)
    try:
        import optuna
        optuna.logging.set_verbosity(optuna.logging.WARNING)
        logging.getLogger("optuna").setLevel(logging.WARNING)
    except ImportError:
        pass

setup_logging(level=logging.WARNING)
logger = logging.getLogger(__name__)

# ============================================================================
# IMPORTS DEL PROYECTO
# ============================================================================
from general.configuracion import (
    ACTIVOS, ACTIVO_PRIMARIO, resolve_archivo_data, resolve_archivo_data_tf,
    resolve_qty_max_activo, resolve_qty_max_activo_range,
    COMBINACION_A_EJECUTAR, CONFIG,
    FECHA_FIN, FECHA_INICIO, N_TRIALS, FECHA_FIN_PLOT, FECHA_INICIO_PLOT,
    GENERAR_PLOTS, MAX_ARCHIVOS_GUARDAR, USAR_EXCEL,
    PERTURBACION_ACTIVAR,
    # Neighborhood Fitness Aggregation
    VECINDARIO_ACTIVAR, VECINDARIO_N_NEIGHBORS, VECINDARIO_PERTURBATION_STD,
    VECINDARIO_LAMBDA_PENALTY, VECINDARIO_SEED, VECINDARIO_EXCEL, VECINDARIO_GUARDAR_MEJORES,
    VECINDARIO_MAX_DISPERSION,
    # Limpieza periódica
    CLEANUP_INTERVAL,
)
from modelox.core.data import load_data
from modelox.core.runner import OptimizationRunner, OptunaConfig
# v7.0: NeighborhoodConfig ahora está en scoring.py unificado
from modelox.core.scoring import NeighborhoodConfig
from modelox.core.types import normalize_timeframe_to_suffix, BacktestConfig, filter_by_date, nuclear_cleanup, full_system_cleanup, TrialArtifacts
from modelox.strategies.registry import instantiate_strategies
from visual.excel import exportar_trades_excel_rapido, convertir_resumen_csv_a_excel
from visual.grafico import plot_trades
from visual.rich import (
    mostrar_cabecera_inicio, mostrar_fin_optimizacion,
    mostrar_panel_elegante, mostrar_top_trials, actualizar_estadisticas,
    mostrar_evolucion_metricas,
    resetear_estadisticas, mostrar_resultado_vecindario,
)

# Configurar intervalo de limpieza desde configuracion.py
os.environ["MODELOX_CLEANUP_INTERVAL"] = str(CLEANUP_INTERVAL)

# ============================================================================
# REPORTERS INTEGRADOS (fusionados de modelox/reporting/)
# ============================================================================

class BaseReporter(Protocol):
    """
    Protocolo base para todos los reporters.
    """

    def needs_dataframe(self, score: float) -> bool:
        """Indica si este reporter necesita df_signals convertido a Pandas."""
        ...

    def on_trial_end(self, artifacts: TrialArtifacts) -> None:
        """Procesa resultados de un trial completado."""
        ...

    def on_strategy_end(self, strategy_name: str, study: Any) -> None:
        """Procesa resultados finales de una estrategia."""
        ...


@dataclass
class ExcelReporter(BaseReporter):
    """
    Excel exporter wrapper - ULTRA OPTIMIZADO (v3.1).
    """

    resumen_path: str = "resultados/excel/resumen.xlsx"
    trades_base_dir: str = "resultados/excel"
    max_archivos: int = 5
    use_fast_mode: bool = True
    
    _csv_resumen_path: Optional[str] = field(default=None, init=False, repr=False)
    _resumen_rows: List[Dict[str, Any]] = field(default_factory=list, init=False, repr=False)
    _trade_candidates: List[Dict[str, Any]] = field(default_factory=list, init=False, repr=False)
    _min_candidate_score: float = field(default=float("-inf"), init=False, repr=False)
    _activo: Optional[str] = field(default=None, init=False, repr=False)

    def needs_dataframe(self, score: float) -> bool:
        return False

    @staticmethod
    def _safe_activo_name(activo: str) -> str:
        return str(activo).strip().replace(" ", "_").upper() if activo else "DEFAULT"

    def _excel_dir_for(self, activo: str) -> str:
        # Ya no creamos subcarpeta por activo, usamos trades_base_dir directamente
        return self.trades_base_dir

    def _update_min_score(self):
        if self._trade_candidates:
            self._min_candidate_score = min(c["score"] for c in self._trade_candidates)
        else:
            self._min_candidate_score = float("-inf")

    def on_trial_end(self, artifacts: TrialArtifacts) -> None:
        params_src = getattr(artifacts, "params_reporting", None) or artifacts.params
        activo = None
        if isinstance(params_src, dict):
            activo = params_src.get("__activo") or params_src.get("ACTIVO") or params_src.get("activo")
        
        self._activo = activo
        score = artifacts.score if artifacts.score is not None else 0.0
        
        params = dict(params_src)
        params["NOMBRE_COMBO"] = artifacts.strategy_name
        
        resumen_row = {
            "trial_number": artifacts.trial_number,
            "score": score,
            "metrics": deepcopy(artifacts.metrics) if artifacts.metrics else {},
            "params": {k: v for k, v in params.items() if not str(k).startswith("__")},
            "perturbado": artifacts.perturbado,
            "perturb_seed": artifacts.perturb_seed,
            "strategy_name": artifacts.strategy_name,
        }
        self._resumen_rows.append(resumen_row)

        try:
            base_dir = self.trades_base_dir
            os.makedirs(base_dir, exist_ok=True)
            if not self._csv_resumen_path:
                self._csv_resumen_path = os.path.join(base_dir, "RESUMEN.csv")
            self._write_resumen_csv(self._csv_resumen_path)
        except Exception:
            pass
        
        is_candidate = (
            len(self._trade_candidates) < self.max_archivos or 
            score > self._min_candidate_score
        )
        
        if is_candidate and artifacts.trades is not None:
            candidate = {
                "score": score,
                "trial_number": artifacts.trial_number,
                "trades": artifacts.trades,
                "params": params,
                "metrics": artifacts.metrics,
                "perturbado": artifacts.perturbado,
                "perturb_seed": artifacts.perturb_seed,
            }
            self._trade_candidates.append(candidate)
            
            if len(self._trade_candidates) > self.max_archivos:
                self._trade_candidates.sort(key=lambda x: x["score"], reverse=True)
                removed = self._trade_candidates.pop()
                del removed
            
            self._update_min_score()

    def on_strategy_end(self, strategy_name: str, study) -> None:
        if not self._resumen_rows:
            return
        
        activo = self._activo
        base_dir = self.trades_base_dir
        os.makedirs(base_dir, exist_ok=True)
        
        activo_safe = self._safe_activo_name(str(activo) if activo else "DEFAULT")
        csv_path = os.path.join(base_dir, "RESUMEN.csv")
        
        self._write_resumen_csv(csv_path)
        
        self._trade_candidates.sort(key=lambda x: x["score"], reverse=True)
        
        # Guardar trades directamente en base_dir, sin subcarpeta trades
        for candidate in self._trade_candidates[:self.max_archivos]:
            try:
                self._write_trades_excel(base_dir, candidate)
            except Exception as e:
                logger.warning(f"Error guardando trades trial {candidate['trial_number']}: {e}")
        
        try:
            convertir_resumen_csv_a_excel(
                csv_path=csv_path,
                strategy_name=strategy_name,
                activo=activo_safe,
                output_dir=base_dir
            )
        except Exception as e:
            logger.warning(f"Error generando Dashboard Excel: {e}")
        
        self._resumen_rows = []
        self._trade_candidates = []
        self._min_candidate_score = float("-inf")

    def _write_resumen_csv(self, csv_path: str):
        if not self._resumen_rows:
            return
        
        all_keys = set()
        for row in self._resumen_rows:
            all_keys.add("trial")
            all_keys.add("score")
            all_keys.add("strategy")
            if row.get("metrics"):
                all_keys.update(row["metrics"].keys())
            if row.get("params"):
                all_keys.update(f"param_{k}" for k in row["params"].keys())
        
        columns = ["trial", "score", "strategy"]
        metric_cols = sorted([k for k in all_keys if k not in columns and not k.startswith("param_")])
        param_cols = sorted([k for k in all_keys if k.startswith("param_")])
        columns.extend(metric_cols)
        columns.extend(param_cols)
        
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            
            for row in self._resumen_rows:
                csv_row = {
                    "trial": row["trial_number"],
                    "score": row["score"],
                    "strategy": row["strategy_name"],
                }
                if row.get("metrics"):
                    csv_row.update(row["metrics"])
                if row.get("params"):
                    csv_row.update({f"param_{k}": v for k, v in row["params"].items()})
                
                writer.writerow(csv_row)

    def _write_trades_excel(self, trades_dir: str, candidate: Dict[str, Any]):
        import pandas as pd
        
        trades = candidate["trades"]
        if trades is None or (hasattr(trades, "empty") and trades.empty):
            return
        
        if hasattr(trades, "to_pandas"):
            df_trades = trades.to_pandas()
        else:
            df_trades = trades

        try:
            df_trades = df_trades.copy()
            if isinstance(df_trades.index, pd.DatetimeIndex) and df_trades.index.tz is not None:
                df_trades.index = df_trades.index.tz_localize(None)
            for col in df_trades.columns:
                try:
                    if isinstance(df_trades[col].dtype, pd.DatetimeTZDtype):
                        df_trades[col] = df_trades[col].dt.tz_localize(None)
                except Exception:
                    continue
            for col in df_trades.columns:
                if df_trades[col].dtype != object:
                    continue
                s = df_trades[col]
                try:
                    converted = pd.to_datetime(s, errors="ignore", utc=True)
                    if isinstance(converted.dtype, pd.DatetimeTZDtype):
                        df_trades[col] = converted.dt.tz_localize(None)
                        continue
                except Exception:
                    pass
                try:
                    sample = next((v for v in s.head(50).tolist() if v is not None), None)
                    if sample is None:
                        continue
                    tzinfo = getattr(sample, "tzinfo", None)
                    if tzinfo is None:
                        continue
                    df_trades[col] = s.apply(
                        lambda v: v.replace(tzinfo=None)
                        if hasattr(v, "tzinfo") and v.tzinfo is not None else v
                    )
                except Exception:
                    continue
        except Exception:
            pass
        
        score_str = f"{candidate['score']:.2f}".replace(".", "_")
        filename = f"TRADES_TRIAL{candidate['trial_number']}_SCORE{score_str}.xlsx"
        filepath = os.path.join(trades_dir, filename)
        
        df_trades.to_excel(filepath, index=False, sheet_name="Trades")


@dataclass
class PlotReporter(BaseReporter):
    """Lightweight Charts (TradingView) HTML exporter - OPTIMIZADO."""

    plot_base: str = "resultados/graficos"
    fecha_inicio_plot: str = "2025-01-01"
    fecha_fin_plot: str = "2025-01-20"
    max_archivos: int = 5
    saldo_inicial: float = 300.0
    activo: Optional[str] = None
    
    _candidates: List[Dict[str, Any]] = field(default_factory=list, init=False, repr=False)
    _min_candidate_score: float = field(default=float("-inf"), init=False, repr=False)

    def needs_dataframe(self, score: float) -> bool:
        if score is None:
            return False
        if len(self._candidates) < self.max_archivos:
            return True
        return score > self._min_candidate_score

    def _update_min_score(self):
        if self._candidates:
            self._min_candidate_score = min(c["score"] for c in self._candidates)
        else:
            self._min_candidate_score = float("-inf")

    def on_trial_end(self, artifacts: TrialArtifacts) -> None:
        if artifacts.trial_number == 0 and os.path.exists(self.plot_base):
            try:
                for f in os.listdir(self.plot_base):
                    if f.startswith("TRIAL-") and f.endswith(".html"):
                        os.remove(os.path.join(self.plot_base, f))
            except Exception:
                pass
            self._candidates = []
            self._min_candidate_score = float("-inf")

        score = artifacts.score
        if score is None:
            return

        is_candidate = (
            len(self._candidates) < self.max_archivos or 
            score > self._min_candidate_score
        )
        
        if not is_candidate:
            return
        
        if getattr(artifacts, "df_signals", None) is None:
            return

        params_for_plot = getattr(artifacts, "params_reporting", None) or artifacts.params
        
        candidate = {
            "score": score,
            "trial_number": artifacts.trial_number,
            "strategy_name": artifacts.strategy_name,
            "params": deepcopy(params_for_plot),
            "metrics": deepcopy(artifacts.metrics) if artifacts.metrics else {},
            "equity_curve": list(artifacts.equity_curve) if artifacts.equity_curve else [],
            "df_signals": artifacts.df_signals,
            "trades": artifacts.trades,
        }
        
        self._candidates.append(candidate)
        
        if len(self._candidates) > self.max_archivos:
            self._candidates.sort(key=lambda x: x["score"], reverse=True)
            removed = self._candidates.pop()
            del removed
        
        self._update_min_score()

    def on_strategy_end(self, strategy_name: str, study) -> None:
        if not self._candidates:
            return
        
        os.makedirs(self.plot_base, exist_ok=True)
        
        self._candidates.sort(key=lambda x: x["score"], reverse=True)
        
        for candidate in self._candidates[:self.max_archivos]:
            try:
                plot_trades(
                    df=candidate["df_signals"],
                    df_trades=candidate["trades"],
                    plot_base=self.plot_base,
                    fecha_inicio_plot=self.fecha_inicio_plot,
                    fecha_fin_plot=self.fecha_fin_plot,
                    trial_number=candidate["trial_number"],
                    params=candidate["params"],
                    score=candidate["score"],
                    combo=candidate["strategy_name"],
                    metrics=candidate["metrics"],
                    equity_curve=candidate["equity_curve"],
                    saldo_inicial=self.saldo_inicial,
                    max_archivos=self.max_archivos,
                    activo=self.activo,
                )
            except Exception as e:
                logger.warning(f"Error generando plot para trial {candidate['trial_number']}: {e}")
        
        self._candidates = []
        self._min_candidate_score = float("-inf")


@dataclass
class ElegantRichReporter(BaseReporter):
    """Bloomberg/TradingView-style Rich console reporter."""

    saldo_inicial: float = 300.0
    activo: str = ""
    mostrar_evolucion_cada: int = 50
    mostrar_evolucion_inline: bool = False  # Disabled - averages now shown in panel
    _best_score: float = field(default=float("-inf"), init=False, repr=False)
    _initialized: bool = field(default=False, init=False, repr=False)

    def needs_dataframe(self, score: float) -> bool:
        return False

    def on_trial_end(self, artifacts: TrialArtifacts) -> None:
        if not self._initialized:
            resetear_estadisticas()
            self._initialized = True

        current_score = artifacts.score or 0
        best_so_far = None

        if current_score > self._best_score:
            self._best_score = current_score
        best_so_far = self._best_score if self._best_score > float("-inf") else None

        indicadores = list(getattr(artifacts, "indicators_used", []))
        if not indicadores and artifacts.params:
            raw = artifacts.params.get("__indicators_used", [])
            if isinstance(raw, (list, tuple)):
                indicadores = [str(x) for x in raw if x]

        metrics = dict(artifacts.metrics or {})
        try:
            eq = getattr(artifacts, "equity_curve", None)
            if isinstance(eq, (list, tuple)) and len(eq) > 0:
                if "saldo_max" not in metrics or metrics.get("saldo_max") in (None, 0, 0.0):
                    metrics["saldo_max"] = float(max(eq))
        except Exception:
            pass

        actualizar_estadisticas(metrics, current_score, artifacts.trial_number)

        tf_entry = ""
        tf_exit = ""
        if artifacts.params:
            tf_entry = str(artifacts.params.get("__timeframe_entry", ""))
            tf_exit = str(artifacts.params.get("__timeframe_exit", ""))

        params_for_reporting = artifacts.params.copy() if artifacts.params else {}
        if '__qty_max_activo' in params_for_reporting:
            params_for_reporting['cantidad'] = params_for_reporting['__qty_max_activo']

        mostrar_panel_elegante(
            metrics=metrics,
            params=params_for_reporting,
            score=artifacts.score or 0,
            trial_num=artifacts.trial_number,
            saldo_inicial=self.saldo_inicial,
            indicadores_activos=indicadores,
            combo_str=artifacts.strategy_name or "",
            activo=self.activo,
            best_so_far=best_so_far,
            timeframe_entry=tf_entry,
            timeframe_exit=tf_exit,
        )

        # Mostrar resultado de vecindario si está disponible
        neighborhood_result = getattr(artifacts, "neighborhood_result", None)
        if neighborhood_result:
            mostrar_resultado_vecindario(neighborhood_result, inline=True)

    def on_strategy_end(self, strategy_name: str, study) -> None:
        self._initialized = False
        if study and hasattr(study, 'trials') and study.trials:
            mostrar_evolucion_metricas(forzar=True)  # SUMMARY solo al final
            mostrar_top_trials(study, n=5)


@dataclass
class NeighborhoodReporter(BaseReporter):
    """
    Genera Excel con resultados detallados del análisis de vecindario.
    
    SOLO guarda trials APROBADOS (robustez verificada).
    Escribe Excel INMEDIATAMENTE durante el trial (no al final).
    
    Formato similar al RESUMEN.xlsx de trials:
    - Trial original + todos los vecinos (subtrials)
    - Métricas completas de cada uno
    - Parámetros usados en cada subtrial
    """
    
    output_dir: str = "resultados/vecindario"
    max_guardar: int = 5
    activo: str = ""
    timeframe: str = "1m"
    
    def __post_init__(self):
        os.makedirs(self.output_dir, exist_ok=True)
        self._run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Inicializar lista de archivos guardados (importante: lista de instancia, no de clase)
        self._saved_files: List[Tuple[float, str]] = []
    
    def needs_dataframe(self, score: float) -> bool:
        return False
    
    def on_trial_end(self, artifacts: TrialArtifacts) -> None:
        """
        Procesa trial: si está APROBADO, escribe Excel inmediatamente.
        """
        neighborhood_result = getattr(artifacts, "neighborhood_result", None)
        if not neighborhood_result:
            return
        
        if hasattr(neighborhood_result, "to_dict"):
            neigh_dict = neighborhood_result.to_dict()
        else:
            neigh_dict = dict(neighborhood_result)
        
        # =====================================================================
        # VERIFICAR SI ESTÁ APROBADO
        # =====================================================================
        n_tested = neigh_dict.get("n_neighbors_tested", 0)
        
        # Códigos especiales: -1 = skip por mal ROI, -2 = skip por baja frecuencia
        if n_tested <= 0:
            return  # No se hizo test o suspenso por requisitos
        
        # Verificar aprobación de robustez
        aprobado = neigh_dict.get("robustness_approved", False)
        avg_dispersion = neigh_dict.get("avg_dispersion", 1.0)
        stability_score = max(0.0, 1.0 - avg_dispersion)  # Inverso de dispersión
        
        if not aprobado:
            return  # Test hecho pero no aprobado
        
        score = artifacts.score if artifacts.score is not None else 0.0
        
        # =====================================================================
        # APROBADO: Escribir Excel INMEDIATAMENTE
        # =====================================================================
        record = {
            "trial": artifacts.trial_number,
            "strategy": artifacts.strategy_name,
            "score": score,
            "stability_score": stability_score,
            "metrics": deepcopy(artifacts.metrics or {}),
            "params": deepcopy(artifacts.params) if artifacts.params else {},
            "neighborhood_info": deepcopy(neigh_dict),
        }
        
        # Escribir archivo Excel
        filepath = self._write_trial_neighborhood_excel(record)
        
        if filepath:
            # Añadir a tracking
            self._saved_files.append((score, filepath))
            
            # Mantener solo TOP N (eliminar los peores si hay más)
            while len(self._saved_files) > self.max_guardar:
                self._saved_files.sort(key=lambda x: x[0], reverse=True)  # Mayor score primero
                worst_score, worst_file = self._saved_files.pop()  # Eliminar el peor
                try:
                    if os.path.exists(worst_file):
                        os.remove(worst_file)
                        print(f"    [ROBUSTEZ] Eliminado archivo con score {worst_score:.0f}: {Path(worst_file).name}")
                except Exception as e:
                    print(f"    [ROBUSTEZ] Error eliminando {worst_file}: {e}")
    
    def on_strategy_end(self, strategy_name: str, study) -> None:
        """Reset al finalizar estrategia."""
        self._saved_files = []
    
    def _write_trial_neighborhood_excel(self, record: Dict[str, Any]) -> Optional[str]:
        """
        Escribe Excel para un trial aprobado con formato mejorado.
        
        Estructura del Excel:
        - Fila 1: Trial original (ORIGINAL)
        - Filas 2-N+1: Vecinos (VECINO_1, VECINO_2, ...)
        - Fila N+2: DISPERSIÓN (con max permitida en paréntesis)
        - Fila N+3: RESUMEN
        
        Columnas de dispersión: disp_ROI (50%), disp_SQN (50%), etc.
        Donde (50%) es la dispersión máxima permitida.
        """
        import pandas as pd
        import numpy as np
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        trial_num = record["trial"]
        strategy = str(record.get("strategy", "STRATEGY")).replace(" ", "_").replace("/", "_")
        score = record["score"]
        stability = record.get("stability_score", 0.0)
        
        # Nombre de archivo con score
        score_str = f"{score:.0f}"
        base_name = f"ROBUSTEZ_T{trial_num}_S{score_str}"
        excel_path = Path(self.output_dir) / f"{base_name}.xlsx"
        
        neigh_info = record.get("neighborhood_info", {}) or {}
        original_metrics = record.get("metrics", {}) or {}
        original_params = record.get("params", {}) or {}
        
        # Importar configuración
        try:
            from general.configuracion import VECINDARIO_MAX_DISPERSION
        except ImportError:
            VECINDARIO_MAX_DISPERSION = 0.50
        
        # =====================================================================
        # HELPER PARA VALORES SEGUROS
        # =====================================================================
        def safe_float(val, default=0.0):
            try:
                f = float(val) if val is not None else default
                return default if (f != f) else f  # NaN check
            except:
                return default
        
        # =====================================================================
        # CONSTRUIR FILAS
        # =====================================================================
        rows = []
        
        # Limpiar parámetros del original
        clean_original_params = {
            k: v for k, v in original_params.items() 
            if not str(k).startswith("__") and not str(k).startswith("exit_")
        }
        
        # ---- FILA ORIGINAL ----
        original_row = {
            "SUBTRIAL": f"T{trial_num}_ORIGINAL",
            "TIPO": "ORIGINAL",
            "SCORE": safe_float(neigh_info.get("original_score", score)),
            "ROI": safe_float(original_metrics.get("roi")),
            "SQN": safe_float(original_metrics.get("sqn")),
            "SHARPE": safe_float(original_metrics.get("sharpe")),
            "DRAWDOWN": safe_float(original_metrics.get("drawdown") or original_metrics.get("max_drawdown"), 50.0),
            "N_TRADES": int(safe_float(original_metrics.get("n_trades") or original_metrics.get("total_trades"), 0)),
            "TRADES_DIA": safe_float(original_metrics.get("trades_por_dia") or original_metrics.get("trades_dia")),
            "PROFIT_FACTOR": safe_float(original_metrics.get("profit_factor"), 1.0),
            "WIN_RATE": safe_float(original_metrics.get("win_rate") or original_metrics.get("porc_ganadoras")),
        }
        # Añadir parámetros
        for k, v in clean_original_params.items():
            original_row[f"P_{k.upper()}"] = v
        rows.append(original_row)
        
        # ---- FILAS DE VECINOS ----
        neighbor_scores = neigh_info.get("neighbor_scores", []) or []
        neighbor_metrics_list = neigh_info.get("neighbor_metrics", []) or []
        neighbor_params_list = neigh_info.get("neighbor_params", []) or []
        
        for i, n_score in enumerate(neighbor_scores):
            n_metrics = neighbor_metrics_list[i] if i < len(neighbor_metrics_list) else {}
            n_params = neighbor_params_list[i] if i < len(neighbor_params_list) else {}
            
            neighbor_row = {
                "SUBTRIAL": f"T{trial_num}_V{i+1}",
                "TIPO": f"VECINO_{i+1}",
                "SCORE": safe_float(n_score),
                "ROI": safe_float(n_metrics.get("roi")),
                "SQN": safe_float(n_metrics.get("sqn")),
                "SHARPE": safe_float(n_metrics.get("sharpe")),
                "DRAWDOWN": safe_float(n_metrics.get("drawdown") or n_metrics.get("max_drawdown"), 50.0),
                "N_TRADES": int(safe_float(n_metrics.get("n_trades") or n_metrics.get("total_trades"), 0)),
                "TRADES_DIA": safe_float(n_metrics.get("trades_por_dia") or n_metrics.get("trades_dia")),
                "PROFIT_FACTOR": safe_float(n_metrics.get("profit_factor"), 1.0),
                "WIN_RATE": safe_float(n_metrics.get("win_rate") or n_metrics.get("porc_ganadoras")),
            }
            # Añadir parámetros del vecino
            for k, v in n_params.items():
                if not str(k).startswith("__") and not str(k).startswith("exit_"):
                    neighbor_row[f"P_{k.upper()}"] = v
            rows.append(neighbor_row)
        
        # =====================================================================
        # USAR DISPERSIONES PRE-CALCULADAS (de neighborhood_fitness.py)
        # Para garantizar consistencia con robustness_approved
        # =====================================================================
        metricas_eval = ["ROI", "SQN", "SHARPE", "DRAWDOWN"]
        
        # Las dispersiones ya vienen calculadas en neigh_info
        dispersiones_precalc = neigh_info.get("dispersions", {})
        
        # Mapear nombres (neighborhood usa minúsculas)
        dispersiones = {
            "ROI": dispersiones_precalc.get("roi", 0.0),
            "SQN": dispersiones_precalc.get("sqn", 0.0),
            "SHARPE": dispersiones_precalc.get("sharpe", 0.0),
            "DRAWDOWN": dispersiones_precalc.get("drawdown", 0.0),
        }
        
        # Verificar aprobación por métrica
        aprobaciones = {m: dispersiones[m] <= VECINDARIO_MAX_DISPERSION for m in metricas_eval}
        
        # ---- FILA DISPERSIÓN ----
        max_disp_pct = int(VECINDARIO_MAX_DISPERSION * 100)
        disp_row = {
            "SUBTRIAL": "DISPERSIÓN",
            "TIPO": "CV%",
        }
        for m in metricas_eval:
            cv_pct = dispersiones[m] * 100
            tick = "✓" if aprobaciones[m] else "✗"
            disp_row[m] = f"{cv_pct:.1f}% {tick} (máx {max_disp_pct}%)"
        
        # Resultado global - DEBE coincidir con robustness_approved de neigh_info
        todas_aprueban = neigh_info.get("robustness_approved", all(aprobaciones.values()))
        disp_row["SCORE"] = "APROBADO ✓" if todas_aprueban else "SUSPENSO ✗"
        rows.append(disp_row)
        
        # ---- FILA RESUMEN ----
        avg_disp = neigh_info.get("avg_dispersion", np.mean(list(dispersiones.values())))
        resumen_row = {
            "SUBTRIAL": "RESUMEN",
            "TIPO": "ESTADÍSTICAS",
            "SCORE": safe_float(neigh_info.get("aggregated_score", score)),
            "ROI": f"μ={neigh_info.get('mean_score', 0):.2f}",
            "SQN": f"σ={neigh_info.get('std_score', 0):.2f}",
            "SHARPE": f"μ_sh={neigh_info.get('mean_sharpe', 0):.3f}",
            "DRAWDOWN": f"disp={avg_disp*100:.1f}%",
            "N_TRADES": f"n={neigh_info.get('n_neighbors_successful', 0)}/{neigh_info.get('n_neighbors_tested', 0)}",
            "TRADES_DIA": f"t={neigh_info.get('execution_time_ms', 0):.0f}ms",
        }
        rows.append(resumen_row)
        
        if not rows:
            return None
        
        # =====================================================================
        # CREAR DATAFRAME
        # =====================================================================
        df = pd.DataFrame(rows)
        
        # Ordenar columnas: SUBTRIAL, TIPO, SCORE, métricas, parámetros
        priority_cols = ["SUBTRIAL", "TIPO", "SCORE", "ROI", "SQN", "SHARPE", "DRAWDOWN",
                        "N_TRADES", "TRADES_DIA", "PROFIT_FACTOR", "WIN_RATE"]
        param_cols = sorted([c for c in df.columns if c.startswith("P_")])
        other_cols = [c for c in df.columns if c not in priority_cols and c not in param_cols]
        
        final_cols = [c for c in priority_cols if c in df.columns]
        final_cols += other_cols
        final_cols += param_cols
        
        df = df[[c for c in final_cols if c in df.columns]]
        
        # =====================================================================
        # ESCRIBIR EXCEL CON FORMATO
        # =====================================================================
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Robustez', index=False)
                worksheet = writer.sheets['Robustez']
                
                # Estilos
                header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                original_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                vecino_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
                disp_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
                resumen_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                border = Border(
                    left=Side(style='thin', color='BDC3C7'),
                    right=Side(style='thin', color='BDC3C7'),
                    top=Side(style='thin', color='BDC3C7'),
                    bottom=Side(style='thin', color='BDC3C7')
                )
                center_align = Alignment(horizontal='center', vertical='center')
                
                # Aplicar estilos al header
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border
                
                # Aplicar estilos a las filas de datos
                for row_idx in range(2, len(df) + 2):
                    tipo_cell = worksheet.cell(row=row_idx, column=2)
                    tipo_val = str(tipo_cell.value or "").upper()
                    
                    # Determinar fill según tipo
                    if "ORIGINAL" in tipo_val:
                        fill = original_fill
                    elif "VECINO" in tipo_val:
                        fill = vecino_fill
                    elif "CV" in tipo_val or "DISPERSIÓN" in tipo_val:
                        fill = disp_fill
                    else:
                        fill = resumen_fill
                    
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = fill
                        cell.alignment = center_align
                        cell.border = border
                
                # Ajustar ancho de columnas
                for idx, col in enumerate(df.columns, 1):
                    max_len = max(
                        len(str(col)),
                        df[col].astype(str).str.len().max() if len(df) > 0 else 0
                    )
                    worksheet.column_dimensions[worksheet.cell(1, idx).column_letter].width = min(max_len + 3, 30)
                
                worksheet.freeze_panes = 'A2'
            
            print(f"    [ROBUSTEZ] ✓ Guardado: {Path(excel_path).name}")
            return str(excel_path)
            
        except Exception as e:
            logger.warning(f"Error escribiendo Excel vecindario trial {trial_num}: {e}")
            return None


# nuclear_cleanup ya importado desde types arriba

# ============================================================================
# LÓGICA DE EJECUCIÓN (SINGLE EXIT TYPE)
# ============================================================================
def run_single_exit_type(
    exit_type: str,
    strategy: object,
    strategy_name: str,
    strategy_safe: str,
    activo: str,
    df_filtrado: object,
    tf_cache: dict[str, object],
    timeframe_base: str,
    cfg: BacktestConfig,
    tf_display: str,
    archivo_data: str,
    periodo_datos: str,
    resolve_archivo_data_tf_func: object = None,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
) -> None:

    # 1. CONFIGURACIÓN
    cfg_dict = cfg.__dict__.copy()
    cfg_dict["exit_type"] = str(exit_type)
    cfg_updated = BacktestConfig(**cfg_dict)

    # 2. DETECCIÓN DE CAPACIDADES
    try:
        indicadores = list(getattr(strategy, "parametros_optuna", {}).keys())
    except Exception:
        indicadores = []

    try:
        strategy_exit_enabled = bool(
            callable(getattr(strategy, "decide_exit", None))
            and bool(getattr(strategy, "ACTIVAR_SALIDA_PERSONALIZADA", False))
        )
    except Exception:
        strategy_exit_enabled = False

    # 3. MOSTRAR HEADER
    mostrar_cabecera_inicio(
        activo=activo,
        combo_nombre=strategy_name,
        indicadores=indicadores,
        n_trials=int(N_TRIALS),
        archivo_data=archivo_data,
        timeframe=tf_display,
        periodo=periodo_datos,
        exit_type=exit_type,
        strategy_exit_enabled=strategy_exit_enabled,
        perturbacion_activar=PERTURBACION_ACTIVAR,
    )

    # 4. RUTAS DE SALIDA
    # Estructura: resultados/ESTRATEGIA/TRAILING|FIXED/ACTIVO/GRAFICA|EXCEL|ROBUSTEZ
    activo_safe = str(activo).upper()
    exit_type_folder = "TRAILING" if "trail" in str(exit_type).lower() else "FIXED"
    strategy_root_dir = os.path.join(
        "resultados",
        strategy_safe,
        exit_type_folder,
        activo_safe,
    )
    excel_dir = os.path.join(strategy_root_dir, "EXCEL")
    graficos_dir = os.path.join(strategy_root_dir, "GRAFICA")
    robustez_dir = os.path.join(strategy_root_dir, "ROBUSTEZ")
    os.makedirs(excel_dir, exist_ok=True)
    os.makedirs(graficos_dir, exist_ok=True)
    os.makedirs(robustez_dir, exist_ok=True)

    # 5. REPORTEROS
    reporters = [ElegantRichReporter(saldo_inicial=cfg_updated.saldo_inicial, activo=activo)]

    if USAR_EXCEL:
        reporters.append(ExcelReporter(
            resumen_path=f"{excel_dir}/RESUMEN.xlsx",
            trades_base_dir=excel_dir,
            max_archivos=int(MAX_ARCHIVOS_GUARDAR)
        ))

    if GENERAR_PLOTS:
        reporters.append(PlotReporter(
            plot_base=graficos_dir,
            fecha_inicio_plot=FECHA_INICIO_PLOT,
            fecha_fin_plot=FECHA_FIN_PLOT,
            max_archivos=int(MAX_ARCHIVOS_GUARDAR),
            saldo_inicial=cfg_updated.saldo_inicial,
            activo=activo,
        ))

    # Reporter de Vecindario (cuando VECINDARIO_ACTIVAR=True y VECINDARIO_EXCEL=True)
    if VECINDARIO_ACTIVAR and VECINDARIO_EXCEL:
        reporters.append(NeighborhoodReporter(
            output_dir=robustez_dir,
            max_guardar=int(VECINDARIO_GUARDAR_MEJORES),
            activo=str(activo),
            timeframe=normalize_timeframe_to_suffix(timeframe_base),
        ))

    # 6. RUNNER
    runner = OptimizationRunner(config=cfg_updated, n_trials=int(N_TRIALS), reporters=reporters)

    # ========================================================================
    # CONFIGURAR SISTEMA DE EVALUACIÓN DE ROBUSTEZ
    # ========================================================================
    # El sistema ahora usa ÚNICAMENTE Neighborhood Fitness Aggregation:
    # - Evalúa la topología local (vecindario) alrededor de los parámetros
    # - Score = μ - λ·σ (media penalizada por varianza)
    # - Genera Excel con todas las métricas de vecinos
    # ========================================================================
    if VECINDARIO_ACTIVAR:
        runner.neighborhood_enabled = True
        # Obtener min_trades_per_day (para skip de neighborhood en trials con baja frecuencia)
        try:
            from general.configuracion import VECINDARIO_MIN_TRADES_DIA
            min_trades_day_cfg = float(VECINDARIO_MIN_TRADES_DIA)
        except (ImportError, AttributeError):
            min_trades_day_cfg = 0.25
        runner.neighborhood_config = NeighborhoodConfig(
            n_neighbors=int(VECINDARIO_N_NEIGHBORS),
            max_dispersion=float(VECINDARIO_MAX_DISPERSION),
            perturbation_std=float(VECINDARIO_PERTURBATION_STD),
            lambda_penalty=float(VECINDARIO_LAMBDA_PENALTY),
            seed=int(VECINDARIO_SEED) if VECINDARIO_SEED is not None else None,
            enabled=True,
            min_trades_per_day=min_trades_day_cfg,
        )
        logger.info(
            f"[NEIGHBORHOOD] Activado: {VECINDARIO_N_NEIGHBORS} vecinos, "
            f"σ={VECINDARIO_PERTURBATION_STD}, λ={VECINDARIO_LAMBDA_PENALTY}, "
            f"min_trades/día={min_trades_day_cfg}, max_dispersion={VECINDARIO_MAX_DISPERSION}"
        )
    else:
        runner.neighborhood_enabled = False

    # [CAMBIO CRÍTICO] n_jobs=1 para evitar que la RAM se multiplique por N núcleos y cause 'Killed'
    runner.optuna = OptunaConfig(seed=None, n_jobs=1, storage=None)

    runner.activo = activo

    try:
        # Carga diferida de timeframes extra
        entry_tf = getattr(strategy, "timeframe_entry", None) or timeframe_base
        exit_tf = getattr(strategy, "timeframe_exit", None) or timeframe_base
        needed_tfs = [timeframe_base, entry_tf, exit_tf]

        for tf in needed_tfs:
            tf_suf = normalize_timeframe_to_suffix(tf)
            if tf_suf in tf_cache:
                continue

            if resolve_archivo_data_tf_func:
                try:
                    path_tf = resolve_archivo_data_tf_func(activo, tf, formato="parquet")
                    df_tf = load_data(path_tf)
                    if fecha_inicio and fecha_fin:
                        df_tf = filter_by_date(df_tf, fecha_inicio, fecha_fin)
                    tf_cache[tf_suf] = df_tf
                except Exception as e:
                    logger.warning(f"No se pudo cargar TF extra {tf}: {e}")

        runner.optimize_strategies(
            df=df_filtrado,
            strategies=[strategy],
            df_by_timeframe=tf_cache,
            base_timeframe=timeframe_base,
        )

        # -------------------------------------------------------------------------
        # [CORRECCIÓN] VISUALIZACIÓN DE RESULTADOS COMPATIBLE CON NSGA-II (MULTI-OBJETIVO)
        # -------------------------------------------------------------------------
        if hasattr(runner, "_last_study") and runner._last_study:
            study = runner._last_study

            # Detectar si hay más de 1 objetivo
            is_multiobj = len(study.directions) > 1

            try:
                best_trial = None
                best_val = 0.0

                if is_multiobj:
                    # NSGA-II: Obtenemos la Frontera de Pareto
                    # Ordenamos por Calidad (values[0]) descendente para coger el "mejor"
                    # Nota: values[0] = Calidad, values[1] = Riesgo (Drawdown)
                    pareto_front = sorted(study.best_trials, key=lambda t: t.values[0], reverse=True)
                    if pareto_front:
                        best_trial = pareto_front[0]
                        best_val = best_trial.values[0]
                else:
                    # TPE Clásico: Existe un único best_trial
                    if study.best_trial:
                        best_trial = study.best_trial
                        best_val = study.best_value

                # Mostrar resultado si se encontró un trial válido
                if best_trial:
                    mostrar_fin_optimizacion(
                        total_trials=len(study.trials),
                        best_score=best_val,
                        best_trial=best_trial.number,
                        estrategia=strategy_name,
                    )
            except Exception as e:
                # Fallback seguro para no detener la ejecución si Optuna cambia
                logger.warning(f"No se pudo extraer el mejor trial para el reporte final: {e}")
        # -------------------------------------------------------------------------

    except KeyboardInterrupt:
        raise
    except Exception as e:
        logger.error(f"Error en {strategy_name}: {e}")
    finally:
        del runner
        del reporters
        # Limpieza nuclear para asegurar que la RAM se devuelve al SO
        nuclear_cleanup()

# ============================================================================
# LIMPIEZA Y CIERRE SEGURO
# ============================================================================
def _purge_pycache(*, root: Path, exclude: set[str]) -> None:
    try:
        for dirpath, dirnames, filenames in os.walk(root):
            p = Path(dirpath)
            if any(part in exclude for part in p.parts):
                dirnames[:] = []
                continue
            if p.name == "__pycache__":
                shutil.rmtree(p, ignore_errors=True)
                dirnames[:] = []
                continue
            for f in filenames:
                if f.endswith(".pyc"):
                    try:
                        (p / f).unlink(missing_ok=True)
                    except OSError:
                        pass
    except OSError:
        pass

class HealthGuard:
    @staticmethod
    def final_cleanup():
        nuclear_cleanup()
        # Usamos la variable global _PROJECT_ROOT capturada al inicio
        if CONFIG.get("PURGE_PYCACHE_ON_EXIT"):
            _purge_pycache(root=_PROJECT_ROOT, exclude={".git", ".venv", "data"})

def main() -> None:
    atexit.register(HealthGuard.final_cleanup)

    # 1. PARSEAR ESTRATEGIAS
    if COMBINACION_A_EJECUTAR == "all":
        ids = None
    elif isinstance(COMBINACION_A_EJECUTAR, (list, tuple)):
        ids = list(COMBINACION_A_EJECUTAR)
    else:
        ids = [int(COMBINACION_A_EJECUTAR)]

    # 2. ACTIVOS Y TIMEFRAMES
    activos = list(ACTIVOS) if ACTIVOS else [ACTIVO_PRIMARIO]

    raw_tfs = CONFIG.get("TIMEFRAMES", None)
    fallback_tf = CONFIG.get("TIMEFRAME", 15)

    def _ensure_int_list(raw, fallback):
        res = []
        candidatos = raw if isinstance(raw, (list, tuple)) else str(raw).split(",")
        for c in candidatos:
            try:
                clean = str(c).lower().replace("m", "")
                val = int(float(clean))
                if val > 0:
                    res.append(val)
            except (ValueError, TypeError):
                pass
        return res if res else [int(fallback)]

    tfs_run = _ensure_int_list(raw_tfs, fallback_tf)

    # 3. EJECUCIÓN
    try:
        for activo in activos:
            for tf_base in tfs_run:

                # Carga de datos
                try:
                    archivo = resolve_archivo_data_tf(activo, tf_base, formato="parquet")
                    if not os.path.exists(archivo):
                        archivo = resolve_archivo_data(activo)
                except Exception:
                    archivo = resolve_archivo_data(activo)

                if not os.path.exists(archivo):
                    print(f"❌ DATA NOT FOUND: {archivo}")
                    continue

                df = load_data(archivo)
                df_filtrado = filter_by_date(df, FECHA_INICIO, FECHA_FIN)

                periodo_str = ""
                try:
                    ts = df_filtrado["timestamp"] if "timestamp" in df_filtrado.columns else df_filtrado["datetime"]
                    periodo_str = f"{ts.min():%Y-%m-%d} -> {ts.max():%Y-%m-%d}"
                except (KeyError, AttributeError):
                    pass

                tf_cache = {normalize_timeframe_to_suffix(tf_base): df_filtrado}

                cfg = BacktestConfig(
                    saldo_inicial=float(CONFIG["SALDO_INICIAL"]),
                    saldo_operativo_max=float(CONFIG["SALDO_OPERATIVO_MAX"]),
                    comision_pct=float(CONFIG["COMISION_PCT"]),
                    comision_sides=int(CONFIG["COMISION_SIDES"]),
                    saldo_minimo_operativo=float(CONFIG["SALDO_MINIMO_OPERATIVO"]),
                    qty_max_activo=float(resolve_qty_max_activo(activo)),
                    saldo_usado=float(CONFIG["SALDO_USADO"]),
                    apalancamiento_max=float(CONFIG["APALANCAMIENTO_MAX"]),
                    riesgo_por_trade_pct=float(CONFIG["RIESGO_POR_TRADE_PCT"]),
                    optimize_qty_max_activo=bool(CONFIG["OPTIMIZAR_QTY_ACTIVO"]),
                    qty_max_activo_range=tuple(resolve_qty_max_activo_range(activo)),
                    exit_sl_pct=float(CONFIG["EXIT_SL_PCT"]),
                    exit_tp_pct=float(CONFIG["EXIT_TP_PCT"]),
                    exit_trail_act_pct=float(CONFIG["EXIT_TRAIL_ACT_PCT"]),
                    exit_trail_dist_pct=float(CONFIG["EXIT_TRAIL_DIST_PCT"]),
                    optimize_exits=bool(CONFIG["OPTIMIZAR_SALIDAS"]),
                    exit_sl_pct_range=tuple(CONFIG["EXIT_SL_PCT_RANGE"]),
                    exit_tp_pct_range=tuple(CONFIG["EXIT_TP_PCT_RANGE"]),
                    exit_trail_act_pct_range=tuple(CONFIG["EXIT_TRAIL_ACT_PCT_RANGE"]),
                    exit_trail_dist_pct_range=tuple(CONFIG["EXIT_TRAIL_DIST_PCT_RANGE"]),
                )

                loop_ids = ids if ids else [None]
                for sid in loop_ids:
                    strategies = instantiate_strategies(only_id=sid if sid else None)
                    if not strategies:
                        continue

                    for strat in strategies:
                        s_name = getattr(strat, 'name', "STRAT")
                        s_safe = re.sub(r"[^A-Z0-9_]+", "_", s_name.upper())

                        base_suf = normalize_timeframe_to_suffix(tf_base)
                        # Fix: use base_suf when strategy timeframe is None
                        entry_raw = getattr(strat, "timeframe_entry", None)
                        exit_raw = getattr(strat, "timeframe_exit", None)
                        entry_suf = normalize_timeframe_to_suffix(entry_raw) if entry_raw else base_suf
                        exit_suf = normalize_timeframe_to_suffix(exit_raw) if exit_raw else base_suf
                        # Simplificar display: solo mostrar timeframe base
                        tf_display = base_suf.upper()

                        e_type = str(CONFIG["EXIT_TYPE"]).lower()
                        types_run = ["pnl_fixed", "pnl_trailing"] if e_type == "all" else [CONFIG["EXIT_TYPE"]]

                        for et in types_run:
                            run_single_exit_type(
                                exit_type=et,
                                strategy=strat,
                                strategy_name=s_name,
                                strategy_safe=s_safe,
                                activo=activo,
                                df_filtrado=df_filtrado,
                                tf_cache=tf_cache,
                                timeframe_base=tf_base,
                                cfg=cfg,
                                tf_display=tf_display,
                                archivo_data=archivo,
                                periodo_datos=periodo_str,
                                resolve_archivo_data_tf_func=resolve_archivo_data_tf,
                                fecha_inicio=FECHA_INICIO,
                                fecha_fin=FECHA_FIN
                            )

                # LIMPIEZA NUCLEAR AL CAMBIAR DE ACTIVO/TF PARA LIBERAR DATOS VIEJOS
                del df, df_filtrado
                nuclear_cleanup()

    except KeyboardInterrupt:
        pass
    finally:
        HealthGuard.final_cleanup()
        # LIMPIEZA TOTAL AL FINALIZAR
        full_system_cleanup()
        print("\n✅ [CLEANUP] Memoria y recursos liberados correctamente.")

if __name__ == "__main__":
    main()
