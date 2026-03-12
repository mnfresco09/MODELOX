"""
================================================================================
PLATEAU.PY — Análisis de Zona de Meseta Multi-Activo
================================================================================

Encuentra regiones del espacio de parámetros donde el rendimiento medio
es estable y aceptable en todos los activos analizados.

USO:
    python plateau.py [archivo1.xlsx archivo2.xlsx ...]

    Modo 1 — Arrastrar archivos al terminal (recomendado):
        Escribe "python plateau.py " y arrastra los RESUMEN directamente.
        El activo se detecta automáticamente del path (BTC, GOLD, SP500…).

    Modo 2 — Sin argumentos (escanea carpeta "entrada/"):
        entrada/
            BTC/
                RESUMEN ID1.xlsx
            GOLD/
                RESUMEN ID2.xlsx

OUTPUT (carpeta "salida/"):
    salida/PLATEAU_BREAKOUT.xlsx
    salida/PLATEAU_REVERSION.xlsx
    ...

NOTA: Agrupa automáticamente por ESTRATEGIA. Cualquier número de archivos
      y activos.
================================================================================
"""

import os
import sys
import glob
import shlex
import warnings
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple

from sklearn.neighbors import KNeighborsRegressor
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

CARPETA_ENTRADA = Path("entrada")   # usado solo si no hay args en CLI
CARPETA_SALIDA  = Path("salida")

# Nombres de activos conocidos para inferir del path
ACTIVOS_CONOCIDOS = ["BTC", "ETH", "GOLD", "SILVER", "SP500", "NASDAQ", "BIST100", "SOL", "EUR"]

# Filtro mínimo de trades por día
MIN_TRADES_DIA = 0.15

# Métrica compuesta: ROI - PENALIZACION_DD * MAX_DD
# Ejemplo: ROI=50%, DD=15% → 50 - 2.0*15 = 20
PENALIZACION_DD = 2.0

# KNN: vecinos para suavizar el espacio de parámetros
KNN_VECINOS = 15

# Penalización de inconsistencia entre activos (mayor = prefiere parámetros
# que funcionan igual en todos los activos)
LAMBDA_STD = 0.5

# Filas en el output (con filtro de diversidad)
MAX_FILAS = 50

# Distancia mínima entre zonas seleccionadas (0-1, espacio normalizado).
# Evita mostrar 50 filas casi idénticas.
DIVERSIDAD_MINIMA = 0.05

# Filas para calcular RANGOS: se toman las mejores N por plateau score
# (sin filtro de diversidad) → rango concentrado en la zona de meseta real
TOP_N_PARA_RANGOS = 15

# Columnas fijas (no son parámetros de estrategia)
COLS_FIJAS = {
    "TRIAL", "ESTRATEGIA", "SCORE",
    "TRADES DIA", "LONG", "SHORT", "PROFIT FACTOR",
    "ROI", "MAX DD", "EXPECTATIVA",
    "BEN BRUTO", "BEN NETO", "SALDO ACTUAL", "COMISIONES",
}

# =============================================================================
# COLORES (mismo estilo que los RESUMEN actuales)
# =============================================================================

C_HEADER_DARK   = "2D3436"   # cabecera oscura
C_HEADER_MED    = "636E72"   # cabecera media (params)
C_HEADER_ACCENT = "2C3E50"   # rendimiento
C_TEXT_WHITE    = "FFFFFF"
C_TEXT_DARK     = "2D3436"
C_BORDER        = "DFE6E9"
C_ROW_ALT       = "F8F9FA"
C_GREEN_LIGHT   = "F0FFF4"
C_RED_LIGHT     = "FFF5F5"
C_ACCENT_GREEN  = "38A169"
C_ACCENT_RED    = "E53E3E"

FONT_NAME = "Calibri"


# =============================================================================
# CARGA DE DATOS
# =============================================================================

def cargar_resumen(ruta: Path) -> pd.DataFrame:
    """Lee un archivo RESUMEN ID*.xlsx y devuelve un DataFrame limpio."""
    try:
        # Fila 0 = sección (DATOS/MÉTRICAS/PARÁMETROS), fila 1 = cabecera real
        df_raw = pd.read_excel(ruta, header=None, engine="openpyxl")
        if df_raw.shape[0] < 3:
            return pd.DataFrame()

        header = df_raw.iloc[1].tolist()
        data   = df_raw.iloc[2:].copy()
        data.columns = header
        data = data.reset_index(drop=True)

        # Convertir a numérico donde aplique
        for col in data.columns:
            if col not in ("ESTRATEGIA",):
                data[col] = pd.to_numeric(data[col], errors="ignore")

        return data

    except Exception as e:
        print(f"  ⚠ Error leyendo {ruta.name}: {e}")
        return pd.DataFrame()


def inferir_activo(ruta: Path) -> str:
    """Infiere el nombre del activo a partir del path del archivo."""
    partes = [p.upper() for p in ruta.parts]
    for parte in reversed(partes):
        for activo in ACTIVOS_CONOCIDOS:
            if activo in parte:
                return activo
    # Fallback: nombre de la carpeta padre
    return ruta.parent.name.upper()


def _procesar_ruta(ruta: Path, datos: Dict) -> None:
    """Carga un archivo RESUMEN y lo agrega al dict datos."""
    activo = inferir_activo(ruta)
    print(f"  Cargando {ruta.name} (activo={activo}) ...", end="")

    df = cargar_resumen(ruta)
    if df.empty:
        print(" (vacío, ignorado)")
        return

    # Filtrar por trades mínimos
    if "TRADES DIA" in df.columns:
        antes = len(df)
        df = df[df["TRADES DIA"].fillna(0) >= MIN_TRADES_DIA]
        filtrados = antes - len(df)
        if filtrados:
            print(f" ({filtrados} filtrados TRADES DIA<{MIN_TRADES_DIA})", end="")

    if df.empty:
        print(" (todos filtrados, ignorado)")
        return

    # Detectar nombre de estrategia
    if "ESTRATEGIA" in df.columns and not df["ESTRATEGIA"].dropna().empty:
        estrategia = str(df["ESTRATEGIA"].dropna().iloc[0]).strip().upper()
    else:
        estrategia = ruta.stem.upper()

    datos.setdefault(estrategia, {})[activo] = df
    print(f" OK ({len(df)} trials, estrategia={estrategia})")


def descubrir_archivos() -> Dict[str, Dict[str, pd.DataFrame]]:
    """
    Retorna: {estrategia: {activo: df}}

    Modo 1 (args CLI): usa sys.argv[1:] como rutas de archivos.
    Modo 2 (sin args): escanea entrada/<ACTIVO>/RESUMEN*.xlsx
    """
    datos: Dict[str, Dict[str, pd.DataFrame]] = {}
    args = [a for a in sys.argv[1:] if not a.startswith("-")]

    if args:
        # ── Modo arrastrar: rutas en argumentos ──
        rutas = [Path(a) for a in args]
        print(f"\nArchivos recibidos: {len(rutas)}")
        for ruta in rutas:
            if not ruta.exists():
                print(f"  ⚠ No encontrado: {ruta}")
                continue
            if ruta.suffix.lower() != ".xlsx":
                print(f"  ⚠ Ignorado (no es .xlsx): {ruta.name}")
                continue
            _procesar_ruta(ruta, datos)
    else:
        # ── Modo interactivo: pedir arrastrar al terminal ──
        print()
        print("  Arrastra aquí los archivos RESUMEN que quieras analizar")
        print("  (puedes soltar varios a la vez) y pulsa Enter:\n")
        try:
            linea = input("  > ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nCancelado.")
            sys.exit(0)

        if not linea:
            print("ERROR: No se recibieron archivos.")
            sys.exit(1)

        # macOS pega múltiples paths juntos sin espacio: 'path1''path2''path3'
        # Insertar espacio entre comillas adyacentes antes de parsear
        linea_fixed = linea.replace("''", "' '")
        try:
            partes = shlex.split(linea_fixed)
        except ValueError:
            partes = linea_fixed.split()

        rutas = [Path(p) for p in partes]
        print(f"\nArchivos recibidos: {len(rutas)}")
        for ruta in rutas:
            if not ruta.exists():
                print(f"  ⚠ No encontrado: {ruta}")
                continue
            if ruta.suffix.lower() != ".xlsx":
                print(f"  ⚠ Ignorado (no es .xlsx): {ruta.name}")
                continue
            _procesar_ruta(ruta, datos)

    return datos


# =============================================================================
# IDENTIFICACIÓN DE COLUMNAS
# =============================================================================

def detectar_columnas_params(df: pd.DataFrame) -> List[str]:
    """Devuelve columnas que son parámetros de estrategia (excluye fijas)."""
    return [c for c in df.columns if str(c).strip().upper() not in COLS_FIJAS
            and not str(c).startswith("Unnamed")]


# =============================================================================
# MÉTRICA COMPUESTA
# =============================================================================

def calcular_metrica(df: pd.DataFrame) -> pd.Series:
    """ROI - PENALIZACION_DD * MAX_DD (ambos en tanto por uno o tanto por ciento)."""
    roi = pd.to_numeric(df.get("ROI", 0), errors="coerce").fillna(0)
    dd  = pd.to_numeric(df.get("MAX DD", 0), errors="coerce").fillna(0)
    # ROI y MAX DD vienen como fracción (e.g. 0.64 = 64%) en los RESUMEN
    # Convertir a porcentaje para que sean comparables
    return roi * 100 - PENALIZACION_DD * dd * 100


# =============================================================================
# NORMALIZACIÓN
# =============================================================================

def normalizar_params(df_params: pd.DataFrame,
                      mins: pd.Series = None,
                      maxs: pd.Series = None) -> Tuple[np.ndarray, pd.Series, pd.Series]:
    """Normaliza a [0,1]. Devuelve (array, mins, maxs)."""
    if mins is None:
        mins = df_params.min()
    if maxs is None:
        maxs = df_params.max()

    rango = (maxs - mins).replace(0, 1)  # evitar división por cero
    norm = (df_params - mins) / rango
    return norm.values, mins, maxs


# =============================================================================
# ALGORITMO DE PLATEAU
# =============================================================================

def calcular_plateau(activos_data: Dict[str, pd.DataFrame],
                     param_cols: List[str]) -> pd.DataFrame:
    """
    Calcula el plateau score para todos los trials candidatos.

    Estrategia:
    - Todos los trials de todos los activos = candidatos
    - Para cada candidato, estima su rendimiento en cada activo via KNN
    - Plateau = media(rendimientos_estimados) - LAMBDA_STD * std(rendimientos_estimados)

    Returns DataFrame con param_cols + columnas de resultado.
    """
    activos = list(activos_data.keys())

    # --- Calcular métrica y extraer params para cada activo ---
    asset_frames = {}
    for activo, df in activos_data.items():
        params = df[param_cols].copy()
        # Rellenar NaN con mediana de esa columna
        params = params.apply(lambda c: c.fillna(c.median()), axis=0)
        metrica = calcular_metrica(df)
        roi = pd.to_numeric(df.get("ROI", 0), errors="coerce").fillna(0) * 100
        dd  = pd.to_numeric(df.get("MAX DD", 0), errors="coerce").fillna(0) * 100
        asset_frames[activo] = {
            "params_raw": params,
            "metrica": metrica,
            "roi": roi,
            "dd": dd,
        }

    # --- Construir espacio global de candidatos (unión de todos los trials) ---
    all_params_raw = pd.concat(
        [af["params_raw"] for af in asset_frames.values()],
        ignore_index=True
    )
    # Origen de cada candidato (para el output)
    all_activos_origen = []
    for activo, af in asset_frames.items():
        all_activos_origen.extend([activo] * len(af["params_raw"]))

    # --- Normalizar usando rango global ---
    mins_global = all_params_raw.min()
    maxs_global = all_params_raw.max()
    all_params_norm, _, _ = normalizar_params(all_params_raw, mins_global, maxs_global)

    # --- Para cada activo, entrenar KNN en su propio espacio normalizado ---
    predicciones = {}   # {activo: array de métricas estimadas para todos los candidatos}
    roi_predicciones = {}
    dd_predicciones  = {}

    for activo, af in asset_frames.items():
        X_train_norm, _, _ = normalizar_params(af["params_raw"], mins_global, maxs_global)
        y_metrica = af["metrica"].values
        y_roi     = af["roi"].values
        y_dd      = af["dd"].values

        n_vecinos = min(KNN_VECINOS, len(y_metrica))

        knn_m = KNeighborsRegressor(n_neighbors=n_vecinos, weights="distance")
        knn_r = KNeighborsRegressor(n_neighbors=n_vecinos, weights="distance")
        knn_d = KNeighborsRegressor(n_neighbors=n_vecinos, weights="distance")

        knn_m.fit(X_train_norm, y_metrica)
        knn_r.fit(X_train_norm, y_roi)
        knn_d.fit(X_train_norm, y_dd)

        predicciones[activo]    = knn_m.predict(all_params_norm)
        roi_predicciones[activo] = knn_r.predict(all_params_norm)
        dd_predicciones[activo]  = knn_d.predict(all_params_norm)

    # --- Calcular plateau score ---
    scores_matrix = np.column_stack([predicciones[a] for a in activos])
    if scores_matrix.ndim == 1:
        scores_matrix = scores_matrix.reshape(-1, 1)

    media  = scores_matrix.mean(axis=1)
    std    = scores_matrix.std(axis=1) if len(activos) > 1 else np.zeros(len(media))
    plateau = media - LAMBDA_STD * std

    # --- Construir DataFrame resultado ---
    resultado = all_params_raw.copy()
    resultado["PLATEAU_SCORE"] = plateau
    resultado["METRICA_MEDIA"] = media
    resultado["METRICA_STD"]   = std
    resultado["ORIGEN_ACTIVO"] = all_activos_origen

    for activo in activos:
        resultado[f"ROI_{activo}"]     = roi_predicciones[activo]
        resultado[f"DD_{activo}"]      = dd_predicciones[activo]
        resultado[f"METRICA_{activo}"] = predicciones[activo]

    return resultado.sort_values("PLATEAU_SCORE", ascending=False).reset_index(drop=True)


def seleccionar_diversas(df: pd.DataFrame,
                         param_cols: List[str],
                         mins: pd.Series,
                         maxs: pd.Series,
                         n: int = MAX_FILAS) -> pd.DataFrame:
    """
    Selecciona hasta n filas diversas: descarta candidatos demasiado
    cercanos a uno ya seleccionado (en espacio normalizado).
    """
    rango = (maxs - mins).replace(0, 1)
    params_norm = (df[param_cols] - mins) / rango

    seleccionadas = []
    seleccionadas_norm = []

    for idx in range(len(df)):
        punto = params_norm.iloc[idx].values
        if not seleccionadas_norm:
            seleccionadas.append(idx)
            seleccionadas_norm.append(punto)
        else:
            mat  = np.array(seleccionadas_norm)
            dist = np.linalg.norm(mat - punto, axis=1).min()
            if dist >= DIVERSIDAD_MINIMA:
                seleccionadas.append(idx)
                seleccionadas_norm.append(punto)

        if len(seleccionadas) >= n:
            break

    return df.iloc[seleccionadas].reset_index(drop=True)


# =============================================================================
# ESCRITURA EXCEL
# =============================================================================

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _font(bold=False, color=C_TEXT_WHITE, size=10) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def _border_thin() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def escribir_excel(df_resultado: pd.DataFrame,
                   df_top_concentrado: pd.DataFrame,
                   param_cols: List[str],
                   activos: List[str],
                   estrategia: str,
                   ruta_salida: Path) -> None:
    """Escribe el archivo Excel de plateau con el estilo minimalista.
    df_resultado       — 50 zonas diversas (hoja PLATEAU)
    df_top_concentrado — top N por score puro (referencia de RANGOS)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PLATEAU"

    # ─── Columnas del output ─────────────────────────────────────────────────
    # RANK | params... | ROI_A | DD_A | ... | METRICA_MEDIA | METRICA_STD | PLATEAU
    col_rank   = 1
    col_params_start = 2
    col_params_end   = col_params_start + len(param_cols) - 1

    col_assets_start = col_params_end + 1
    # Por activo: ROI, DD, METRICA → 3 cols cada uno
    cols_por_activo = 3
    col_assets_end  = col_assets_start + len(activos) * cols_por_activo - 1

    col_resumen_start = col_assets_end + 1
    col_resumen_end   = col_resumen_start + 2   # MEDIA, STD, PLATEAU

    total_cols = col_resumen_end

    # ─── Fila 1: Título ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=total_cols)
    cell = ws.cell(row=1, column=1,
                   value=f"PLATEAU — {estrategia}  |  Activos: {', '.join(activos)}  |  "
                         f"KNN={KNN_VECINOS}  λ={LAMBDA_STD}  "
                         f"Métrica=ROI-{PENALIZACION_DD}×DD")
    cell.fill    = _fill(C_HEADER_DARK)
    cell.font    = _font(bold=True, size=12)
    cell.alignment = _center()

    # ─── Fila 2: Cabeceras de sección ────────────────────────────────────────
    ws.row_dimensions[2].height = 18

    def seccion(c_ini, c_fin, texto, color):
        if c_ini == c_fin:
            ws.cell(row=2, column=c_ini, value=texto).fill = _fill(color)
        else:
            ws.merge_cells(start_row=2, start_column=c_ini,
                           end_row=2,   end_column=c_fin)
            ws.cell(row=2, column=c_ini, value=texto).fill = _fill(color)
        for c in range(c_ini, c_fin + 1):
            cel = ws.cell(row=2, column=c)
            cel.fill      = _fill(color)
            cel.font      = _font(bold=True, size=9)
            cel.alignment = _center()
            cel.border    = _border_thin()

    seccion(col_rank,           col_rank,           "ID",             C_HEADER_DARK)
    seccion(col_params_start,   col_params_end,     "PARÁMETROS",     C_HEADER_MED)
    if activos:
        seccion(col_assets_start, col_assets_end,   "RENDIMIENTO POR ACTIVO", C_HEADER_ACCENT)
    seccion(col_resumen_start,  col_resumen_end,    "RESUMEN PLATEAU", C_HEADER_DARK)

    # ─── Fila 3: Nombres de columna ──────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    headers = ["RANK"] + list(param_cols)
    for activo in activos:
        headers += [f"ROI {activo} %", f"DD {activo} %", f"MÉTRICA {activo}"]
    headers += ["MÉTRICA MEDIA", "MÉTRICA STD", "PLATEAU SCORE"]

    for ci, h in enumerate(headers, start=1):
        cel = ws.cell(row=3, column=ci, value=h)
        # Color fondo según sección
        if ci == col_rank:
            bg = C_HEADER_DARK
        elif col_params_start <= ci <= col_params_end:
            bg = "4A4A4A"
        elif col_assets_start <= ci <= col_assets_end:
            bg = "3D5A7A"
        else:
            bg = C_HEADER_DARK
        cel.fill      = _fill(bg)
        cel.font      = _font(bold=True, size=9)
        cel.alignment = _center()
        cel.border    = _border_thin()

    # ─── Filas de datos ──────────────────────────────────────────────────────
    def fmt_val(v, col_name=""):
        """Formatea valores numéricos."""
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return None
        col_up = str(col_name).upper()
        if "ROI" in col_up or "DD" in col_up or "MÉTRICA" in col_up or "PLATEAU" in col_up:
            return round(float(v), 2)
        # Param: si es entero, sin decimales
        if isinstance(v, (int, np.integer)):
            return int(v)
        fv = float(v)
        return int(fv) if fv == int(fv) else round(fv, 4)

    for i, row_data in df_resultado.iterrows():
        xl_row = i + 4   # fila 4 en adelante
        ws.row_dimensions[xl_row].height = 16

        bg_row = "FFFFFF" if i % 2 == 0 else C_ROW_ALT

        # RANK
        cel = ws.cell(row=xl_row, column=col_rank, value=i + 1)
        cel.fill = _fill(bg_row); cel.font = _font(color=C_TEXT_DARK, bold=True, size=9)
        cel.alignment = _center(); cel.border = _border_thin()

        # PARAMS
        for pi, pcol in enumerate(param_cols):
            ci  = col_params_start + pi
            val = fmt_val(row_data.get(pcol), pcol)
            cel = ws.cell(row=xl_row, column=ci, value=val)
            cel.fill = _fill(bg_row); cel.font = _font(color=C_TEXT_DARK, size=9)
            cel.alignment = _center(); cel.border = _border_thin()

        # RENDIMIENTO POR ACTIVO
        for ai, activo in enumerate(activos):
            base_ci = col_assets_start + ai * cols_por_activo
            for ji, (clave, col_name) in enumerate([
                (f"ROI_{activo}",     f"ROI {activo} %"),
                (f"DD_{activo}",      f"DD {activo} %"),
                (f"METRICA_{activo}", f"MÉTRICA {activo}"),
            ]):
                val = fmt_val(row_data.get(clave), clave)
                ci  = base_ci + ji
                cel = ws.cell(row=xl_row, column=ci, value=val)
                cel.fill      = _fill(bg_row)
                cel.font      = _font(color=C_TEXT_DARK, size=9)
                cel.alignment = _center()
                cel.border    = _border_thin()
                # Color rojo/verde para ROI y DD
                if val is not None:
                    if "ROI" in clave and val < 0:
                        cel.fill = _fill("FFF0F0")
                    elif "ROI" in clave and val > 0:
                        cel.fill = _fill("F0FFF4")
                    elif "DD" in clave and val > 25:
                        cel.fill = _fill("FFF0F0")

        # RESUMEN
        media    = fmt_val(row_data.get("METRICA_MEDIA"), "MÉTRICA")
        std_val  = fmt_val(row_data.get("METRICA_STD"),   "STD")
        plateau  = fmt_val(row_data.get("PLATEAU_SCORE"), "PLATEAU")

        for ji, (ci, val) in enumerate([
            (col_resumen_start,     media),
            (col_resumen_start + 1, std_val),
            (col_resumen_end,       plateau),
        ]):
            cel = ws.cell(row=xl_row, column=ci, value=val)
            cel.border    = _border_thin()
            cel.alignment = _center()
            cel.font      = _font(color=C_TEXT_DARK, bold=(ji == 2), size=9)
            if ji == 2 and val is not None:   # PLATEAU SCORE — gradiente visual
                cel.fill = _fill("E8F5E9") if val >= 0 else _fill("FFEBEE")
            else:
                cel.fill = _fill(bg_row)

    # ─── Ancho de columnas ───────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(col_rank)].width = 6
    for pi, pcol in enumerate(param_cols):
        ci = col_params_start + pi
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(pcol)) + 2, 12)
    for ai, activo in enumerate(activos):
        for ji in range(cols_por_activo):
            ci = col_assets_start + ai * cols_por_activo + ji
            ws.column_dimensions[get_column_letter(ci)].width = max(len(activo) + 8, 13)
    for ci in range(col_resumen_start, col_resumen_end + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    # ─── Formato condicional en PLATEAU SCORE (escala verde) ─────────────────
    if len(df_resultado) > 1:
        rng = f"{get_column_letter(col_resumen_end)}4:{get_column_letter(col_resumen_end)}{3 + len(df_resultado)}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min",   start_color="FFCDD2",
            mid_type="num",     mid_value=0, mid_color="FFFFFF",
            end_type="max",     end_color="C8E6C9",
        ))

    # ─── Hoja _TOP — top N concentrado (fuente de RANGOS) ───────────────────
    # Se escribe oculta; RANGOS apunta aquí para calcular rangos estrechos.
    ws_top = wb.create_sheet("_TOP")
    ws_top.sheet_state = "hidden"

    # Cabecera en fila 1, datos desde fila 2
    for ci, col in enumerate(param_cols, start=1):
        ws_top.cell(row=1, column=ci, value=col)
    for ri, (_, row_data) in enumerate(df_top_concentrado[param_cols].iterrows(), start=2):
        for ci, col in enumerate(param_cols, start=1):
            v = row_data[col]
            ws_top.cell(row=ri, column=ci,
                        value=(None if pd.isna(v) else (int(v) if float(v) == int(v) else round(float(v), 6))))

    top_fila_ini = 2
    top_fila_fin = 1 + len(df_top_concentrado)

    # ─── Hoja RANGOS — tabla de rangos recomendados con fórmulas ────────────
    ws_rng = wb.create_sheet("RANGOS")

    fila_datos_ini = 4                  # usado para PLATEAU (referencia legacy)
    fila_datos_fin = 3 + len(df_resultado)

    # Columnas de la hoja RANGOS
    # PARÁMETRO | RANGO RECOMENDADO | MÍNIMO | MÁXIMO | MEDIANA | AMPLITUD
    HDR_RANGOS = ["PARÁMETRO", "RANGO RECOMENDADO", "MÍNIMO", "MÁXIMO", "MEDIANA", "AMPLITUD"]

    # ── Fila 1: título ────────────────────────────────────────────────────
    ws_rng.row_dimensions[1].height = 26
    ws_rng.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HDR_RANGOS))
    cel = ws_rng.cell(row=1, column=1,
                      value=f"RANGOS RECOMENDADOS — {estrategia}  "
                            f"(basado en top {len(df_top_concentrado)} zonas por plateau score)")
    cel.fill = _fill(C_HEADER_DARK); cel.font = _font(bold=True, size=12); cel.alignment = _center()

    # ── Fila 2: cabeceras ─────────────────────────────────────────────────
    ws_rng.row_dimensions[2].height = 20
    COL_COLORS = [C_HEADER_DARK, C_HEADER_ACCENT, "3A7D44", "C0392B", "7D6608", C_HEADER_MED]
    for ci, (h, col_bg) in enumerate(zip(HDR_RANGOS, COL_COLORS), start=1):
        cel = ws_rng.cell(row=2, column=ci, value=h)
        cel.fill = _fill(col_bg); cel.font = _font(bold=True, size=10)
        cel.alignment = _center(); cel.border = _border_thin()

    # ── Filas de datos: una por parámetro ────────────────────────────────
    for pi, pcol in enumerate(param_cols):
        # _TOP tiene params desde columna A (índice 1)
        xl_top_col = get_column_letter(pi + 1)
        rng_ref = f"_TOP!{xl_top_col}{top_fila_ini}:{xl_top_col}{top_fila_fin}"

        xl_row = pi + 3
        ws_rng.row_dimensions[xl_row].height = 18
        bg = "FFFFFF" if pi % 2 == 0 else C_ROW_ALT

        # Determinar si el parámetro es entero o decimal (del primer valor no-nulo)
        sample = df_resultado[pcol].dropna()
        es_entero = (sample == sample.round(0)).all() if len(sample) > 0 else True
        fmt_excel = '0' if es_entero else '0.00'

        f_min    = f"=MIN({rng_ref})"
        f_max    = f"=MAX({rng_ref})"
        f_median = f"=MEDIAN({rng_ref})"
        f_amp    = f"=MAX({rng_ref})-MIN({rng_ref})"
        # Rango legible: "MIN — MAX" como texto concatenado
        f_rango  = (f'=TEXT(MIN({rng_ref}),"{fmt_excel}")'
                    f'&" — "'
                    f'&TEXT(MAX({rng_ref}),"{fmt_excel}")')

        valores = [
            (1, pcol,     None,       False, True),    # PARÁMETRO (texto fijo)
            (2, f_rango,  None,       False, True),    # RANGO (fórmula texto)
            (3, f_min,    fmt_excel,  True,  False),   # MÍNIMO
            (4, f_max,    fmt_excel,  True,  False),   # MÁXIMO
            (5, f_median, fmt_excel,  True,  False),   # MEDIANA
            (6, f_amp,    fmt_excel,  True,  False),   # AMPLITUD
        ]

        for ci, val, num_fmt, is_num, is_text in valores:
            cel = ws_rng.cell(row=xl_row, column=ci, value=val)
            cel.fill   = _fill(bg)
            cel.border = _border_thin()
            if ci == 1:  # nombre parámetro
                cel.font      = _font(color=C_TEXT_DARK, bold=True, size=10)
                cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif ci == 2:  # rango legible — destacado
                cel.font      = _font(color="1A5276", bold=True, size=10)
                cel.alignment = _center()
            else:
                cel.font      = _font(color=C_TEXT_DARK, size=10)
                cel.alignment = _center()
                if num_fmt:
                    cel.number_format = num_fmt

    # ── Anchos ───────────────────────────────────────────────────────────
    anchos = [28, 22, 12, 12, 12, 12]
    for ci, w in enumerate(anchos, start=1):
        ws_rng.column_dimensions[get_column_letter(ci)].width = w

    # ── Nota al pie ──────────────────────────────────────────────────────
    fila_nota = len(param_cols) + 4
    ws_rng.merge_cells(start_row=fila_nota, start_column=1, end_row=fila_nota, end_column=len(HDR_RANGOS))
    cel_nota = ws_rng.cell(row=fila_nota, column=1,
                           value=f"* Rangos calculados sobre las {len(df_top_concentrado)} zonas con mayor "
                                 f"plateau score (hoja _TOP). Las fórmulas son vivas.")
    cel_nota.font      = Font(name=FONT_NAME, italic=True, size=9, color="636E72")
    cel_nota.alignment = Alignment(horizontal="left", vertical="center")

    # ─── Hoja de info / parámetros usados ────────────────────────────────────
    ws_info = wb.create_sheet("INFO")
    info_rows = [
        ("Parámetro",          "Valor"),
        ("Estrategia",         estrategia),
        ("Activos analizados", ", ".join(activos)),
        ("Min trades/día",     MIN_TRADES_DIA),
        ("KNN vecinos",        KNN_VECINOS),
        ("Penalización DD",    PENALIZACION_DD),
        ("Lambda STD",         LAMBDA_STD),
        ("Diversidad mínima",  DIVERSIDAD_MINIMA),
        ("Filas output",       len(df_resultado)),
        ("Fórmula métrica",    f"ROI% - {PENALIZACION_DD} × DD%"),
        ("Fórmula plateau",    f"media(métricas_activos) - {LAMBDA_STD} × std(métricas_activos)"),
    ]
    for r, (k, v) in enumerate(info_rows, start=1):
        ws_info.cell(row=r, column=1, value=k).font  = Font(bold=(r == 1), name=FONT_NAME)
        ws_info.cell(row=r, column=2, value=v).font  = Font(name=FONT_NAME)
    ws_info.column_dimensions["A"].width = 25
    ws_info.column_dimensions["B"].width = 45

    wb.save(ruta_salida)
    print(f"  → Guardado: {ruta_salida}")


# =============================================================================
# MAIN
# =============================================================================

def procesar_estrategia(estrategia: str,
                         activos_data: Dict[str, pd.DataFrame]) -> None:
    activos = list(activos_data.keys())
    print(f"\n{'='*60}")
    print(f"  ESTRATEGIA: {estrategia}  |  Activos: {activos}")
    print(f"{'='*60}")

    # Detectar columnas de parámetros (intersección: solo las que existen en todos)
    sets_params = [set(detectar_columnas_params(df)) for df in activos_data.values()]
    if not sets_params:
        print("  Sin datos, omitiendo.")
        return

    # Usar unión: si un param no existe en un activo, se rellena con mediana
    param_cols = sorted(sets_params[0].union(*sets_params[1:]) if len(sets_params) > 1 else sets_params[0])

    # Asegurar que todos los DFs tienen las mismas columnas de params
    for activo, df in activos_data.items():
        for pc in param_cols:
            if pc not in df.columns:
                df[pc] = np.nan

    print(f"  Parámetros detectados ({len(param_cols)}): {param_cols}")
    for activo, df in activos_data.items():
        print(f"  {activo}: {len(df)} trials válidos")

    # Calcular plateau
    df_resultado = calcular_plateau(activos_data, param_cols)

    # Top N concentrado (sin diversidad) → fuente para RANGOS
    df_top_concentrado = df_resultado.head(TOP_N_PARA_RANGOS)

    # Seleccionar zonas diversas para PLATEAU (exploración visual)
    mins_global = df_resultado[param_cols].min()
    maxs_global = df_resultado[param_cols].max()
    df_top = seleccionar_diversas(df_resultado, param_cols, mins_global, maxs_global, MAX_FILAS)

    print(f"  Zonas plateau seleccionadas: {len(df_top)}")
    print(f"  Top 5 plateau scores: "
          f"{df_top_concentrado['PLATEAU_SCORE'].head(5).round(2).tolist()}")
    print(f"  Rango RANGOS: top {len(df_top_concentrado)} más concentrados")

    # Guardar Excel
    CARPETA_SALIDA.mkdir(exist_ok=True)
    nombre_archivo = f"PLATEAU_{estrategia.replace(' ', '_')}.xlsx"
    ruta_salida = CARPETA_SALIDA / nombre_archivo

    escribir_excel(
        df_resultado       = df_top,
        df_top_concentrado = df_top_concentrado,
        param_cols         = param_cols,
        activos            = activos,
        estrategia         = estrategia,
        ruta_salida        = ruta_salida,
    )


def main():
    print("=" * 60)
    print("  PLATEAU.PY — Análisis de Zona de Meseta Multi-Activo")
    print("=" * 60)

    datos = descubrir_archivos()

    if not datos:
        print("ERROR: No se pudieron cargar datos.")
        sys.exit(1)

    print(f"\nEstrategias encontradas: {list(datos.keys())}")

    for estrategia, activos_data in sorted(datos.items()):
        procesar_estrategia(estrategia, activos_data)

    print(f"\n{'='*60}")
    print(f"  Análisis completado. Resultados en: {CARPETA_SALIDA}/")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
