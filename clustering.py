"""
================================================================================
CLUSTERING.PY - Agrupación Inteligente de Trials por Vecindad de Parámetros
================================================================================
Analiza un Excel de resultados MODELOX, detecta automáticamente los parámetros
de la estrategia, y agrupa los trials en clusters de vecinos usando HDBSCAN
con distancias adaptativas por tipo de parámetro.

Uso:
    python clustering.py
    → Pedirá arrastrar el archivo Excel de resultados

Autor: Sistema MODELOX
================================================================================
"""

import os
import sys
import re
import warnings
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from scipy import stats
from sklearn.preprocessing import RobustScaler
from sklearn.neighbors import NearestNeighbors
from kneed import KneeLocator

import hdbscan

from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

# ==============================================================================
# CONFIGURACIÓN DE ESTILO (Idéntico a visual/excel.py)
# ==============================================================================

COLORS = {
    "header_bg_metrics": "1A5276",  # Azul Oscuro (Métricas)
    "header_bg_params":  "566573",  # Gris Plomo (Parámetros)
    "header_bg_id":      "212F3D",  # Negro Azulado (IDs)
    "header_bg_cluster": "7D3C98",  # Púrpura (Cluster ID)
    "text_white":        "FFFFFF",
    "text_dark":         "212F3D",
    "border_color":      "BDC3C7",
    "success_bg":        "D5F5E3",  # Verde claro
    "danger_bg":         "FADBD8",  # Rojo claro
}

# Paleta de colores para separadores de cluster (20 colores distintos)
CLUSTER_PALETTE = [
    "2E86C1",  # Azul
    "28B463",  # Verde
    "D4AC0D",  # Dorado
    "CB4335",  # Rojo
    "8E44AD",  # Púrpura
    "E67E22",  # Naranja
    "1ABC9C",  # Turquesa
    "EC7063",  # Coral
    "5DADE2",  # Celeste
    "45B39D",  # Verde Mar
    "F4D03F",  # Amarillo
    "AF7AC5",  # Lavanda
    "EB984E",  # Melocotón
    "85C1E9",  # Azul Claro
    "82E0AA",  # Verde Claro
    "F1948A",  # Rosa
    "BB8FCE",  # Lila
    "F0B27A",  # Arena
    "76D7C4",  # Menta
    "AEB6BF",  # Gris
]

FONT_TITLE = "Arial"
FONT_BODY = "Arial"

# Métricas conocidas del sistema MODELOX (para detectar params vs metrics)
KNOWN_METRICS = {
    "SALDO_ACTUAL", "ROI_PCT", "ROI%", "PROFIT_FACTOR", "WINRATE_PCT",
    "WINRATE%", "TOTAL_TRADES", "TRADES_DIA", "MAX_DD_PCT", "MAX_DD%",
    "SHARPE", "SQN", "ESTABILIDAD", "AVG_TRADE", "EXPECTATIVA",
    "WIN_STREAK", "LOSS_STREAK", "NUM_LONGS", "NUM_SHORTS",
    "SHARPE_RATIO", "SORTINO", "CALMAR", "KELLY",
    "NET_PROFIT", "PNL_NETO", "NET_PNL",
}

KNOWN_IDS = {"TRIAL", "ESTRATEGIA", "SCORE", "STRATEGY"}

# Keywords que indican que una columna es métrica, NO parámetro
METRIC_KEYWORDS = [
    "PROFIT", "LOSS", "PNL", "NET", "GROSS", "SALDO", "BALANCE", "RETORNO",
    "RETURN", "ROI", "BENEFICIO", "RIESGO", "RISK", "REWARD", "COMISION",
    "FEES", "WIN", "GANADORA", "PERDEDORA", "ACIERTO", "RATE",
    "DRAWDOWN", "DD", "RACHA", "STREAK", "UNDERWATER",
    "RATIO", "FACTOR", "SHARPE", "SORTINO", "CALMAR", "SQN",
    "EXPECTATIVA", "KELLY", "COUNT", "NUM_", "N_TRADES",
    "LONGS", "SHORTS", "TRADES", "METRIC", "RESULT",
    "BEST", "WORST", "DIA_OPERADO", "DURATION",
]

# Excepciones: keywords de parámetros legítimos
PARAM_EXCEPTIONS = [
    "STOP", "SL", "TP", "TRAIL", "PERIOD", "LEN", "FAST", "SLOW",
    "SIGNAL", "LIMIT", "THRESHOLD", "SIGMA", "OFFSET", "ATR",
    "MA", "EMA", "SMA", "ZLEMA", "RSI", "MACD", "BB", "BOLL",
    "LOOKBACK", "LOOKBAR", "WINDOW", "DIST", "MULT", "FACTOR_",
    "EXIT_SL", "EXIT_TP", "TAKE", "LOSS_",
]


# ==============================================================================
# 1. LECTURA INTELIGENTE DEL ARCHIVO
# ==============================================================================

def solicitar_archivo() -> str:
    """Solicita al usuario arrastrar un archivo Excel."""
    print("\n" + "=" * 70)
    print("  MODELOX - CLUSTERING INTELIGENTE DE TRIALS")
    print("=" * 70)
    print("\n  Arrastra aquí el archivo Excel de resultados y pulsa Enter:\n")

    path = input("  📂 Archivo: ").strip()

    # Limpiar comillas y escapes de drag-and-drop en macOS
    path = path.strip("'\"")
    path = path.replace("\\ ", " ")

    if not os.path.isfile(path):
        print(f"\n  ❌ No se encontró el archivo: {path}")
        sys.exit(1)

    if not path.lower().endswith((".xlsx", ".xls")):
        print(f"\n  ❌ El archivo debe ser .xlsx o .xls")
        sys.exit(1)

    print(f"\n  ✅ Archivo cargado: {os.path.basename(path)}")
    return path


def leer_excel_modelox(filepath: str) -> Tuple[pd.DataFrame, List[str], List[str], List[str]]:
    """
    Lee un Excel MODELOX con headers en fila 2 (fila 1 = títulos de grupo).
    Retorna: (df, id_cols, metric_cols, param_cols)
    """
    # Intentar leer con header en fila 1 (0-indexed)
    df = pd.read_excel(filepath, header=1)

    # Si las columnas tienen formato multi-level, aplanar
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [str(c[-1]).strip() for c in df.columns]

    # Normalizar nombres
    df.columns = [str(c).strip() for c in df.columns]

    # Detectar automáticamente IDs, métricas y parámetros
    id_cols = []
    metric_cols = []
    param_cols = []

    for col in df.columns:
        col_upper = col.upper().replace("%", "_PCT")

        # ¿Es columna de ID?
        if col_upper in KNOWN_IDS or col == "TRIAL":
            id_cols.append(col)
            continue

        # ¿Es métrica conocida?
        if col_upper in KNOWN_METRICS or col in KNOWN_METRICS:
            metric_cols.append(col)
            continue

        # Heurístico: ¿parece métrica?
        is_metric = False
        is_param_exception = False

        for kw in PARAM_EXCEPTIONS:
            if kw in col_upper:
                is_param_exception = True
                break

        if not is_param_exception:
            for kw in METRIC_KEYWORDS:
                if kw in col_upper:
                    is_metric = True
                    break

        if is_metric:
            metric_cols.append(col)
        else:
            # Verificar que sea numérica y tenga variación
            if df[col].dtype in [np.float64, np.int64, np.float32, np.int32, float, int]:
                if df[col].nunique() > 1:
                    param_cols.append(col)
                else:
                    metric_cols.append(col)  # Constante → no es parámetro útil
            elif pd.to_numeric(df[col], errors="coerce").notna().sum() > len(df) * 0.8:
                df[col] = pd.to_numeric(df[col], errors="coerce")
                if df[col].nunique() > 1:
                    param_cols.append(col)
            else:
                metric_cols.append(col)

    print(f"\n  📊 Estructura detectada:")
    print(f"     • {len(df)} trials")
    print(f"     • {len(id_cols)} columnas ID: {id_cols}")
    print(f"     • {len(metric_cols)} métricas: {metric_cols}")
    print(f"     • {len(param_cols)} parámetros: {param_cols}")

    return df, id_cols, metric_cols, param_cols


# ==============================================================================
# 2. CLUSTERING INTELIGENTE CON DISTANCIAS ADAPTATIVAS
# ==============================================================================

def analizar_tipo_parametro(series: pd.Series, name: str) -> Dict:
    """
    Analiza un parámetro: clasifica su tipo, asigna peso adaptativo,
    y calcula tolerancia ROBUSTA basada en la distribución real.

    Tolerancia: Ensemble de 3 estimadores robustos:
    1. Freedman-Diaconis: 2 × IQR × n^(-1/3)
    2. MAD: 1.4826 × median(|xi - median(x)|)
    3. Scott adaptado: 3.49 × MAD × n^(-1/3)
    Resultado = mediana de los 3, con safety clamps.
    """
    clean = series.dropna()
    if len(clean) == 0:
        return {"type": "unknown", "weight": 1.0}

    values = clean.values.astype(float)
    n = len(values)
    vmin, vmax = float(values.min()), float(values.max())
    rango = vmax - vmin
    n_unique = clean.nunique()
    std = float(values.std())
    name_upper = name.upper()

    # ── Estadísticas robustas ──
    q25, q50, q75 = np.percentile(values, [25, 50, 75])
    iqr = q75 - q25
    p5, p95 = np.percentile(values, [5, 95])
    rango_efectivo = p95 - p5

    mad_raw = float(np.median(np.abs(values - q50)))
    mad_sigma = 1.4826 * mad_raw

    info = {
        "name": name, "min": vmin, "max": vmax, "range": rango,
        "n_unique": n_unique, "std": std, "iqr": iqr,
        "mad": mad_sigma, "rango_efectivo": rango_efectivo,
    }

    # ── Clasificación del tipo + pesos adaptativos (propios de HDBSCAN) ──
    if rango == 0:
        info.update({"type": "constant", "weight": 0.0, "neighbor_tolerance": 0})
        return info

    if ("%" in name or "PCT" in name_upper or "SL" in name_upper or "TP" in name_upper) and vmax <= 100:
        info["type"] = "percentage"
        info["weight"] = 1.5
    elif n_unique <= 10:
        info["type"] = "discrete_few"
        info["weight"] = 2.0
    elif rango <= 30:
        info["type"] = "narrow_range"
        info["weight"] = 1.8
    elif rango <= 200:
        info["type"] = "medium_range"
        info["weight"] = 1.2
    else:
        info["type"] = "wide_range"
        info["weight"] = 1.0

    # ── Caso especial: discreto con pocos valores ──
    if n_unique <= 10:
        sorted_unique = np.sort(clean.unique())
        if len(sorted_unique) >= 2:
            steps = np.diff(sorted_unique)
            min_step = float(steps.min())
            tol = max(min_step, 0.5)
        else:
            tol = 1.0
        info["neighbor_tolerance"] = tol
    else:
        # ── Cálculo robusto de tolerancia (3 estimadores) ──
        n_cbrt = n ** (1.0 / 3.0)

        fd_width = 2.0 * iqr / n_cbrt if iqr > 0 else rango / n_cbrt
        mad_tol = mad_sigma
        scott_width = 3.49 * mad_sigma / n_cbrt if mad_sigma > 0 else rango / n_cbrt

        candidates = sorted([fd_width, mad_tol, scott_width])
        tol = candidates[1]  # Mediana

        # Safety clamps
        sorted_unique = np.sort(clean.unique())
        if len(sorted_unique) >= 2:
            min_step = float(np.diff(sorted_unique).min())
            tol = max(tol, min_step)
        tol = max(tol, 0.5)

        if rango_efectivo > 0:
            tol = min(tol, rango_efectivo * 0.15)
        tol = max(tol, 0.5)

        info["neighbor_tolerance"] = tol

    # Ajuste por nombre: indicadores técnicos → más peso
    technical_kw = ["RSI", "MACD", "BB", "BOLL", "ATR", "ADX"]
    if any(kw in name_upper for kw in technical_kw):
        info["weight"] *= 1.3

    return info


def construir_matriz_distancia_adaptativa(
    df_params: pd.DataFrame,
    param_info: Dict[str, Dict],
) -> np.ndarray:
    """
    Construye matriz de features normalizada con pesos adaptativos
    usando RobustScaler (resistente a outliers) + ponderación por tipo.
    """
    features = []
    weights = []

    for col in df_params.columns:
        info = param_info[col]
        if info.get("type") == "constant" or info.get("weight", 0) == 0:
            continue

        values = df_params[col].values.reshape(-1, 1)

        # RobustScaler: usa mediana e IQR en lugar de media/std
        # Más robusto ante outliers en datos financieros
        scaler = RobustScaler()
        scaled = scaler.fit_transform(values).ravel()

        features.append(scaled)
        weights.append(info["weight"])

    if not features:
        raise ValueError("No se encontraron parámetros válidos para clustering")

    X = np.column_stack(features)

    # Aplicar pesos
    weights = np.array(weights)
    weights = weights / weights.sum() * len(weights)  # Normalizar preservando escala
    X = X * weights[np.newaxis, :]

    return X


def determinar_min_cluster_size(n_samples: int) -> int:
    """
    Determina el tamaño mínimo de cluster adaptativo.
    Más pequeño = más clusters, más granularidad.
    """
    if n_samples < 50:
        return max(3, n_samples // 5)
    elif n_samples < 200:
        return max(5, n_samples // 10)
    elif n_samples < 1000:
        return max(8, n_samples // 20)
    elif n_samples < 5000:
        return max(10, n_samples // 100)
    else:
        return max(15, n_samples // 150)


def determinar_min_samples(n_samples: int, n_params: int) -> int:
    """
    min_samples adaptativo basado en dimensionalidad.
    Más bajo = más sensible a estructura local.
    """
    base = max(2, n_params - 1)
    if n_samples < 100:
        return min(base, max(2, n_samples // 5))
    elif n_samples < 1000:
        return min(base + 1, max(3, n_samples // 30))
    else:
        return min(base + 2, max(3, n_samples // 60))


def ejecutar_hdbscan(
    X: np.ndarray,
    param_info: Dict[str, Dict],
) -> np.ndarray:
    """
    Ejecuta HDBSCAN con parámetros optimizados automáticamente.

    HDBSCAN es superior a DBSCAN/K-Means porque:
    1. No requiere especificar K (número de clusters)
    2. Detecta clusters de densidad variable
    3. Maneja ruido naturalmente (-1 = outlier)
    4. Estable con hiperparámetros por defecto
    """
    n_samples = X.shape[0]
    n_features = X.shape[1]

    min_cluster = determinar_min_cluster_size(n_samples)
    min_samp = determinar_min_samples(n_samples, n_features)

    print(f"\n  🔬 HDBSCAN Config:")
    print(f"     • min_cluster_size = {min_cluster}")
    print(f"     • min_samples = {min_samp}")
    print(f"     • metric = euclidean")
    print(f"     • cluster_selection = leaf (granular)")

    # Intentar primero con 'leaf' (más granular, mejor para hiperparámetros)
    clusterer = hdbscan.HDBSCAN(
        min_cluster_size=min_cluster,
        min_samples=min_samp,
        metric="euclidean",
        cluster_selection_method="leaf",  # Leaf = más clusters, más granular
        cluster_selection_epsilon=0.0,
        allow_single_cluster=False,
        prediction_data=True,
    )

    labels = clusterer.fit_predict(X)

    n_clusters = len(set(labels) - {-1})
    n_noise = (labels == -1).sum()

    # Si leaf no encuentra suficiente estructura, probar eom
    if n_clusters < 3:
        print(f"     • leaf: {n_clusters} clusters → probando eom...")
        clusterer_eom = hdbscan.HDBSCAN(
            min_cluster_size=max(5, min_cluster // 2),
            min_samples=max(2, min_samp // 2),
            metric="euclidean",
            cluster_selection_method="eom",
            allow_single_cluster=False,
            prediction_data=True,
        )
        labels_eom = clusterer_eom.fit_predict(X)
        n_clusters_eom = len(set(labels_eom) - {-1})

        if n_clusters_eom > n_clusters:
            labels = labels_eom
            n_clusters = n_clusters_eom
            n_noise = (labels == -1).sum()
            print(f"     • eom: {n_clusters} clusters")

    n_clusters = len(set(labels) - {-1})
    n_noise = (labels == -1).sum()

    print(f"\n  📦 Resultado:")
    print(f"     • {n_clusters} clusters detectados")
    print(f"     • {n_noise} trials sin cluster (ruido/outliers)")

    # Si demasiados outliers (>40%), reasignar al cluster más cercano
    if n_noise > 0.4 * n_samples and n_clusters > 0:
        print(f"\n  ⚠️  Reasignando {n_noise} outliers al cluster más cercano...")
        labels = _reasignar_outliers(X, labels)
        n_noise_after = (labels == -1).sum()
        print(f"     • Outliers restantes: {n_noise_after}")

    # Si HDBSCAN no encuentra estructura (todo -1), fallback a K-Means adaptativo
    if n_clusters == 0:
        print("\n  ⚠️  HDBSCAN no encontró estructura → Fallback a Birch + KMeans")
        labels = _fallback_clustering(X, n_samples)

    return labels


def _reasignar_outliers(X: np.ndarray, labels: np.ndarray) -> np.ndarray:
    """Reasigna outliers al cluster más cercano usando k-NN."""
    labels = labels.copy()
    mask_cluster = labels >= 0
    mask_noise = labels == -1

    if mask_cluster.sum() == 0:
        return labels

    # Construir k-NN con puntos clusterizados
    nn = NearestNeighbors(n_neighbors=5, metric="euclidean")
    nn.fit(X[mask_cluster])

    distances, indices = nn.kneighbors(X[mask_noise])

    # Para cada outlier, asignar el cluster mayoritario de sus vecinos
    cluster_labels_valid = labels[mask_cluster]
    noise_indices = np.where(mask_noise)[0]

    for i, idx in enumerate(noise_indices):
        neighbor_labels = cluster_labels_valid[indices[i]]
        # Voto mayoritario
        unique, counts = np.unique(neighbor_labels, return_counts=True)
        labels[idx] = unique[counts.argmax()]

    return labels


def _fallback_clustering(X: np.ndarray, n_samples: int) -> np.ndarray:
    """Fallback: estima K óptimo con Silhouette y usa KMeans."""
    from sklearn.cluster import KMeans
    from sklearn.metrics import silhouette_score

    # Estimar K con silueta
    max_k = min(30, max(3, n_samples // 20))
    min_k = 3

    if max_k < min_k:
        max_k = min_k + 1

    best_k = min_k
    best_score = -1

    print(f"     Probando K de {min_k} a {max_k}...")

    for k in range(min_k, max_k + 1):
        km = KMeans(n_clusters=k, random_state=42, n_init=10, max_iter=300)
        lbl = km.fit_predict(X)
        score = silhouette_score(X, lbl, sample_size=min(3000, n_samples))
        if score > best_score:
            best_score = score
            best_k = k

    print(f"     • K óptimo (Silhouette): {best_k} (score={best_score:.3f})")

    km = KMeans(n_clusters=best_k, random_state=42, n_init=20, max_iter=500)
    return km.fit_predict(X)


# ==============================================================================
# 3. ANÁLISIS DE CLUSTERS
# ==============================================================================

def analizar_clusters(
    df: pd.DataFrame,
    labels: np.ndarray,
    param_cols: List[str],
    metric_cols: List[str],
    param_info: Dict[str, Dict],
) -> pd.DataFrame:
    """
    Genera resumen estadístico de cada cluster y ordena por score medio.
    """
    df = df.copy()
    df["__CLUSTER__"] = labels

    resumen_clusters = []

    for cluster_id in sorted(df["__CLUSTER__"].unique()):
        mask = df["__CLUSTER__"] == cluster_id
        subset = df[mask]
        n = len(subset)

        info = {
            "CLUSTER": cluster_id if cluster_id >= 0 else "OUTLIERS",
            "N_TRIALS": n,
        }

        # Score medio
        if "SCORE" in df.columns:
            info["SCORE_MEDIO"] = subset["SCORE"].mean()
            info["SCORE_MAX"] = subset["SCORE"].max()

        # Rangos de parámetros
        for col in param_cols:
            if col in subset.columns:
                info[f"{col}_MIN"] = subset[col].min()
                info[f"{col}_MAX"] = subset[col].max()
                info[f"{col}_MEDIA"] = subset[col].mean()

        resumen_clusters.append(info)

    df_resumen = pd.DataFrame(resumen_clusters)

    # Ordenar clusters por score medio descendente
    sort_col = "SCORE_MEDIO" if "SCORE_MEDIO" in df_resumen.columns else "N_TRIALS"
    df_resumen = df_resumen.sort_values(sort_col, ascending=False).reset_index(drop=True)

    return df_resumen


def imprimir_resumen(
    df_resumen: pd.DataFrame,
    param_cols: List[str],
    param_info: Dict[str, Dict],
):
    """Imprime resumen de clusters en consola."""
    print("\n" + "=" * 70)
    print("  RESUMEN DE CLUSTERS")
    print("=" * 70)

    for _, row in df_resumen.iterrows():
        cid = row["CLUSTER"]
        n = int(row["N_TRIALS"])
        label = f"Cluster {cid}" if cid != "OUTLIERS" else "🔸 OUTLIERS"

        print(f"\n  ── {label} ({n} trials) ──")

        if "SCORE_MEDIO" in row:
            print(f"     Score: media={row['SCORE_MEDIO']:.2f}, max={row['SCORE_MAX']:.2f}")

        for col in param_cols:
            min_key = f"{col}_MIN"
            max_key = f"{col}_MAX"
            if min_key in row and max_key in row:
                vmin = row[min_key]
                vmax = row[max_key]
                vmedia = row.get(f"{col}_MEDIA", (vmin + vmax) / 2)
                tipo = param_info.get(col, {}).get("type", "?")
                print(f"     {col}: [{vmin:.2f} - {vmax:.2f}] (media={vmedia:.2f}) [{tipo}]")

    print()


# ==============================================================================
# 4. GENERACIÓN DE EXCEL CON FORMATO PROFESIONAL
# ==============================================================================

def generar_excel_clustered(
    df_original: pd.DataFrame,
    labels: np.ndarray,
    id_cols: List[str],
    metric_cols: List[str],
    param_cols: List[str],
    df_resumen_clusters: pd.DataFrame,
    output_path: str,
    saldo_inicial: float = 300.0,
):
    """
    Genera Excel con el mismo formato que MODELOX original,
    pero con trials agrupados por cluster y separadores visuales.
    """
    df = df_original.copy()
    df["__CLUSTER__"] = labels

    # Orden de clusters por score medio (del resumen)
    cluster_order = df_resumen_clusters["CLUSTER"].tolist()

    # Columna CLUSTER para el Excel
    final_id_cols = ["CLUSTER"] + [c for c in id_cols if c in df.columns]
    final_metric_cols = [c for c in metric_cols if c in df.columns]
    final_param_cols = [c for c in param_cols if c in df.columns]

    all_cols = final_id_cols + final_metric_cols + final_param_cols

    # Construir DataFrame final con separadores
    rows_data = []
    separator_rows = []  # (row_index, cluster_id) para colorear
    current_row = 0  # Fila actual (0-based en los datos)

    for cluster_id in cluster_order:
        if cluster_id == "OUTLIERS":
            mask = df["__CLUSTER__"] == -1
        else:
            mask = df["__CLUSTER__"] == cluster_id

        subset = df[mask].copy()
        if subset.empty:
            continue

        # Ordenar dentro del cluster por SCORE desc
        if "SCORE" in subset.columns:
            subset = subset.sort_values("SCORE", ascending=False)

        # Fila separadora del cluster
        sep_row = {col: "" for col in all_cols}
        n_trials = len(subset)

        if "SCORE" in subset.columns:
            score_medio = subset["SCORE"].mean()
            sep_row["CLUSTER"] = f"━━ CLUSTER {cluster_id} ━━ ({n_trials} trials, Score medio: {score_medio:.2f}) ━━"
        else:
            sep_row["CLUSTER"] = f"━━ CLUSTER {cluster_id} ━━ ({n_trials} trials) ━━"

        rows_data.append(sep_row)
        separator_rows.append((current_row, cluster_id))
        current_row += 1

        # Filas de datos
        for _, trial_row in subset.iterrows():
            data_row = {}
            data_row["CLUSTER"] = cluster_id if cluster_id != "OUTLIERS" else -1

            for col in id_cols:
                if col in trial_row:
                    data_row[col] = trial_row[col]

            for col in metric_cols:
                if col in trial_row:
                    data_row[col] = trial_row[col]

            for col in param_cols:
                if col in trial_row:
                    data_row[col] = trial_row[col]

            rows_data.append(data_row)
            current_row += 1

    df_final = pd.DataFrame(rows_data, columns=all_cols)

    # Guardar Excel base con openpyxl directamente (evita save/reload para archivos grandes)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clusters"

    # Escribir fila 1 (grupo titles) y fila 2 (headers) manualmente
    # Primero, headers en fila 2
    for j, col_name in enumerate(all_cols, 1):
        ws.cell(row=2, column=j, value=col_name)

    # Escribir datos desde fila 3
    for i, row_data in enumerate(rows_data, 3):
        for j, col_name in enumerate(all_cols, 1):
            val = row_data.get(col_name, "")
            ws.cell(row=i, column=j, value=val)

    # Aplicar estilos directamente (sin save+reload)
    _aplicar_estilo_clustered(
        wb,
        ws,
        df_final,
        final_id_cols,
        final_metric_cols,
        final_param_cols,
        separator_rows,
        saldo_inicial,
    )

    # Hoja 2: Resumen de Clusters
    _agregar_hoja_resumen(wb, df_resumen_clusters, param_cols)

    # Guardar una sola vez
    wb.save(output_path)

    return output_path


def _aplicar_estilo_clustered(
    wb,
    ws,
    df: pd.DataFrame,
    id_cols: List[str],
    metric_cols: List[str],
    param_cols: List[str],
    separator_rows: List[Tuple[int, any]],
    saldo_ini: float,
):
    """Aplica el estilo profesional idéntico al de MODELOX + separadores de cluster."""
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    max_row = ws.max_row

    n_ids = len(id_cols)
    n_metrics = len(metric_cols)
    n_params = len(param_cols)

    start_metrics = n_ids + 1
    end_metrics = start_metrics + n_metrics - 1
    start_params = end_metrics + 1
    end_params = start_params + n_params - 1

    border_full = Border(
        left=Side(style="thin", color=COLORS["border_color"]),
        right=Side(style="thin", color=COLORS["border_color"]),
        top=Side(style="thin", color=COLORS["border_color"]),
        bottom=Side(style="thin", color=COLORS["border_color"]),
    )

    # ── ROW 1: TÍTULOS DE GRUPO ──
    font_group = Font(name=FONT_TITLE, size=12, bold=True, color=COLORS["text_white"])
    align_group = Alignment(horizontal="center", vertical="center")

    # IDs (incluye CLUSTER)
    for c in range(1, start_metrics):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_id"])
        cell.font = font_group
        cell.alignment = align_group

    # Métricas
    if n_metrics > 0:
        cell = ws.cell(row=1, column=start_metrics)
        cell.value = "MÉTRICAS CLAVE"
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
        cell.font = font_group
        cell.alignment = align_group
        if n_metrics > 1:
            ws.merge_cells(
                start_row=1, start_column=start_metrics,
                end_row=1, end_column=end_metrics,
            )
        for c in range(start_metrics, end_metrics + 1):
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])

    # Parámetros
    if n_params > 0:
        cell = ws.cell(row=1, column=start_params)
        cell.value = "PARÁMETROS ESTRATEGIA"
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_params"])
        cell.font = font_group
        cell.alignment = align_group
        if n_params > 1:
            ws.merge_cells(
                start_row=1, start_column=start_params,
                end_row=1, end_column=end_params,
            )
        for c in range(start_params, end_params + 1):
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=COLORS["header_bg_params"])

    # ── ROW 2: HEADERS ──
    font_header = Font(name=FONT_TITLE, size=12, bold=True, color=COLORS["text_white"])

    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_full

        if col < start_metrics:
            cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_id"])
        elif col <= end_metrics:
            cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
        else:
            cell.fill = PatternFill("solid", fgColor=COLORS["header_bg_params"])

        val_len = len(str(cell.value)) if cell.value else 0
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, val_len + 2), 22)

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    # ── DATOS (ROW 3+) ──
    font_body = Font(name=FONT_BODY, size=12, color=COLORS["text_dark"])

    # Mapear separadores: fila Excel = separator_data_row + 3 (header=2 rows)
    separator_excel_rows = set()
    for data_row_idx, cluster_id in separator_rows:
        excel_row = data_row_idx + 3  # +1 for group title + 1 for header + 1 for 1-based
        separator_excel_rows.add((excel_row, cluster_id))

    sep_row_set = {r for r, _ in separator_excel_rows}
    cluster_for_row = {r: cid for r, cid in separator_excel_rows}

    for r in range(3, max_row + 1):
        is_separator = r in sep_row_set

        if is_separator:
            # ── FILA SEPARADORA ──
            cluster_id = cluster_for_row[r]
            if isinstance(cluster_id, str) and "OUTLIER" in cluster_id:
                color_idx = len(CLUSTER_PALETTE) - 1
            elif isinstance(cluster_id, (int, np.integer)):
                color_idx = int(cluster_id) % len(CLUSTER_PALETTE)
            else:
                color_idx = 0

            sep_color = CLUSTER_PALETTE[color_idx]
            sep_fill = PatternFill("solid", fgColor=sep_color)
            sep_font = Font(name=FONT_TITLE, size=11, bold=True, color=COLORS["text_white"])

            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = sep_fill
                cell.font = sep_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_full

            # Merge la fila del separador para el texto
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
            ws.row_dimensions[r].height = 25
        else:
            # ── FILA DE DATOS ──
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = font_body
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_full

                header_val = str(ws.cell(2, c).value or "").upper()

                if isinstance(cell.value, (int, float)):
                    if "TRADES" in header_val and "DIA" in header_val:
                        cell.number_format = "0.00"
                    elif "TRADES" in header_val or "NUM_" in header_val or "CLUSTER" in header_val:
                        cell.number_format = "0"
                    elif any(kw in header_val for kw in ["SCORE", "%", "PCT", "RATIO", "SHARPE", "FACTOR", "ESTABILIDAD"]):
                        cell.number_format = "0.00"
                    elif any(kw in header_val for kw in ["SALDO", "PROFIT", "PNL"]):
                        cell.number_format = "#,##0.00"

    # ── CONDITIONAL FORMATTING (solo filas de datos) ──
    col_map = {}
    for c in range(1, max_col + 1):
        val = str(ws.cell(2, c).value or "").strip()
        if val:
            col_map[val] = get_column_letter(c)

    if "SCORE" in col_map:
        col_score = col_map["SCORE"]
        ws.conditional_formatting.add(
            f"{col_score}3:{col_score}{max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    if "ROI_PCT" in col_map or "ROI%" in col_map:
        col_roi = col_map.get("ROI_PCT") or col_map.get("ROI%")
        ws.conditional_formatting.add(
            f"{col_roi}3:{col_roi}{max_row}",
            DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True),
        )

    if "SALDO_ACTUAL" in col_map:
        l_idx = list(col_map.keys()).index("SALDO_ACTUAL")
        col_num = list(range(1, max_col + 1))[l_idx]
        for row in range(3, max_row + 1):
            if row in sep_row_set:
                continue
            cell = ws.cell(row=row, column=col_num)
            try:
                val = float(cell.value)
                if val >= saldo_ini * 1.5:
                    cell.fill = PatternFill("solid", fgColor=COLORS["success_bg"])
                    cell.font = Font(name=FONT_BODY, size=12, color="006100", bold=True)
                elif val < saldo_ini:
                    cell.fill = PatternFill("solid", fgColor=COLORS["danger_bg"])
                    cell.font = Font(name=FONT_BODY, size=12, color="9C0006")
            except Exception:
                pass

    # Freeze panes (usar string para evitar problema con merged cells)
    freeze_col_letter = get_column_letter(start_metrics)
    ws.freeze_panes = f"{freeze_col_letter}3"


def _agregar_hoja_resumen(
    wb,
    df_resumen: pd.DataFrame,
    param_cols: List[str],
):
    """Agrega hoja 'Resumen Clusters' con estadísticas por cluster."""

    if "Resumen Clusters" in wb.sheetnames:
        del wb["Resumen Clusters"]

    ws = wb.create_sheet("Resumen Clusters")

    # Headers
    headers = list(df_resumen.columns)
    font_header = Font(name=FONT_TITLE, size=11, bold=True, color=COLORS["text_white"])
    fill_header = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_full = Border(
        left=Side(style="thin", color=COLORS["border_color"]),
        right=Side(style="thin", color=COLORS["border_color"]),
        top=Side(style="thin", color=COLORS["border_color"]),
        bottom=Side(style="thin", color=COLORS["border_color"]),
    )

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_full
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, len(h) + 2), 22)

    # Datos
    font_body = Font(name=FONT_BODY, size=11, color=COLORS["text_dark"])

    for i, (_, row) in enumerate(df_resumen.iterrows(), 2):
        for j, col in enumerate(headers, 1):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = font_body
            cell.alignment = align_center
            cell.border = border_full

            if isinstance(val, float):
                cell.number_format = "0.00"

    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.sheet_view.showGridLines = False


# ==============================================================================
# 5. PIPELINE PRINCIPAL
# ==============================================================================

def main():
    """Pipeline completo de clustering."""

    # 1. Solicitar archivo
    filepath = solicitar_archivo()

    # 2. Leer datos
    df, id_cols, metric_cols, param_cols = leer_excel_modelox(filepath)

    if not param_cols:
        print("\n  ❌ No se detectaron columnas de parámetros. No se puede clusterizar.")
        sys.exit(1)

    # 3. Analizar parámetros
    print("\n  🔍 Análisis de parámetros:")
    param_info = {}
    for col in param_cols:
        info = analizar_tipo_parametro(df[col], col)
        param_info[col] = info
        print(f"     • {col}: tipo={info['type']}, peso={info['weight']:.2f}, "
              f"rango=[{info['min']:.2f}-{info['max']:.2f}], tolerancia_vecino={info.get('neighbor_tolerance', 0):.2f}")

    # 4. Construir features
    print("\n  ⚙️  Construyendo matriz de distancias adaptativas...")
    X = construir_matriz_distancia_adaptativa(df[param_cols], param_info)
    print(f"     • Matriz: {X.shape[0]} samples × {X.shape[1]} features")

    # 5. Clustering
    print("\n  🧠 Ejecutando clustering HDBSCAN...")
    labels = ejecutar_hdbscan(X, param_info)

    # 6. Análisis
    df_resumen = analizar_clusters(df, labels, param_cols, metric_cols, param_info)
    imprimir_resumen(df_resumen, param_cols, param_info)

    # 7. Generar Excel
    base_dir = os.path.dirname(filepath)
    base_name = os.path.splitext(os.path.basename(filepath))[0]
    output_path = os.path.join(base_dir, f"{base_name}_CLUSTERS.xlsx")

    print(f"\n  📊 Generando Excel: {os.path.basename(output_path)}...")

    generar_excel_clustered(
        df_original=df,
        labels=labels,
        id_cols=id_cols,
        metric_cols=metric_cols,
        param_cols=param_cols,
        df_resumen_clusters=df_resumen,
        output_path=output_path,
    )

    print(f"\n  ✅ ¡Excel generado exitosamente!")
    print(f"     📁 {output_path}")
    print(f"\n     Hojas:")
    print(f"       1. 'Clusters' - Trials agrupados con separadores")
    print(f"       2. 'Resumen Clusters' - Estadísticas por cluster")
    print("\n" + "=" * 70)


if __name__ == "__main__":
    main()
