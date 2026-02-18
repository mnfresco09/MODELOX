
import sys
import os
import types
from typing import Dict, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
import pandas as pd
import polars as pl
import logging

# Configurar path
current_dir = os.getcwd()
if current_dir not in sys.path:
    sys.path.append(current_dir)

# Imports del proyecto
from modelox.core.data import load_data, resample_to_base_timeframe
from modelox.core.types import normalize_timeframe_to_suffix
from modelox.strategies.registry import instantiate_strategies
from general.configuracion import (
    ACTIVO_PRIMARIO, CONFIG, resolve_archivo_data, resolve_archivo_data_tf
)

# Configuración de Logging
logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("DEBUG")

# Estilos (Minimalistas)
COLORS = {
    "header_bg": "262626", # Gris Oscuro
    "text_white": "FFFFFF",
    "text_dark": "333333",
    "border_color": "E0E0E0",
    "success_bg": "E6F4EA", # Verde suave
    "danger_bg": "FCE8E6",  # Rojo suave
    "neutral_bg": "FFFFFF",
    "success_text": "006100",
    "danger_text": "9C0006",
}

FONT_HEADER = Font(name="Arial", size=10, bold=True, color=COLORS["text_white"])
FONT_BODY = Font(name="Arial", size=10, color=COLORS["text_dark"])
FONT_SUCCESS = Font(name="Arial", size=12, color=COLORS["success_text"], bold=True)
FONT_DANGER = Font(name="Arial", size=11, color=COLORS["danger_text"], bold=True)

BORDER_FULL = Border(
    left=Side(style='thin', color=COLORS["border_color"]),
    right=Side(style='thin', color=COLORS["border_color"]),
    top=Side(style='thin', color=COLORS["border_color"]),
    bottom=Side(style='thin', color=COLORS["border_color"])
)

# === CONFIGURACIÓN DE USUARIO ===
STRATEGY_ID = 1      # <--- CAMBIAR ID AQUÍ
LIMIT_CANDLES = 5000  # <--- VELAS A REVISAR
DEBUG_START_DATE = "2025-01-01" # <--- INICIO DEL REPORTE (Dejar None para usar tail)

class MinParamTrial:
    """Mock de Optuna Trial que devuelve siempre el valor MÍNIMO del rango."""
    def suggest_int(self, name, low, high, step=1, log=False): return low
    def suggest_float(self, name, low, high, step=None, log=False): return low
    def suggest_categorical(self, name, choices): return choices[0]

def get_debug_params(strategy_id: int) -> Dict[str, Any]:
    """Obtiene los parámetros MÍNIMOS definidos en la estrategia."""
    try:
        strategies = instantiate_strategies(only_id=strategy_id)
        if not strategies:
            logger.warning(f"No se pudo instanciar estrategia {strategy_id} para obtener params.")
            return {}
        
        strategy = strategies[0]
        dummy_trial = MinParamTrial()
        
        # Ejecuta suggest_params con el mock para capturar mínimos
        params = strategy.suggest_params(dummy_trial)
        logger.info(f"🔹 Parámetros Mínimos Capturados del Código: {params}")
        return params
        
    except Exception as e:
        logger.error(f"Error obteniendo params mínimos: {e}")
        return {}

def debug_finalize_patch(self, q: pl.LazyFrame, **kwargs) -> pl.DataFrame:
    """Monkey-patch para interceptar TODAS las columnas intermedias."""
    return q.collect()

def generate_excel_report(pdf: pd.DataFrame, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "DEBUG_REPORT"
    ws.sheet_view.showGridLines = False 
    ws.freeze_panes = "D2" # Congelar Fecha, Precio y Entrada
    
    headers = list(pdf.columns)
    
    # 1. HEADERS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_FULL

    # 2. DATA
    for r_idx, row in pdf.iterrows():
        excel_row = r_idx + 2
        for c_idx, col_name in enumerate(headers):
            val = row[col_name]
            cell = ws.cell(row=excel_row, column=c_idx+1, value=val)
            cell.font = FONT_BODY
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER_FULL
            
            # Formatos de Fecha/Precio
            if col_name == "TIME":
                cell.number_format = "YYYY-MM-DD HH:MM"
                cell.font = Font(name="Arial", size=9, color="555555")
            elif "PRICE" in col_name:
                cell.number_format = "#,##0.00"
            
            # --- ESTILOS VISUALES ---
            
            # TENDENCIA
            if col_name == "TENDENCIA":
                if val == "ALCISTA":
                    cell.font = Font(name="Arial", size=10, color="006100", bold=True)
                    cell.fill = PatternFill("solid", fgColor="E6F4EA") # Verde
                else:
                    cell.font = Font(name="Arial", size=10, color="9C0006", bold=True)
                    cell.fill = PatternFill("solid", fgColor="FCE8E6") # Rojo

            # ESTADOS (APROBADO/FALLO)
            elif "ESTADO" in col_name or "COND" in col_name or "CHECK" in col_name:
                if val in ["✓", "APROBADO", True]:
                    cell.value = "✓ OK"
                    cell.font = Font(name="Arial", size=10, color="006100", bold=True)
                    cell.fill = PatternFill("solid", fgColor="E6F4EA")
                elif val in ["✕", "FALLO", False]:
                    cell.value = "✕ NO"
                    cell.font = Font(name="Arial", size=10, color="9C0006", bold=True)
                    cell.fill = PatternFill("solid", fgColor="FCE8E6")
            
            # ENTRADA FINAL
            elif "ENTRADA" in col_name or "SIG" in col_name:
                if "🚀" in str(val) or "LONG" in str(val) or "SHORT" in str(val) or val is True:
                     # Si es booleano True o texto de señal
                     if val is True: cell.value = "🚀"
                     cell.font = Font(name="Arial", size=12, color="FFFFFF", bold=True)
                     cell.fill = PatternFill("solid", fgColor="2E7D32") # Verde Intenso
                else:
                    cell.font = Font(name="Arial", size=10, color="DDDDDD") # Gris muy claro
                    if val is False: cell.value = "-"
                    
            # PARAMETROS (P_)
            elif col_name.startswith("P_"):
                cell.font = Font(name="Arial", size=9, color="666666", italic=True)
                cell.fill = PatternFill("solid", fgColor="F5F5F5") # Gris muy suave
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # GENERIC BOOLEAN CONDITIONS
            elif isinstance(val, bool) or str(val) in ["✓", "✕"]:
                cell.number_format = "@"
                if val in [True, "✓"]:
                    cell.value = "✓"
                    cell.font = FONT_SUCCESS
                    cell.fill = PatternFill("solid", fgColor=COLORS["success_bg"])
                else:
                    cell.value = "✕"
                    cell.font = FONT_DANGER
                    cell.fill = PatternFill("solid", fgColor=COLORS["danger_bg"])

            # NUMEROS
            elif isinstance(val, (int, float)):
                cell.number_format = "0.0000"

    # 3. AUTO-WIDTH INTELIGENTE
    for col in range(1, len(headers) + 1):
        column = get_column_letter(col)
        # Anchos fijos estimados para limpieza visual
        width = 15
        header = str(ws.cell(row=1, column=col).value)
        
        if "TIME" in header: width = 18
        elif "PRICE" in header: width = 12
        elif "VELAS" in header: width = 10
        elif "DIST" in header: width = 12
        elif "ESTADO" in header: width = 15
        elif "ENTRADA" in header: width = 15
        elif "COND" in header: width = 10
        elif "P_" in header: width = 16
        
        ws.column_dimensions[column].width = width

    # 4. DATA BARS (VISUALIZACIÓN DE MAGNITUDES)
    
    # Identificar columnas por letra
    col_map = {h: get_column_letter(i+1) for i, h in enumerate(headers)}
    max_row = len(pdf) + 1
    
    if "C_VELAS_ACTUAL" in col_map:
        c = col_map["C_VELAS_ACTUAL"]
        # Escala de Color: Blanco (0) -> Rojo (Max Limit)
        ws.conditional_formatting.add(f"{c}2:{c}{max_row}", ColorScaleRule(
            start_type='num', start_value=0, start_color='FFFFFF',
            end_type='percentile', end_value=90, end_color='FFCDD2'
        ))

    if "D_DIST_ACTUAL" in col_map:
        c = col_map["D_DIST_ACTUAL"]
        # Data Bar Azul para visualizar magnitud de distancia
        ws.conditional_formatting.add(f"{c}2:{c}{max_row}", DataBarRule(
            start_type='min', end_type='max', color="64B5F6", showValue=True
        ))

    wb.save(filename)
    logger.info(f"✅ Reporte guardado: {filename}")

def main():
    logger.info(f"🔹 Iniciando Depurador de Estrategia ID: {STRATEGY_ID}")
    
    try:
        # 1. Cargar Estrategia
        strategies = instantiate_strategies(only_id=STRATEGY_ID)
        if not strategies:
            logger.error(f"❌ No se encontró estrategia con ID {STRATEGY_ID}")
            return
        strategy = strategies[0]
        logger.info(f"   Estrategia Cargada: {strategy.name}")
        
        # 2. Monkey Patch para extraer columnas internas
        strategy.finalize_signals = types.MethodType(debug_finalize_patch, strategy)
        
        # 3. Cargar Datos
        tf = CONFIG.get("TIMEFRAME", "5m")
        logger.info(f"🔹 Cargando datos {ACTIVO_PRIMARIO} ({tf})...")
        
        archivo = resolve_archivo_data_tf(ACTIVO_PRIMARIO, tf)
        if not os.path.exists(archivo):
             archivo = resolve_archivo_data(ACTIVO_PRIMARIO)
        
        df = load_data(archivo)
        # Resamplear si es necesario
        tf_suffix = normalize_timeframe_to_suffix(tf)
        df = resample_to_base_timeframe(df, tf_suffix)

        logger.info(f"   Datos cargados: {len(df)} velas")

        # 4. Ejecutar Estrategia
        logger.info("🔹 Ejecutando lógica de estrategia...")
        params = get_debug_params(STRATEGY_ID)
        
        # Esto retornará TODAS las columnas gracias al patch
        result_df = strategy.generate_signals(df, params)
    
    except Exception as e:
        logger.error(f"❌ Error Ejecución: {e}")
        import traceback
        traceback.print_exc()
        return

    # 5. Procesar Resultados
    # Convertir a Pandas para manipulación de reportes y fechas
    pdf = result_df.to_pandas()
    
    # 5.1 Normalizar Timestamp (CRÍTICO antes de filtrar por fecha)
    if "timestamp" in pdf.columns:
        if pdf["timestamp"].dtype == 'int64':
             pdf["timestamp"] = pd.to_datetime(pdf["timestamp"], unit='ms', errors='ignore')
        else:
             pdf["timestamp"] = pd.to_datetime(pdf["timestamp"])
        
        if hasattr(pdf["timestamp"].dt, "tz") and pdf["timestamp"].dt.tz is not None:
            pdf["timestamp"] = pdf["timestamp"].dt.tz_localize(None)
    
    # 5.2 Filtrado por Fecha y Límite
    if DEBUG_START_DATE:
        start_dt = pd.to_datetime(DEBUG_START_DATE)
        pdf = pdf[pdf["timestamp"] >= start_dt]
        # Tomamos las PRIMERAS velas desde esa fecha (inicio del periodo)
        pdf = pdf.head(LIMIT_CANDLES)
        logger.info(f"   Filtrado desde {DEBUG_START_DATE}: {len(pdf)} velas capturadas.")
    else:
        # Si no hay fecha, comportamiento clásico: ultimas N velas
        pdf = pdf.tail(LIMIT_CANDLES)
        logger.info(f"   Tomando últimas {LIMIT_CANDLES} velas.")
    
    # 6. Construir DataFrame de Reporte
    
    # --- 6.1 INYECTAR PARÁMETROS (MODULAR PARA CUALQUIER ID) ---
    for k, v in params.items():
        if not k.startswith("__"): # Ignorar internos
            col_name = f"P_{k.upper()}"
            pdf[col_name] = v

    # --- 6.2 CONSTRUCCION DEL DATAFRAME FINAL ---
    final_pdf = pd.DataFrame()
    final_pdf["TIME"] = pdf["timestamp"]
    final_pdf["PRICE"] = pdf["close"]

    # --- 6.3 COLUMNA UNIFICADA "ENTRADA" ---
    def _get_entry(row):
        if "signal_long" in row and row["signal_long"]: return "🚀 LONG"
        if "signal_short" in row and row["signal_short"]: return "🚀 SHORT"
        return "-"
    
    if "signal_long" in pdf.columns:
         final_pdf["ENTRADA"] = pdf.apply(_get_entry, axis=1)

    # --- LÓGICA DE COLUMNAS SEGUN ID O GENÉRICA ---
    # Detectar todas las columnas disponibles excepto las ya usadas
    skip_cols = ["timestamp", "close", "open", "high", "low", "volume", "cycle_id", "signal_long", "signal_short", "signal_exit"]
    skip_cols += [c for c in pdf.columns if c.startswith("P_")] 

    # Lógica específica ID 1 (Mejorada para convivir con genérico)
    if STRATEGY_ID == 1:
        # TENDENCIA
        if "is_bullish" in pdf.columns:
            final_pdf["TENDENCIA"] = pdf["is_bullish"].apply(lambda x: "ALCISTA" if x else "BAJISTA")
        
        # LOGICA TIEMPO
        if "bars_in_cycle" in pdf.columns:
            final_pdf["VELAS_ACTUAL"] = pdf["bars_in_cycle"]
            final_pdf["COND_TIEMPO"] = pdf.apply(
                lambda r: "✓" if r["bars_in_cycle"] <= params.get("lookbar", 999) else "✕", axis=1
            )

        # LOGICA DISTANCIA
        if "curr_dist" in pdf.columns and "target_dist" in pdf.columns:
            final_pdf["DIST_ACTUAL"] = pdf["curr_dist"]
            final_pdf["DIST_OBJETIVO"] = pdf["target_dist"]
            final_pdf["COND_DIST"] = pdf.apply(
                lambda r: "✓" if r["curr_dist"] >= r["target_dist"] else "✕", axis=1
            )
            
    # COPY REST OF COLUMNS GENERICALLY
    handled_cols = ["timestamp", "close", "is_bullish", "bars_in_cycle", "curr_dist", "target_dist"]
    
    for c in pdf.columns:
        if c not in skip_cols and c not in handled_cols:
            col_name = c.upper()
            if "SIGNAL" in col_name: continue # Skip other signals if managed
            final_pdf[col_name] = pdf[c]

    # --- 6.4 INYECTAR PARÁMETROS AL FINAL ---
    for c in pdf.columns:
        if c.startswith("P_"):
            final_pdf[c] = pdf[c]

    # 7. Generar Excel
    output_file = f"DEBUG_ESTRATEGIA_ID{STRATEGY_ID}.xlsx"
    generate_excel_report(final_pdf.reset_index(drop=True), output_file)

if __name__ == "__main__":
    main()
