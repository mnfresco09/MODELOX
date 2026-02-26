"""
TRADE LOGGER - REGISTRO DE TRADES EN EXCEL
============================================
Guarda cada trade cerrado en un archivo Excel con formato profesional.

Archivos:
  - TRADES_REAL.xlsx  → trading con dinero real  (USDT)
  - TRADES_DEMO.xlsx  → trading demo             (VST)

Flujo:
  1. Al arrancar la sesión → initialize_session()  crea / sobreescribe el archivo
  2. Al cerrar un trade   → log_trade()             agrega una fila

Estructura del Excel:
  Fila 1 : etiquetas de sesión  (SALDO INICIAL | APALANCAMIENTO | SALDO POR TRADE)
  Fila 2 : valores  de sesión
  Fila 3 : separador vacío
  Fila 4 : cabeceras de columnas de trades
  Fila 5+: trades cerrados
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

EXCEL_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "resultados")
EXCEL_REAL = "TRADES_REAL.xlsx"
EXCEL_DEMO = "TRADES_DEMO.xlsx"
SHEET_NAME = "Trades"

# Filas del layout
ROW_META_LABEL  = 1
ROW_META_VALUE  = 2
ROW_SEPARATOR   = 3
ROW_HEADER      = 4
ROW_DATA_START  = 5

# Cabecera de sesión
META_FIELDS = ["SALDO INICIAL", "APALANCAMIENTO", "SALDO POR TRADE"]

# Columnas de trades (orden definitivo)
TRADE_COLUMNS = [
    "ESTRATEGIA",
    "ACTIVO",
    "FECHA ENTRADA",
    "FECHA SALIDA",
    "MOTIVO",
    "DIRECCIÓN",
    "PNL BRUTO",
    "COMISIONES",
    "PNL NETO",
    "BALANCE TRAS CIERRE",
]

# Paleta de colores
_C = {
    "meta_label_bg":  "2C3E50",
    "meta_label_fg":  "FFFFFF",
    "meta_value_bg":  "1A2530",
    "meta_value_fg":  "ECF0F1",
    "col_header_bg":  "1A5276",
    "col_header_fg":  "FFFFFF",
    "cell_fg":        "212F3D",
    "border":         "BDC3C7",
    "profit_bg":      "D5F5E3",
    "profit_fg":      "006100",
    "loss_bg":        "FADBD8",
    "loss_fg":        "922B21",
}

_FONT  = "Arial"
_SZ_M  = 12   # meta
_SZ_H  = 11   # header
_SZ_B  = 11   # body

_BORDER = Border(
    left  =Side(style="thin", color=_C["border"]),
    right =Side(style="thin", color=_C["border"]),
    top   =Side(style="thin", color=_C["border"]),
    bottom=Side(style="thin", color=_C["border"]),
)

_COL_WIDTHS = {
    "ESTRATEGIA":          16,
    "ACTIVO":              10,
    "FECHA ENTRADA":       22,
    "FECHA SALIDA":        22,
    "MOTIVO":              10,
    "DIRECCIÓN":           10,
    "PNL BRUTO":           14,
    "COMISIONES":          14,
    "PNL NETO":            14,
    "BALANCE TRAS CIERRE": 20,
}


# =============================================================================
# HELPERS INTERNOS
# =============================================================================

def _get_path(quote_currency: str) -> str:
    os.makedirs(EXCEL_DIR, exist_ok=True)
    fname = EXCEL_REAL if quote_currency == "USDT" else EXCEL_DEMO
    return os.path.join(EXCEL_DIR, fname)


def _read_initial_balance(ws) -> float:
    """Lee el saldo inicial almacenado en la celda de metadatos."""
    raw = str(ws.cell(row=ROW_META_VALUE, column=1).value or "0")
    raw = raw.replace("$", "").replace(",", "").strip()
    try:
        return float(raw)
    except (ValueError, TypeError):
        return 0.0


def _read_last_balance(ws) -> float:
    """
    Devuelve el último BALANCE TRAS CIERRE del sheet.
    Si no hay trades todavía, devuelve el saldo inicial.
    """
    bal_col = TRADE_COLUMNS.index("BALANCE TRAS CIERRE") + 1
    for row in range(ws.max_row, ROW_DATA_START - 1, -1):
        val = ws.cell(row=row, column=bal_col).value
        if val is not None:
            try:
                return float(val)
            except (ValueError, TypeError):
                pass
    return _read_initial_balance(ws)


def _color_cell(ws, row: int, col: int, positive: bool) -> None:
    """Aplica color verde/rojo a una celda según si es positivo."""
    cell = ws.cell(row=row, column=col)
    if cell.value is None:
        return
    if positive:
        cell.fill = PatternFill("solid", fgColor=_C["profit_bg"])
        cell.font = Font(name=_FONT, size=_SZ_B, color=_C["profit_fg"], bold=True)
    else:
        cell.fill = PatternFill("solid", fgColor=_C["loss_bg"])
        cell.font = Font(name=_FONT, size=_SZ_B, color=_C["loss_fg"], bold=True)


# =============================================================================
# API PÚBLICA
# =============================================================================

def initialize_session(
    initial_balance: float,
    leverage: int,
    amount_per_trade: float,
    quote_currency: str,
) -> None:
    """
    Crea (o sobreescribe) el archivo Excel para la sesión actual.

    Args:
        initial_balance:  Saldo disponible al inicio de la sesión
        leverage:         Apalancamiento configurado
        amount_per_trade: Margen usado por trade
        quote_currency:   "USDT" (real) o "VST" (demo)
    """
    try:
        path = _get_path(quote_currency)

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.sheet_view.showGridLines = False

        meta_values = [
            f"${initial_balance:,.2f}",
            f"{leverage}x",
            f"${amount_per_trade:,.2f}",
        ]

        # ── Fila 1: etiquetas ────────────────────────────────────────────
        for c, label in enumerate(META_FIELDS, 1):
            cell = ws.cell(row=ROW_META_LABEL, column=c)
            cell.value = label
            cell.font      = Font(name=_FONT, size=_SZ_M, bold=True, color=_C["meta_label_fg"])
            cell.fill      = PatternFill("solid", fgColor=_C["meta_label_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _BORDER

        # ── Fila 2: valores ──────────────────────────────────────────────
        for c, val in enumerate(meta_values, 1):
            cell = ws.cell(row=ROW_META_VALUE, column=c)
            cell.value = val
            cell.font      = Font(name=_FONT, size=_SZ_M, bold=True, color=_C["meta_value_fg"])
            cell.fill      = PatternFill("solid", fgColor=_C["meta_value_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _BORDER

        # ── Fila 3: separador (vacía) ────────────────────────────────────

        # ── Fila 4: cabeceras de trades ───────────────────────────────────
        for c, col in enumerate(TRADE_COLUMNS, 1):
            cell = ws.cell(row=ROW_HEADER, column=c)
            cell.value = col
            cell.font      = Font(name=_FONT, size=_SZ_H, bold=True, color=_C["col_header_fg"])
            cell.fill      = PatternFill("solid", fgColor=_C["col_header_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = _BORDER

        # ── Alturas ───────────────────────────────────────────────────────
        ws.row_dimensions[ROW_META_LABEL].height = 24
        ws.row_dimensions[ROW_META_VALUE].height = 24
        ws.row_dimensions[ROW_SEPARATOR].height  = 6
        ws.row_dimensions[ROW_HEADER].height     = 28

        # ── Anchos ────────────────────────────────────────────────────────
        for c, col in enumerate(TRADE_COLUMNS, 1):
            w = max(_COL_WIDTHS.get(col, 14), 20 if c <= len(META_FIELDS) else 0)
            ws.column_dimensions[get_column_letter(c)].width = w

        wb.save(path)

    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"Error inicializando Excel de sesión: {e}")


def log_trade(
    trade: Any,
    config: Any,
    strategy_name: str = "",
) -> None:
    """
    Registra un trade cerrado en el Excel de la sesión actual.

    Args:
        trade:         TradeRecord con los datos del trade
        config:        TradingConfig de la sesión
        strategy_name: Nombre completo de la estrategia
    """
    try:
        quote_currency = getattr(config, "quote_currency", "VST")
        path = _get_path(quote_currency)

        # Crear el archivo si no existe (fallback)
        if not os.path.exists(path):
            initialize_session(
                initial_balance=0.0,
                leverage=getattr(config, "leverage", 1),
                amount_per_trade=getattr(config, "amount_per_trade", 0.0),
                quote_currency=quote_currency,
            )

        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

        # ── Datos del trade ───────────────────────────────────────────────
        pnl_bruto    = getattr(trade, "pnl", 0) or 0
        trading_fees = getattr(trade, "trading_fees", 0) or 0
        funding_fees = getattr(trade, "funding_fees", 0) or 0
        comisiones   = trading_fees + funding_fees
        pnl_neto     = getattr(trade, "total_pnl", None) or (pnl_bruto + comisiones)

        # Fechas con HH:MM:SS
        entry_time = getattr(trade, "timestamp", None)
        exit_time  = getattr(trade, "exit_time", None) or datetime.now()
        _fmt = "%d/%m/%Y %H:%M:%S"
        fecha_entrada = entry_time.strftime(_fmt) if isinstance(entry_time, datetime) else "--"
        fecha_salida  = exit_time.strftime(_fmt)  if isinstance(exit_time,  datetime) else "--"

        # Motivo de cierre
        reason = getattr(trade, "reason", None) or "MANUAL"
        motivo = {
            "SL": "SL",
            "TP": "TP",
            "TRAILING": "TRAILING",
            "EXCHANGE_CLOSE": "MANUAL",
        }.get(reason, reason)

        # Activo (nombre corto)
        from real.config import DISPLAY_NAMES
        activo = DISPLAY_NAMES.get(trade.symbol, trade.symbol.replace("-USDT", ""))

        # Estrategia en formato IDX
        strat_id = getattr(config, "strategy_id", None)
        estrategia = f"ID{strat_id}" if strat_id else strategy_name

        # Balance acumulado
        prev_balance      = _read_last_balance(ws)
        balance_tras_cierre = prev_balance + pnl_neto

        row_data = {
            "ESTRATEGIA":          estrategia,
            "ACTIVO":              activo,
            "FECHA ENTRADA":       fecha_entrada,
            "FECHA SALIDA":        fecha_salida,
            "MOTIVO":              motivo,
            "DIRECCIÓN":           trade.side,
            "PNL BRUTO":           round(pnl_bruto, 4),
            "COMISIONES":          round(comisiones, 4),
            "PNL NETO":            round(pnl_neto, 4),
            "BALANCE TRAS CIERRE": round(balance_tras_cierre, 4),
        }

        # ── Escribir fila ─────────────────────────────────────────────────
        next_row = max(ws.max_row + 1, ROW_DATA_START)

        for c, col in enumerate(TRADE_COLUMNS, 1):
            cell = ws.cell(row=next_row, column=c)
            cell.value     = row_data[col]
            cell.font      = Font(name=_FONT, size=_SZ_B, color=_C["cell_fg"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _BORDER

            # Formato numérico
            if isinstance(cell.value, (int, float)):
                col_up = col.upper()
                if any(k in col_up for k in ("PNL", "COMISIONES", "BALANCE")):
                    cell.number_format = "#,##0.0000"

        # Colores PNL NETO
        pnl_col = TRADE_COLUMNS.index("PNL NETO") + 1
        _color_cell(ws, next_row, pnl_col, float(pnl_neto) >= 0)

        # Color BALANCE TRAS CIERRE (verde si >= saldo inicial)
        bal_col = TRADE_COLUMNS.index("BALANCE TRAS CIERRE") + 1
        _color_cell(ws, next_row, bal_col, float(balance_tras_cierre) >= _read_initial_balance(ws))

        ws.row_dimensions[next_row].height = 22
        wb.save(path)

    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"Error guardando trade en Excel: {e}")
